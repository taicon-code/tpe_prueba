# tpe_app/views/ayudante_views.py
"""
Vistas para el rol AYUDANTE - Registro de datos históricos y búsqueda de antecedentes
"""

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.db import transaction
from django.db.models import Q, Count
from django.utils import timezone
from django.http import JsonResponse
from django.urls import reverse
from datetime import date
from ..decorators import rol_requerido
from ..models import (
    SIM, PM, PM_SIM, AUTOTPE, AUTOTSP, VOCAL_TPE, Resolucion, RecursoTSP, Notificacion, Memorandum,
    DocumentoAdjunto,
)
from ..forms import (
    RESForm, NotificacionForm, RAPForm, RAEEForm, AUTOTPEHistoricoForm, MemorandumForm,
    PMSIMFormSet, WizardSIMForm, WizardRESForm, WizardRRForm, WizardAUTOTPEForm, WizardRAPForm, WizardRAEEForm, WizardAUTOTSPForm,
    BuscarSIMHistoricoForm, EditarSIMHistoricoForm
)


@rol_requerido('AYUDANTE', 'ADMIN3_NOTIFICADOR')
def ayudante_dashboard(request):
    """Dashboard principal del AYUDANTE"""
    from ..models import DocumentoAdjunto

    # Últimos SIM registrados
    ultimos_sim = SIM.objects.all().order_by('-fecha_registro')[:10]

    # Últimas Resoluciones PRIMERA registradas
    ultimas_res = Resolucion.objects.filter(instancia='PRIMERA').order_by('-fecha')[:10]

    # RES sin PDF (panel principal) — solo PRIMERA
    res_con_pdf = set(
        DocumentoAdjunto.objects.filter(resolucion__isnull=False).values_list('resolucion_id', flat=True)
    )
    res_sin_pdf = (
        Resolucion.objects.filter(instancia='PRIMERA')
        .exclude(id__in=res_con_pdf)
        .select_related('sim', 'abogado').order_by('-fecha')[:10]
    )
    total_res_sin_pdf = (
        Resolucion.objects.filter(instancia='PRIMERA')
        .exclude(id__in=res_con_pdf).count()
    )

    # Contadores generales
    total_sim = SIM.objects.count()
    total_res = Resolucion.objects.filter(instancia='PRIMERA').count()
    total_autotpe = AUTOTPE.objects.count()

    # Autos sin PDF
    autos_con_pdf = set(
        DocumentoAdjunto.objects.filter(autotpe__isnull=False).values_list('autotpe_id', flat=True)
    )
    total_autotpe_sin_pdf = AUTOTPE.objects.exclude(id__in=autos_con_pdf).count()

    # RES por año (solo instancia PRIMERA)
    res_por_anio = list(
        Resolucion.objects.filter(instancia='PRIMERA')
        .values('fecha__year')
        .annotate(total=Count('id'))
        .order_by('-fecha__year')
    )

    # Autos TPE por año
    autos_por_anio = list(
        AUTOTPE.objects.values('fecha__year')
        .annotate(total=Count('id'))
        .order_by('-fecha__year')
    )

    context = {
        'ultimos_sim': ultimos_sim,
        'ultimas_res': ultimas_res,
        'res_sin_pdf': res_sin_pdf,
        'total_res_sin_pdf': total_res_sin_pdf,
        'total_sim': total_sim,
        'total_res': total_res,
        'total_autotpe': total_autotpe,
        'total_autotpe_sin_pdf': total_autotpe_sin_pdf,
        'res_por_anio': res_por_anio,
        'autos_por_anio': autos_por_anio,
    }

    return render(request, 'tpe_app/ayudante/dashboard.html', context)


@rol_requerido('AYUDANTE', 'ADMIN3_NOTIFICADOR')
def ayudante_lista_res(request):
    """Lista de RES filtrada por gestión (año) con: grado, nombres, tipo, notificado"""

    # Parámetro GET: gestion (año), default = año actual
    gestion = request.GET.get('gestion', str(date.today().year))

    try:
        gestion = int(gestion)
    except (ValueError, TypeError):
        gestion = date.today().year

    # Query: Resoluciones PRIMERA del año especificado, ordenadas por grado/apellido
    resoluciones = (
        Resolucion.objects
        .filter(instancia='PRIMERA', fecha__year=gestion)
        .select_related('pm', 'sim')
        .order_by('pm__grado', 'pm__paterno', 'pm__nombre')
    )

    # Opciones de gestión disponibles (años únicos)
    gestiones_disponibles = (
        Resolucion.objects.filter(instancia='PRIMERA')
        .values_list('fecha__year', flat=True)
        .distinct()
        .order_by('-fecha__year')
    )

    context = {
        'resoluciones': resoluciones,
        'gestion_actual': gestion,
        'gestiones_disponibles': gestiones_disponibles,
    }

    return render(request, 'tpe_app/ayudante/lista_res.html', context)


@rol_requerido('AYUDANTE', 'ADMIN3_NOTIFICADOR')
def ayudante_registrar_res(request):
    """Registrar una Resolución histórica (sin dictamen previo) + notificación opcional"""
    from ..forms import ResolucionConNotificacionForm

    if request.method == 'POST':
        form = ResolucionConNotificacionForm(request.POST)
        if form.is_valid():
            try:
                with transaction.atomic():
                    # Crear la Resolución
                    res = Resolucion(
                        sim=form.cleaned_data['sim'],
                        pm=form.cleaned_data['pm'],
                        numero=form.cleaned_data['numero'],
                        fecha=form.cleaned_data['fecha'],
                        tipo=form.cleaned_data['tipo'],
                        texto=form.cleaned_data['texto'],
                        instancia='PRIMERA'
                    )
                    res.save()

                    # Si se proporcionan datos de notificación, crear el registro
                    if form.cleaned_data.get('notif_tipo'):
                        notif = Notificacion(
                            resolucion=res,
                            tipo=form.cleaned_data['notif_tipo'],
                            notificado_a=form.cleaned_data.get('notif_notificado_a', ''),
                            fecha=form.cleaned_data.get('notif_fecha'),
                            hora=form.cleaned_data.get('notif_hora')
                        )
                        notif.save()
                        messages.success(request, f'Resolución {res.numero} registrada con notificación')
                    else:
                        messages.success(request, f'Resolución {res.numero} registrada exitosamente')

                    # Actualizar la fase del SIM si es necesario
                    sim = res.sim
                    if sim.fase not in ['1RA_RESOLUCION', '2DA_RESOLUCION', 'ELEVADO_TSP', 'CONCLUIDO']:
                        sim.fase = '1RA_RESOLUCION'
                        sim.estado = 'PROCESO_EN_EL_TPE'
                        sim.save()

                    return redirect('ayudante_lista_res')
            except Exception as e:
                messages.error(request, f'Error al registrar la resolución: {str(e)}')
    else:
        form = ResolucionConNotificacionForm()

    context = {
        'form': form,
        'titulo': 'Registrar Resolución Histórica',
    }

    return render(request, 'tpe_app/ayudante/registrar_res.html', context)


@rol_requerido('AYUDANTE', 'ADMIN1_AGENDADOR', 'ADMIN3_NOTIFICADOR')
def ayudante_registrar_notificacion(request, res_id):
    """Registrar la notificación de una RES existente"""
    from ..models import Notificacion

    res = get_object_or_404(Resolucion, id=res_id)
    notif_existente = getattr(res, 'notificacion', None)

    next_url = request.GET.get('next', '').strip()

    if request.method == 'POST':
        form = NotificacionForm(request.POST, instance=notif_existente)
        if form.is_valid():
            try:
                with transaction.atomic():
                    notif = form.save(commit=False)
                    notif.resolucion = res
                    notif.save()
                    messages.success(request, f'Notificación de RES {res.numero} registrada exitosamente')
                    if next_url:
                        return redirect(next_url)
                    rol = getattr(request.user.perfilusuario, 'rol', 'AYUDANTE')
                    if rol in ['ADMIN3_NOTIFICADOR', 'ADMIN1_AGENDADOR']:
                        return redirect('admin3_dashboard')
                    else:
                        return redirect('ayudante_lista_res')
            except Exception as e:
                messages.error(request, f'Error al registrar notificación: {str(e)}')
    else:
        form = NotificacionForm(instance=notif_existente)

    context = {
        'form': form,
        'res': res,
        'next_url': next_url,
        'titulo': f'Registrar Notificación - RES {res.numero}',
    }

    return render(request, 'tpe_app/ayudante/registrar_notificacion.html', context)


@rol_requerido('AYUDANTE', 'ADMIN3_NOTIFICADOR')
def ayudante_registrar_rap(request):
    """Registrar un Recurso de Apelación al TSP histórico + notificación opcional"""
    from ..forms import RAPConNotificacionForm

    if request.method == 'POST':
        form = RAPConNotificacionForm(request.POST)
        if form.is_valid():
            try:
                with transaction.atomic():
                    # Crear el RAP
                    rap = RecursoTSP(
                        sim=form.cleaned_data['sim'],
                        pm=form.cleaned_data['pm'],
                        resolucion=form.cleaned_data.get('resolucion'),
                        fecha_presentacion=form.cleaned_data['fecha_presentacion'],
                        numero_oficio=form.cleaned_data.get('numero_oficio'),
                        fecha_oficio=form.cleaned_data.get('fecha_oficio'),
                        numero=form.cleaned_data['numero'],
                        fecha=form.cleaned_data['fecha'],
                        texto=form.cleaned_data['texto'],
                        tipo=form.cleaned_data['tipo'],
                        instancia='APELACION'
                    )
                    rap.save()

                    # Si se proporcionan datos de notificación
                    if form.cleaned_data.get('notif_tipo'):
                        notif = Notificacion(
                            recurso_tsp=rap,
                            tipo=form.cleaned_data['notif_tipo'],
                            notificado_a=form.cleaned_data.get('notif_notificado_a', ''),
                            fecha=form.cleaned_data.get('notif_fecha'),
                            hora=form.cleaned_data.get('notif_hora')
                        )
                        notif.save()
                        messages.success(request, f'RAP {rap.numero} registrado con notificación')
                    else:
                        messages.success(request, f'Recurso de Apelación {rap.numero} registrado exitosamente')

                    # Actualizar fase del SIM
                    sim = rap.sim
                    if sim.fase not in ['ELEVADO_TSP', 'CONCLUIDO']:
                        sim.fase = 'ELEVADO_TSP'
                        sim.estado = 'PROCESO_EN_EL_TSP'
                        sim.save()

                    return redirect('ayudante_dashboard')
            except Exception as e:
                messages.error(request, f'Error al registrar RAP: {str(e)}')
    else:
        form = RAPConNotificacionForm()

    context = {
        'form': form,
        'titulo': 'Registrar Recurso de Apelación (RAP) Histórico',
    }

    return render(request, 'tpe_app/ayudante/registrar_rap.html', context)


@rol_requerido('AYUDANTE', 'ADMIN3_NOTIFICADOR')
def ayudante_registrar_raee(request):
    """Registrar un RAEE (Aclaración, Explicación y Enmienda) histórico"""

    if request.method == 'POST':
        form = RAEEForm(request.POST)
        if form.is_valid():
            try:
                with transaction.atomic():
                    raee = form.save(commit=False)
                    raee.save()

                    messages.success(
                        request,
                        f'RAEE {raee.numero} registrado exitosamente'
                    )
                    return redirect('ayudante_dashboard')
            except Exception as e:
                messages.error(request, f'Error al registrar RAEE: {str(e)}')
    else:
        form = RAEEForm()

    context = {
        'form': form,
        'titulo': 'Registrar RAEE (Aclaración, Explicación y Enmienda) Histórico',
    }

    return render(request, 'tpe_app/ayudante/registrar_raee.html', context)


@rol_requerido('AYUDANTE', 'ADMIN3_NOTIFICADOR')
def ayudante_registrar_autotpe(request):
    """Registrar un Auto del TPE histórico (con notificación y memorándum opcional)"""
    from ..forms import AUTOTPEHistoricoConNotificacionForm

    if request.method == 'POST':
        form = AUTOTPEHistoricoConNotificacionForm(request.POST)
        if form.is_valid():
            try:
                with transaction.atomic():
                    # Crear el Auto TPE
                    autotpe = AUTOTPE(
                        sim=form.cleaned_data['sim'],
                        pm=form.cleaned_data['pm'],
                        numero=form.cleaned_data['numero'],
                        fecha=form.cleaned_data['fecha'],
                        tipo=form.cleaned_data['tipo'],
                        texto=form.cleaned_data['texto'],
                        resolucion=form.cleaned_data.get('resolucion')
                    )
                    autotpe.save()

                    # Si se proporcionan datos de notificación
                    if form.cleaned_data.get('notif_tipo'):
                        notif = Notificacion(
                            autotpe=autotpe,
                            tipo=form.cleaned_data['notif_tipo'],
                            notificado_a=form.cleaned_data.get('notif_notificado_a', ''),
                            fecha=form.cleaned_data.get('notif_fecha'),
                            hora=form.cleaned_data.get('notif_hora')
                        )
                        notif.save()

                    # Si es ejecutoria y se proporciona memorándum
                    has_memo = bool(form.cleaned_data.get('memo_numero')) and autotpe.tipo == 'AUTO_EJECUTORIA'
                    if has_memo:
                        memo = Memorandum(
                            autotpe=autotpe,
                            numero=form.cleaned_data['memo_numero'],
                            fecha=form.cleaned_data.get('memo_fecha')
                        )
                        memo.save()
                        messages.success(request, f'Auto TPE {autotpe.numero} registrado con memorándum')
                    else:
                        messages.success(request, f'Auto TPE {autotpe.numero} registrado exitosamente')

                    # Advertencia si Auto de Ejecutoria sin resolución vinculada
                    if autotpe.tipo == 'AUTO_EJECUTORIA' and not autotpe.resolucion:
                        messages.warning(request,
                            '⚠️ Auto de Ejecutoria registrado sin resolución vinculada. '
                            'Esto puede afectar los reportes de ejecutoria pendientes.')

                    # Actualizar fase/estado del SIM según el tipo de auto registrado.
                    # Guardia multi-persona: si el SIM ya está en un estado de proceso
                    # externo activo (TSP, cumplimiento), no avanzar por el auto de
                    # otro militar — el SIM solo avanza cuando ese proceso externo concluya.
                    ESTADOS_ACTIVOS_EXTERNOS = {'PROCESO_EN_EL_TSP', 'CUMPLIMIENTO_EN_TPE'}
                    if autotpe.tipo == 'AUTO_EJECUTORIA':
                        sim = form.cleaned_data['sim']
                        if sim.estado not in ESTADOS_ACTIVOS_EXTERNOS:
                            if has_memo:
                                sim.fase = 'MEMORANDUM_RETORNADO'
                            elif form.cleaned_data.get('notif_tipo'):
                                sim.fase = 'EJECUTORIA_NOTIFICADA'
                            else:
                                sim.fase = 'EN_EJECUTORIA'
                            sim.save()

                    return redirect('ayudante_dashboard')
            except Exception as e:
                messages.error(request, f'Error al registrar Auto TPE: {str(e)}')
    else:
        form = AUTOTPEHistoricoConNotificacionForm()

    context = {
        'form': form,
        'titulo': 'Registrar Auto del TPE Histórico (con Notificación y Memorándum)',
    }

    return render(request, 'tpe_app/ayudante/registrar_autotpe.html', context)


@rol_requerido('AYUDANTE', 'ADMIN1_AGENDADOR', 'ADMIN3_NOTIFICADOR')
def ayudante_registrar_notificacion_rr(request, rr_id):
    """Registrar la notificación de un RR existente"""

    rr = get_object_or_404(Resolucion, id=rr_id, instancia='RECONSIDERACION')

    from ..models import Notificacion
    notif_existente = getattr(rr, 'notificacion', None)
    next_url = request.GET.get('next', '').strip()

    if request.method == 'POST':
        form = NotificacionForm(request.POST, instance=notif_existente)
        if form.is_valid():
            try:
                with transaction.atomic():
                    notif = form.save(commit=False)
                    notif.resolucion = rr
                    notif.save()
                    messages.success(request, f'Notificación de RR {rr.numero} registrada exitosamente')
                    if next_url:
                        return redirect(next_url)
                    rol = getattr(request.user.perfilusuario, 'rol', 'AYUDANTE')
                    if rol in ['ADMIN3_NOTIFICADOR', 'ADMIN1_AGENDADOR']:
                        return redirect('admin3_dashboard')
                    else:
                        return redirect('ayudante_dashboard')
            except Exception as e:
                messages.error(request, f'Error al registrar notificación: {str(e)}')
    else:
        form = NotificacionForm(instance=notif_existente)

    rr.RR_NUM = rr.numero
    rr.RR_FECPRESEN = rr.fecha_presentacion
    context = {
        'form': form,
        'rr': rr,
        'next_url': next_url,
        'titulo': f'Registrar Notificación - RR {rr.numero}',
    }

    return render(request, 'tpe_app/admin3/registrar_notificacion_rr.html', context)


@rol_requerido('AYUDANTE', 'ADMIN3_NOTIFICADOR')
def ayudante_registrar_notificacion_auto(request, auto_id):
    """Registrar la notificación de un Auto TPE existente"""
    from ..models import Notificacion

    auto = get_object_or_404(AUTOTPE, id=auto_id)
    notif_existente = getattr(auto, 'notificacion', None)
    next_url = request.GET.get('next', '').strip()

    if request.method == 'POST':
        form = NotificacionForm(request.POST, instance=notif_existente)
        if form.is_valid():
            try:
                with transaction.atomic():
                    notif = form.save(commit=False)
                    notif.autotpe = auto
                    notif.save()
                    if auto.tipo == 'AUTO_EJECUTORIA' and auto.sim:
                        sim = auto.sim
                        ESTADOS_ACTIVOS_EXTERNOS = {'PROCESO_EN_EL_TSP', 'CUMPLIMIENTO_EN_TPE'}
                        if (sim.estado not in ESTADOS_ACTIVOS_EXTERNOS and
                                sim.fase not in ['EJECUTORIA_NOTIFICADA', 'PENDIENTE_ARCHIVO', 'CONCLUIDO']):
                            sim.fase = 'EJECUTORIA_NOTIFICADA'
                            sim.save()
                    messages.success(request, f'Notificación de Auto {auto.numero} registrada exitosamente')
                    if next_url:
                        return redirect(next_url)
                    rol = getattr(request.user.perfilusuario, 'rol', 'AYUDANTE')
                    if rol in ['ADMIN3_NOTIFICADOR', 'ADMIN1_AGENDADOR']:
                        return redirect('admin3_dashboard')
                    else:
                        return redirect('ayudante_dashboard')
            except Exception as e:
                messages.error(request, f'Error al registrar notificación: {str(e)}')
    else:
        form = NotificacionForm(instance=notif_existente)

    context = {
        'form': form,
        'auto': auto,
        'next_url': next_url,
        'titulo': f'Registrar Notificación - Auto {auto.numero}',
    }

    return render(request, 'tpe_app/ayudante/registrar_notificacion_auto.html', context)


@rol_requerido('AYUDANTE', 'ADMIN3_NOTIFICADOR')
def ayudante_lista_res_sin_pdf(request):
    """Lista de RES pendientes de subir PDF"""

    from ..models import DocumentoAdjunto
    res_con_pdf = DocumentoAdjunto.objects.filter(
        resolucion__isnull=False
    ).values_list('resolucion_id', flat=True)

    resoluciones_sin_pdf = (
        Resolucion.objects
        .filter(instancia='PRIMERA')
        .exclude(id__in=res_con_pdf)
        .select_related('pm', 'sim')
        .order_by('-fecha')
    )

    # Contadores
    total_sin_pdf = resoluciones_sin_pdf.count()
    total_con_pdf = Resolucion.objects.filter(
        instancia='PRIMERA', id__in=res_con_pdf
    ).count()
    total_res = Resolucion.objects.filter(instancia='PRIMERA').count()

    context = {
        'resoluciones': resoluciones_sin_pdf,
        'total_sin_pdf': total_sin_pdf,
        'total_con_pdf': total_con_pdf,
        'total_res': total_res,
    }

    return render(request, 'tpe_app/ayudante/lista_res_sin_pdf.html', context)


# ============================================================================
# WIZARD DE INGRESO RÁPIDO — 6 VISTAS NUEVAS
# ============================================================================

@rol_requerido('AYUDANTE')
def ayudante_wizard_buscar_sim(request):
    """AJAX: busca SIM por código y devuelve TODAS las versiones"""
    codigo = (request.GET.get('q') or '').strip().upper()
    if not codigo:
        return JsonResponse({'found': False})

    sims = list(SIM.objects.filter(codigo=codigo).order_by('version').values('id', 'codigo', 'version', 'resumen', 'estado'))

    if sims:
        # Por defecto seleccionar la última versión
        ultima_version = sims[-1]

        # Preparar lista de versiones para mostrar
        versiones = []
        for sim in sims:
            versiones.append({
                'id': sim['id'],
                'version': sim['version'],
                'label': f"v{sim['version']} - {sim['resumen'][:50] if sim['resumen'] else 'Sin resumen'} ({sim['estado']})",
                'resumen': sim['resumen'][:100] if sim['resumen'] else '',
            })

        return JsonResponse({
            'found': True,
            'versions': versiones,
            'ultima_version_id': ultima_version['id'],
            'ultima_version': ultima_version['version'],
            'sim_cod': codigo,
        })
    return JsonResponse({'found': False})


@rol_requerido('AYUDANTE')
def ayudante_wizard_paso1(request):
    """PASO 1 — Crear o seleccionar SIM (solo datos del sumario)"""
    if request.method == 'POST':
        sim_existente_id = request.POST.get('sim_existente_id')
        if sim_existente_id:
            sim = get_object_or_404(SIM, pk=sim_existente_id)
            messages.info(request, f'Usando SIM existente: {sim.codigo}')
            return redirect('ayudante_wizard_paso2', sim_id=sim.pk)

        form = WizardSIMForm(request.POST)

        if form.is_valid():
            try:
                with transaction.atomic():
                    sim = form.save()
                messages.success(request, f'SIM {sim.codigo} creado. Ahora agregue los militares.')
                return redirect('ayudante_wizard_paso2', sim_id=sim.pk)
            except Exception as e:
                messages.error(request, f'Error al guardar SIM: {str(e)}')
        else:
            # Debug: mostrar todos los errores en mensajes
            error_list = []
            for field, errors in form.errors.items():
                for error in errors:
                    error_list.append(f"{field}: {error}")
            if error_list:
                for err in error_list:
                    messages.error(request, err)
            else:
                messages.error(request, 'Por favor corrija los errores.')
    else:
        form = WizardSIMForm()

    return render(request, 'tpe_app/ayudante/wizard/paso1_sim.html', {
        'form': form,
        'paso_actual': 1,
        'total_pasos': 4,
    })


@rol_requerido('AYUDANTE')
def ayudante_wizard_paso2(request, sim_id):
    """PASO 2 — Agregar/quitar militares al SIM (un formulario simple por militar)"""
    sim = get_object_or_404(SIM, pk=sim_id)
    form_data = {}

    if request.method == 'POST':
        action = request.POST.get('action', '')

        if action == 'continue':
            if sim.militares.exists():
                return redirect('ayudante_wizard_paso2b', sim_id=sim.pk)
            else:
                messages.warning(request, 'Debe agregar al menos un militar antes de continuar.')

        elif action == 'remove':
            pm_id = request.POST.get('pm_id')
            if pm_id:
                PM_SIM.objects.filter(sim=sim, pm_id=pm_id).delete()
                messages.info(request, 'Militar quitado del sumario.')
            return redirect('ayudante_wizard_paso2', sim_id=sim.pk)

        elif action == 'add_one':
            ci_raw      = (request.POST.get('ci') or '').strip()
            nombre      = (request.POST.get('nombre') or '').strip().upper()
            paterno     = (request.POST.get('paterno') or '').strip().upper()
            materno     = (request.POST.get('materno') or '').strip().upper() or None
            grado_fecha = request.POST.get('grado_en_fecha') or None
            escalafon   = request.POST.get('escalafon') or None
            arma        = request.POST.get('arma') or None
            especialidad = (request.POST.get('especialidad') or '').strip() or None
            anio_raw    = request.POST.get('anio_promocion') or None

            # Guardar datos para repoblar el formulario en caso de error
            form_data = request.POST.dict()

            if not nombre or not paterno:
                messages.error(request, 'Nombre y Apellido Paterno son obligatorios.')
            else:
                ci_val = None
                if ci_raw:
                    if not ci_raw.isdigit():
                        messages.error(request, 'El CI debe contener solo números.')
                        ci_raw = None
                    else:
                        ci_val = int(ci_raw)

                anio_val = None
                if anio_raw:
                    try:
                        anio_val = int(anio_raw)
                        if not (1950 <= anio_val <= 2100):
                            messages.error(request, 'Año de egreso fuera de rango.')
                            anio_val = None
                    except ValueError:
                        messages.error(request, 'Año de egreso inválido.')

                if not messages.get_messages(request):
                    try:
                        with transaction.atomic():
                            # Buscar PM existente por CI o por nombre+paterno
                            pm = None
                            pm_existia = False
                            if ci_val:
                                pm = PM.objects.filter(ci=ci_val).first()
                            if not pm:
                                q = PM.objects.filter(nombre=nombre, paterno=paterno)
                                if materno:
                                    q = q.filter(materno=materno)
                                pm = q.first()

                            if pm:
                                pm_existia = True

                            # Si no existe, crearlo
                            if not pm:
                                pm = PM.objects.create(
                                    ci=ci_val,
                                    nombre=nombre,
                                    paterno=paterno,
                                    materno=materno,
                                    grado=grado_fecha,
                                    escalafon=escalafon,
                                    arma=arma,
                                    especialidad=especialidad,
                                    anio_promocion=anio_val,
                                )

                            # Crear o actualizar la relación PM_SIM
                            pm_sim, created = PM_SIM.objects.get_or_create(sim=sim, pm=pm)
                            if grado_fecha:
                                pm_sim.grado_en_fecha = grado_fecha
                                pm_sim.save(update_fields=['grado_en_fecha'])

                            if created:
                                messages.success(request, f'{pm.grado or ""} {pm.paterno} {pm.nombre} agregado al sumario.')
                            else:
                                messages.info(request, f'{pm.paterno} {pm.nombre} ya estaba en el sumario.')

                            # Advertir si el PM ya tiene otros sumarios en el sistema
                            if pm_existia:
                                otros = PM_SIM.objects.filter(pm=pm).exclude(sim=sim).select_related('sim').order_by('-sim__fecha_ingreso')
                                if otros.exists():
                                    codigos = ', '.join(ps.sim.codigo for ps in otros[:5])
                                    total = otros.count()
                                    sufijo = f' (y {total - 5} más)' if total > 5 else ''
                                    messages.warning(
                                        request,
                                        f'ANTECEDENTE: {pm.grado or ""} {pm.paterno} {pm.nombre} '
                                        f'ya figura en {total} sumario(s) previo(s): {codigos}{sufijo}.'
                                    )

                            form_data = {}  # Limpiar formulario tras éxito
                    except Exception as e:
                        messages.error(request, f'Error al agregar militar: {str(e)}')

    militares_actuales = sim.militares.all().order_by('paterno', 'nombre')
    return render(request, 'tpe_app/ayudante/wizard/paso2_pm.html', {
        'sim': sim,
        'militares_actuales': militares_actuales,
        'form_data': form_data,
        'grado_choices': PM.GRADO_CHOICES,
        'escalafon_choices': PM.ESCALAFON_CHOICES,
        'arma_choices': PM.ARMA_CHOICES,
        'paso_actual': 2,
        'total_pasos': 4,
    })


@rol_requerido('AYUDANTE')
def ayudante_wizard_paso2b(request, sim_id):
    """PASO 2.5 — Seleccionar militar para llenar documentos"""
    sim = get_object_or_404(SIM, pk=sim_id)
    militares = sim.militares.all().order_by('paterno', 'nombre')

    if not militares.exists():
        messages.warning(request, 'Debe agregar al menos un militar antes de continuar.')
        return redirect('ayudante_wizard_paso2', sim_id=sim.pk)

    if request.method == 'POST':
        pm_id = request.POST.get('pm_id')
        if pm_id:
            try:
                pm = PM.objects.get(id=pm_id, pm_sim__sim=sim)
                return redirect('ayudante_wizard_paso3', sim_id=sim.pk, pm_id=pm.pk)
            except PM.DoesNotExist:
                messages.error(request, 'Militar no encontrado en este sumario.')
        else:
            messages.error(request, 'Debe seleccionar un militar.')

    return render(request, 'tpe_app/ayudante/wizard/paso2b_militar.html', {
        'sim': sim,
        'militares': militares,
        'paso_actual': '2.5',
        'total_pasos': 4,
    })


@rol_requerido('AYUDANTE')
def ayudante_wizard_paso3(request, sim_id, pm_id=None):
    """PASO 3 — Primera Resolución + RR opcional (para un militar específico)"""
    sim = get_object_or_404(SIM, pk=sim_id)

    # Determinar qué militar estamos procesando
    if pm_id:
        pm = get_object_or_404(PM, id=pm_id, pm_sim__sim=sim)
    else:
        # Retrocompatibilidad: usar el primer militar si no se especifica
        pm = sim.militares.first()
        if not pm:
            messages.warning(request, 'No hay militares en este sumario.')
            return redirect('ayudante_wizard_paso2', sim_id=sim.pk)

    res_existente = Resolucion.objects.filter(sim=sim, pm=pm, instancia='PRIMERA').first()
    rr_existente = Resolucion.objects.filter(sim=sim, pm=pm, instancia='RECONSIDERACION').first()

    if request.method == 'POST':
        action = request.POST.get('action', 'save')

        if action == 'skip':
            return redirect('ayudante_wizard_paso4', sim_id=sim.pk)

        guardar_res = request.POST.get('guardar_res') == '1'
        guardar_rr = request.POST.get('guardar_rr') == '1'

        res_form = WizardRESForm(request.POST if guardar_res else None, instance=res_existente, prefix='res')
        rr_form = WizardRRForm(request.POST if guardar_rr else None, instance=rr_existente, prefix='rr')

        errores = False
        try:
            with transaction.atomic():
                if guardar_res and res_form.is_valid():
                    res = res_form.save(commit=False)
                    res.sim = sim
                    res.pm = pm
                    res.save()
                    res_existente = res

                    # Guardar notificación RES si se proporciona
                    res_notif_tipo = request.POST.get('res_notif_tipo')
                    if res_notif_tipo:
                        # Eliminar notificación anterior si existe
                        res.notificacion.delete() if hasattr(res, 'notificacion') else None
                        notif = Notificacion(
                            resolucion=res,
                            tipo=res_notif_tipo,
                            notificado_a=request.POST.get('res_notif_notificado_a', ''),
                            fecha=request.POST.get('res_notif_fecha') or None,
                            hora=request.POST.get('res_notif_hora') or None
                        )
                        notif.save()

                    if sim.fase not in ['1RA_RESOLUCION', '2DA_RESOLUCION', 'NOTIFICADO_1RA', 'NOTIFICADO_RR', 'ELEVADO_TSP', 'CONCLUIDO']:
                        sim.fase = '1RA_RESOLUCION'
                        sim.save()

                elif guardar_res and not res_form.is_valid():
                    errores = True

                if guardar_rr and res_existente:
                    if rr_form.is_valid():
                        rr = rr_form.save(commit=False)
                        rr.sim = sim
                        rr.pm = res_existente.pm
                        rr.resolucion_origen = res_existente
                        rr.save()

                        # Guardar notificación RR si se proporciona
                        rr_notif_tipo = request.POST.get('rr_notif_tipo')
                        if rr_notif_tipo:
                            # Eliminar notificación anterior si existe
                            rr.notificacion.delete() if hasattr(rr, 'notificacion') else None
                            notif = Notificacion(
                                resolucion=rr,
                                tipo=rr_notif_tipo,
                                notificado_a=request.POST.get('rr_notif_notificado_a', ''),
                                fecha=request.POST.get('rr_notif_fecha') or None,
                                hora=request.POST.get('rr_notif_hora') or None
                            )
                            notif.save()

                        if sim.fase not in ['2DA_RESOLUCION', 'NOTIFICADO_RR', 'ELEVADO_TSP', 'CONCLUIDO']:
                            sim.fase = '2DA_RESOLUCION'
                            sim.save()
                    else:
                        errores = True

                if not errores:
                    messages.success(request, 'Resoluciones guardadas correctamente.')
                    return redirect('ayudante_wizard_paso4', sim_id=sim.pk, pm_id=pm.pk)

        except Exception as e:
            messages.error(request, f'Error al guardar: {str(e)}')

        if errores:
            messages.error(request, 'Por favor corrija los errores.')

    else:
        res_form = WizardRESForm(instance=res_existente, prefix='res')
        rr_form = WizardRRForm(instance=rr_existente, prefix='rr')

    # Solo mostrar el militar actual, no permitir cambio
    res_form.fields['pm'].queryset = PM.objects.filter(id=pm.id)
    res_form.fields['pm'].initial = pm

    # Obtener otros militares para el botón "Siguiente"
    otros_militares = sim.militares.exclude(id=pm.id).order_by('paterno', 'nombre')

    return render(request, 'tpe_app/ayudante/wizard/paso3_resoluciones.html', {
        'sim': sim,
        'pm': pm,
        'res_form': res_form,
        'rr_form': rr_form,
        'res_existente': res_existente,
        'rr_existente': rr_existente,
        'otros_militares': otros_militares,
        'paso_actual': 3,
        'total_pasos': 4,
    })


@rol_requerido('AYUDANTE')
def ayudante_wizard_paso4(request, sim_id, pm_id=None):
    """PASO 4 — Auto TPE, RAP, RAEE, Auto TSP (para un militar específico)"""
    sim = get_object_or_404(SIM, pk=sim_id)

    # Determinar qué militar estamos procesando
    if pm_id:
        pm = get_object_or_404(PM, id=pm_id, pm_sim__sim=sim)
    else:
        # Retrocompatibilidad: usar el primer militar si no se especifica
        pm = sim.militares.first()
        if not pm:
            messages.warning(request, 'No hay militares en este sumario.')
            return redirect('ayudante_wizard_paso2', sim_id=sim.pk)

    autotpe_existente = AUTOTPE.objects.filter(sim=sim, pm=pm).first()
    rap_existente = RecursoTSP.objects.filter(sim=sim, pm=pm, instancia='APELACION').first()
    raee_existente = RecursoTSP.objects.filter(sim=sim, pm=pm, instancia='ACLARACION_ENMIENDA').first()
    autotsp_existente = AUTOTSP.objects.filter(sim=sim).first()

    # Obtener otros militares para el botón "Siguiente"
    otros_militares = sim.militares.exclude(id=pm.id).order_by('paterno', 'nombre')

    if request.method == 'POST':
        action = request.POST.get('action', 'save')

        if action == 'skip':
            # Si hay otros militares, volver a paso 2.5, si no ir al resumen
            if otros_militares.exists():
                return redirect('ayudante_wizard_paso2b', sim_id=sim.pk)
            else:
                return redirect('ayudante_wizard_resumen', sim_id=sim.pk)

        guardar_autotpe = request.POST.get('guardar_autotpe') == '1'
        guardar_rap = request.POST.get('guardar_rap') == '1'
        guardar_raee = request.POST.get('guardar_raee') == '1'
        guardar_autotsp = request.POST.get('guardar_autotsp') == '1'

        autotpe_form = WizardAUTOTPEForm(request.POST if guardar_autotpe else None, instance=autotpe_existente, prefix='autotpe')
        rap_form = WizardRAPForm(request.POST if guardar_rap else None, instance=rap_existente, prefix='rap', sim=sim)
        raee_form = WizardRAEEForm(request.POST if guardar_raee else None, instance=raee_existente, prefix='raee', sim=sim)
        autotsp_form = WizardAUTOTSPForm(request.POST if guardar_autotsp else None, instance=autotsp_existente, prefix='autotsp')

        errores = False
        try:
            with transaction.atomic():
                if guardar_autotpe:
                    if autotpe_form.is_valid():
                        auto = autotpe_form.save(commit=False)
                        auto.sim = sim
                        auto.pm = pm
                        auto.save()
                        autotpe_existente = auto

                        # Procesar notificación de Auto TPE
                        autotpe_notif_tipo = request.POST.get('autotpe_notif_tipo', '').strip()
                        if autotpe_notif_tipo:
                            Notificacion.objects.update_or_create(
                                autotpe=auto,
                                defaults={
                                    'tipo': autotpe_notif_tipo,
                                    'notificado_a': request.POST.get('autotpe_notif_notificado_a', '').strip(),
                                    'fecha': request.POST.get('autotpe_notif_fecha') or None,
                                    'hora': request.POST.get('autotpe_notif_hora') or None,
                                }
                            )

                        # Procesar memorándum si es AUTO_EJECUTORIA
                        if auto.tipo == 'AUTO_EJECUTORIA':
                            autotpe_memo_numero = request.POST.get('autotpe_memo_numero', '').strip()
                            autotpe_memo_fecha = request.POST.get('autotpe_memo_fecha', '')
                            autotpe_memo_fecha_entrega = request.POST.get('autotpe_memo_fecha_entrega', '')

                            if autotpe_memo_numero:
                                Memorandum.objects.update_or_create(
                                    autotpe=auto,
                                    defaults={
                                        'numero': autotpe_memo_numero,
                                        'fecha': autotpe_memo_fecha or None,
                                        'fecha_entrega': autotpe_memo_fecha_entrega or None,
                                    }
                                )
                    else:
                        errores = True

                if guardar_rap:
                    if rap_form.is_valid():
                        rap = rap_form.save(commit=False)
                        rap.sim = sim
                        rap.pm = pm
                        rap.instancia = 'APELACION'
                        rap.save()
                        if sim.fase not in ['ELEVADO_TSP', 'CONCLUIDO']:
                            sim.fase = 'ELEVADO_TSP'
                            sim.save()
                        rap_existente = rap

                        # Procesar notificación de RAP
                        rap_notif_tipo = request.POST.get('rap_notif_tipo', '').strip()
                        if rap_notif_tipo:
                            Notificacion.objects.update_or_create(
                                recurso_tsp=rap,
                                defaults={
                                    'tipo': rap_notif_tipo,
                                    'notificado_a': request.POST.get('rap_notif_notificado_a', '').strip(),
                                    'fecha': request.POST.get('rap_notif_fecha') or None,
                                    'hora': request.POST.get('rap_notif_hora') or None,
                                }
                            )
                    else:
                        errores = True

                if guardar_raee:
                    if raee_form.is_valid():
                        raee = raee_form.save(commit=False)
                        raee.sim = sim
                        raee.pm = pm
                        raee.instancia = 'ACLARACION_ENMIENDA'
                        raee.save()
                        raee_existente = raee

                        # Procesar notificación de RAEE
                        raee_notif_tipo = request.POST.get('raee_notif_tipo', '').strip()
                        if raee_notif_tipo:
                            Notificacion.objects.update_or_create(
                                recurso_tsp=raee,
                                defaults={
                                    'tipo': raee_notif_tipo,
                                    'notificado_a': request.POST.get('raee_notif_notificado_a', '').strip(),
                                    'fecha': request.POST.get('raee_notif_fecha') or None,
                                    'hora': request.POST.get('raee_notif_hora') or None,
                                }
                            )
                    else:
                        errores = True

                if guardar_autotsp:
                    if autotsp_form.is_valid():
                        autotsp = autotsp_form.save(commit=False)
                        autotsp.sim = sim
                        autotsp.save()
                        autotsp_existente = autotsp

                        # Procesar notificación de Auto TSP
                        autotsp_notif_tipo = request.POST.get('autotsp_notif_tipo', '').strip()
                        if autotsp_notif_tipo:
                            Notificacion.objects.update_or_create(
                                autotsp=autotsp,
                                defaults={
                                    'tipo': autotsp_notif_tipo,
                                    'notificado_a': request.POST.get('autotsp_notif_notificado_a', '').strip(),
                                    'fecha': request.POST.get('autotsp_notif_fecha') or None,
                                    'hora': request.POST.get('autotsp_notif_hora') or None,
                                }
                            )
                    else:
                        errores = True

                if not errores:
                    messages.success(request, 'Documentos guardados correctamente.')
                    # Si hay otros militares, volver a paso 2.5, si no ir al resumen
                    if otros_militares.exists():
                        return redirect('ayudante_wizard_paso2b', sim_id=sim.pk)
                    else:
                        return redirect('ayudante_wizard_resumen', sim_id=sim.pk)

        except Exception as e:
            messages.error(request, f'Error al guardar: {str(e)}')

        if errores:
            messages.error(request, 'Por favor corrija los errores en los formularios activos.')

    else:
        autotpe_form = WizardAUTOTPEForm(instance=autotpe_existente, prefix='autotpe')
        rap_form = WizardRAPForm(instance=rap_existente, prefix='rap', sim=sim)
        raee_form = WizardRAEEForm(instance=raee_existente, prefix='raee', sim=sim)
        autotsp_form = WizardAUTOTSPForm(instance=autotsp_existente, prefix='autotsp')

    # Solo mostrar el militar actual, no permitir cambio
    autotpe_form.fields['pm'].queryset = PM.objects.filter(id=pm.id)
    autotpe_form.fields['pm'].initial = pm

    return render(request, 'tpe_app/ayudante/wizard/paso4_autos.html', {
        'sim': sim,
        'pm': pm,
        'autotpe_form': autotpe_form,
        'rap_form': rap_form,
        'raee_form': raee_form,
        'autotsp_form': autotsp_form,
        'autotpe_existente': autotpe_existente,
        'rap_existente': rap_existente,
        'raee_existente': raee_existente,
        'autotsp_existente': autotsp_existente,
        'otros_militares': otros_militares,
        'paso_actual': 4,
        'total_pasos': 4,
    })


@rol_requerido('AYUDANTE')
def ayudante_wizard_resumen(request, sim_id):
    """Vista de resumen final del wizard — solo lectura"""
    sim = get_object_or_404(SIM.objects.prefetch_related('militares'), pk=sim_id)
    resoluciones = Resolucion.objects.filter(sim=sim).order_by('fecha')
    autos_tpe = AUTOTPE.objects.filter(sim=sim).order_by('fecha')
    recursos_tsp = RecursoTSP.objects.filter(sim=sim).order_by('fecha')
    autos_tsp = AUTOTSP.objects.filter(sim=sim).order_by('fecha')

    return render(request, 'tpe_app/ayudante/wizard/resumen.html', {
        'sim': sim,
        'resoluciones': resoluciones,
        'autos_tpe': autos_tpe,
        'recursos_tsp': recursos_tsp,
        'autos_tsp': autos_tsp,
        'paso_actual': 5,
        'total_pasos': 4,
    })


# ============================================================================
# EDICIÓN DE PERSONAL MILITAR — Grado actual, año de egreso, no ascendió
# ============================================================================

@rol_requerido('AYUDANTE', 'ADMIN1_AGENDADOR')
def ayudante_editar_pm(request, pm_id):
    """Permite al Ayudante actualizar grado actual, año de egreso y estado de ascenso."""
    pm = get_object_or_404(PM, pk=pm_id)

    if request.method == 'POST':
        grado      = request.POST.get('grado') or None
        escalafon  = request.POST.get('escalafon') or None
        estado     = request.POST.get('estado') or pm.estado
        promocion  = request.POST.get('anio_promocion') or None
        no_asc     = request.POST.get('no_ascendio') == 'on'
        arma       = request.POST.get('arma') or None
        ci_raw     = request.POST.get('ci') or None
        foto       = request.FILES.get('foto')

        if promocion:
            try:
                promocion = int(promocion)
                if not (1950 <= promocion <= 2100):
                    messages.error(request, 'Año de egreso fuera de rango (1950-2100).')
                    promocion = pm.anio_promocion
            except ValueError:
                messages.error(request, 'Año de egreso inválido.')
                promocion = pm.anio_promocion
        else:
            promocion = None

        ci_val = None
        if ci_raw:
            try:
                ci_val = int(ci_raw)
            except ValueError:
                messages.error(request, 'Número de CI inválido.')
                ci_val = pm.ci

        pm.grado          = grado
        pm.escalafon      = escalafon
        pm.estado         = estado
        pm.anio_promocion = promocion
        pm.no_ascendio    = no_asc
        pm.arma           = arma
        pm.ci             = ci_val

        update_fields = ['grado', 'escalafon', 'estado', 'anio_promocion', 'no_ascendio', 'arma', 'ci']
        if foto:
            content_type = getattr(foto, 'content_type', '') or ''
            if content_type.startswith('image/'):
                if pm.foto:
                    pm.foto.delete(save=False)
                pm.foto = foto
                update_fields.append('foto')
            else:
                messages.error(request, 'La foto debe ser una imagen (JPG, PNG).')

        pm.save(update_fields=update_fields)

        messages.success(request, f'Datos de {pm.nombre} {pm.paterno} actualizados.')
        next_url = request.POST.get('next') or request.GET.get('next') or 'ayudante_dashboard'
        return redirect(next_url)

    return render(request, 'tpe_app/ayudante/editar_pm.html', {
        'pm': pm,
        'grado_choices': PM.GRADO_CHOICES,
        'escalafon_choices': PM.ESCALAFON_CHOICES,
        'estado_choices': PM.ESTADO_CHOICES,
        'arma_choices': PM.ARMA_CHOICES,
        'grado_esperado': pm.grado_esperado,
        'estado_calculado': pm.estado_carrera_calculado,
        'años_servicio': pm.años_servicio,
    })


# ============================================================================
# TABLA DE DOCUMENTOS — Filtrado por tipo y gestión con exportación
# ============================================================================

@rol_requerido('AYUDANTE', 'ADMIN3_NOTIFICADOR', 'ADMIN1_AGENDADOR')
def ayudante_tabla_documentos(request):
    """Tabla filtrada de Resoluciones, RR y Autos TPE con datos del militar"""
    from django.core.paginator import Paginator
    from urllib.parse import urlencode

    tipo_doc = request.GET.get('tipo_doc', 'resolucion')
    gestion = request.GET.get('gestion', '')
    page_num = request.GET.get('page', 1)

    if tipo_doc == 'reconsideracion':
        qs = (Resolucion.objects
              .filter(instancia='RECONSIDERACION')
              .select_related('pm', 'sim'))
        tipo_label = 'Reconsideración (RR)'
        gestiones_disponibles = list(
            Resolucion.objects.filter(instancia='RECONSIDERACION')
            .exclude(fecha__isnull=True)
            .values_list('fecha__year', flat=True)
            .distinct().order_by('-fecha__year')
        )
    elif tipo_doc == 'autotpe':
        qs = (AUTOTPE.objects.all()
              .select_related('pm', 'sim'))
        tipo_label = 'Auto TPE'
        gestiones_disponibles = list(
            AUTOTPE.objects.exclude(fecha__isnull=True)
            .values_list('fecha__year', flat=True)
            .distinct().order_by('-fecha__year')
        )
    else:
        tipo_doc = 'resolucion'
        qs = (Resolucion.objects
              .filter(instancia='PRIMERA')
              .select_related('pm', 'sim'))
        tipo_label = 'Resolución (1ª Instancia)'
        gestiones_disponibles = list(
            Resolucion.objects.filter(instancia='PRIMERA')
            .exclude(fecha__isnull=True)
            .values_list('fecha__year', flat=True)
            .distinct().order_by('-fecha__year')
        )

    if gestion:
        try:
            qs = qs.filter(fecha__year=int(gestion))
        except (ValueError, TypeError):
            gestion = ''

    qs = qs.order_by('-fecha', 'pm__paterno', 'pm__nombre')

    paginator = Paginator(qs, 25)
    page_obj = paginator.get_page(page_num)
    items = list(page_obj)
    ids_page = [obj.pk for obj in items]

    # Pre-fetch PDFs para esta página
    if tipo_doc in ('resolucion', 'reconsideracion'):
        docs_con_pdf = set(
            DocumentoAdjunto.objects.filter(resolucion_id__in=ids_page)
            .values_list('resolucion_id', flat=True)
        )
        notifs_map = {
            n.resolucion_id: n
            for n in Notificacion.objects.filter(resolucion_id__in=ids_page)
        }
    else:
        docs_con_pdf = set(
            DocumentoAdjunto.objects.filter(autotpe_id__in=ids_page)
            .values_list('autotpe_id', flat=True)
        )
        notifs_map = {
            n.autotpe_id: n
            for n in Notificacion.objects.filter(autotpe_id__in=ids_page)
        }

    # Para documentos sin PM directo, buscar el PM por cualquier vía (registros históricos)
    sims_sin_pm = [obj.sim_id for obj in items if not obj.pm_id and obj.sim_id]
    pm_por_sim = {}
    if sims_sin_pm:
        # 1) Tabla puente PM_SIM
        for ps in PM_SIM.objects.filter(sim_id__in=sims_sin_pm).select_related('pm'):
            if ps.sim_id not in pm_por_sim:
                pm_por_sim[ps.sim_id] = ps.pm
        # 2) Otros documentos del mismo SIM que sí tienen PM
        still_missing = [s for s in sims_sin_pm if s not in pm_por_sim]
        if still_missing:
            for res in Resolucion.objects.filter(
                    sim_id__in=still_missing, pm__isnull=False).select_related('pm'):
                if res.sim_id not in pm_por_sim:
                    pm_por_sim[res.sim_id] = res.pm
            for auto in AUTOTPE.objects.filter(
                    sim_id__in=still_missing, pm__isnull=False).select_related('pm'):
                if auto.sim_id not in pm_por_sim:
                    pm_por_sim[auto.sim_id] = auto.pm

    for item in items:
        item.notif_obj = notifs_map.get(item.pk)
        item.has_pdf = item.pk in docs_con_pdf
        item.pm_efectivo = item.pm if item.pm_id else pm_por_sim.get(item.sim_id)

    params = {k: v for k, v in request.GET.items() if k != 'page'}
    back_url = request.path + ('?' + urlencode(params) if params else '')

    context = {
        'items': items,
        'page_obj': page_obj,
        'tipo_doc': tipo_doc,
        'tipo_label': tipo_label,
        'gestion': gestion,
        'gestiones_disponibles': gestiones_disponibles,
        'back_url': back_url,
        'total_count': paginator.count,
    }
    return render(request, 'tpe_app/ayudante/tabla_documentos.html', context)


@rol_requerido('AYUDANTE', 'ADMIN3_NOTIFICADOR')
def subir_pdf_autotpe(request, auto_id):
    """Sube o reemplaza el PDF de un Auto TPE (acepta next para redirigir)"""
    auto = get_object_or_404(AUTOTPE, pk=auto_id)

    if request.method == 'POST':
        archivo_pdf = request.FILES.get('archivo_pdf')
        next_url = request.POST.get('next', '').strip()

        if not archivo_pdf:
            messages.error(request, 'Selecciona un archivo PDF')
        elif not archivo_pdf.name.lower().endswith('.pdf'):
            messages.error(request, 'Solo se permiten archivos PDF')
        else:
            try:
                with transaction.atomic():
                    DocumentoAdjunto.objects.filter(autotpe_id=auto.pk).delete()
                    pm_label = (f'{auto.pm.grado} {auto.pm.paterno}' if auto.pm else 'S/N')
                    DocumentoAdjunto.objects.create(
                        autotpe=auto,
                        tipo='auto',
                        archivo=archivo_pdf,
                        nombre=f'AUTO {auto.numero} - {pm_label}',
                    )
                    messages.success(request, f'PDF del Auto {auto.numero} subido correctamente')
                    return redirect(next_url or 'ayudante_dashboard')
            except Exception as e:
                messages.error(request, f'Error al subir PDF: {str(e)}')

        return redirect(next_url or 'ayudante_dashboard')

    return redirect('ayudante_dashboard')


@rol_requerido('AYUDANTE', 'ADMIN3_NOTIFICADOR', 'ADMIN1_AGENDADOR')
def ayudante_tabla_export_pdf(request):
    """Exporta a PDF la tabla filtrada de documentos"""
    from django.http import HttpResponse
    from io import BytesIO
    import os
    from reportlab.lib.pagesizes import letter, landscape
    from reportlab.lib.units import cm
    from reportlab.lib import colors
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont

    _fonts_dir = r'C:\Windows\Fonts'
    _arial = os.path.join(_fonts_dir, 'arial.ttf')
    _arial_bold = os.path.join(_fonts_dir, 'arialbd.ttf')
    if os.path.exists(_arial):
        try:
            pdfmetrics.registerFont(TTFont('Arial', _arial))
            pdfmetrics.registerFont(TTFont('Arial-Bold', _arial_bold))
        except Exception:
            pass
        FONT_NORMAL = 'Arial'
        FONT_BOLD = 'Arial-Bold'
    else:
        FONT_NORMAL = 'Helvetica'
        FONT_BOLD = 'Helvetica-Bold'

    tipo_doc = request.GET.get('tipo_doc', 'resolucion')
    gestion = request.GET.get('gestion', '')

    if tipo_doc == 'reconsideracion':
        qs = Resolucion.objects.filter(instancia='RECONSIDERACION').select_related('pm', 'sim')
        titulo_tipo = 'RECONSIDERACIONES (RR)'
    elif tipo_doc == 'autotpe':
        qs = AUTOTPE.objects.all().select_related('pm', 'sim')
        titulo_tipo = 'AUTOS TPE'
    else:
        qs = Resolucion.objects.filter(instancia='PRIMERA').select_related('pm', 'sim')
        titulo_tipo = 'RESOLUCIONES (1RA INSTANCIA)'

    if gestion:
        try:
            qs = qs.filter(fecha__year=int(gestion))
        except (ValueError, TypeError):
            pass

    qs = qs.order_by('-fecha', 'pm__paterno', 'pm__nombre')
    items = list(qs)
    ids_all = [obj.pk for obj in items]

    if tipo_doc in ('resolucion', 'reconsideracion'):
        docs_con_pdf = set(
            DocumentoAdjunto.objects.filter(resolucion_id__in=ids_all)
            .values_list('resolucion_id', flat=True)
        )
        notifs_map = {
            n.resolucion_id: n
            for n in Notificacion.objects.filter(resolucion_id__in=ids_all)
        }
    else:
        docs_con_pdf = set(
            DocumentoAdjunto.objects.filter(autotpe_id__in=ids_all)
            .values_list('autotpe_id', flat=True)
        )
        notifs_map = {
            n.autotpe_id: n
            for n in Notificacion.objects.filter(autotpe_id__in=ids_all)
        }

    buf = BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=landscape(letter),
        rightMargin=1.5 * cm, leftMargin=1.5 * cm,
        topMargin=1.5 * cm, bottomMargin=1.5 * cm,
    )

    style_title = ParagraphStyle('title', fontName=FONT_BOLD, fontSize=11, spaceAfter=4)
    style_sub = ParagraphStyle('sub', fontName=FONT_NORMAL, fontSize=8, spaceAfter=8, textColor=colors.grey)

    titulo_gestion = f'GESTIÓN {gestion}' if gestion else 'TODAS LAS GESTIONES'
    story = [
        Paragraph(f'TRIBUNAL DE PERSONAL DEL EJÉRCITO — {titulo_tipo}', style_title),
        Paragraph(f'{titulo_gestion}  |  Total: {len(items)} registros', style_sub),
    ]

    headers = ['N°', 'MILITAR', 'FECHA', 'NOTIFICACIÓN', 'PDF', 'FOTO', 'AÑO PROM.', 'ARMA', 'CI', 'ESTADO SIM']
    data = [headers]

    def _fd(d):
        return d.strftime('%d/%m/%y') if d else '—'

    for obj in items:
        notif = notifs_map.get(obj.pk)
        pm = obj.pm
        if pm:
            militar = f'{pm.grado or ""} {pm.paterno or ""} {pm.nombre or ""}'.strip()
            foto_str = 'SI' if pm.foto else '—'
            año_str = str(pm.anio_promocion) if pm.anio_promocion else '—'
            arma_str = pm.arma or '—'
            ci_str = str(int(pm.ci)) if pm.ci else '—'
        else:
            militar = '(SIN MILITAR)'
            foto_str = '—'
            año_str = '—'
            arma_str = '—'
            ci_str = '—'

        notif_str = _fd(notif.fecha) if notif else 'PENDIENTE'
        pdf_str = 'SI' if obj.pk in docs_con_pdf else '—'
        estado_str = obj.sim.get_estado_display() if obj.sim else '—'
        data.append([
            obj.numero or '—', militar, _fd(obj.fecha),
            notif_str, pdf_str, foto_str,
            año_str, arma_str, ci_str, estado_str,
        ])

    col_widths = [1.6 * cm, 6.5 * cm, 1.8 * cm, 2.4 * cm, 1.2 * cm,
                  1.2 * cm, 2 * cm, 1.6 * cm, 2 * cm, 3.5 * cm]
    tbl = Table(data, colWidths=col_widths, repeatRows=1)
    tbl.setStyle(TableStyle([
        ('FONTNAME',    (0, 0), (-1, 0),  FONT_BOLD),
        ('FONTSIZE',    (0, 0), (-1, 0),  7),
        ('FONTNAME',    (0, 1), (-1, -1), FONT_NORMAL),
        ('FONTSIZE',    (0, 1), (-1, -1), 6.5),
        ('BACKGROUND',  (0, 0), (-1, 0),  colors.HexColor('#2c3e50')),
        ('TEXTCOLOR',   (0, 0), (-1, 0),  colors.white),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8f9fa')]),
        ('GRID',        (0, 0), (-1, -1), 0.25, colors.HexColor('#dee2e6')),
        ('VALIGN',      (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING',  (0, 0), (-1, -1), 3),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
        ('LEFTPADDING', (0, 0), (-1, -1), 3),
        ('RIGHTPADDING', (0, 0), (-1, -1), 3),
    ]))
    story.append(tbl)
    doc.build(story)

    buf.seek(0)
    fname = f'TPE_{tipo_doc.upper()}{"_" + gestion if gestion else ""}.pdf'
    response = HttpResponse(buf.read(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{fname}"'
    return response


@rol_requerido('AYUDANTE', 'ADMIN3_NOTIFICADOR', 'ADMIN1_AGENDADOR')
def ayudante_tabla_export_excel(request):
    """Exporta a Excel la tabla filtrada de documentos"""
    from django.http import HttpResponse
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    tipo_doc = request.GET.get('tipo_doc', 'resolucion')
    gestion = request.GET.get('gestion', '')

    if tipo_doc == 'reconsideracion':
        qs = Resolucion.objects.filter(instancia='RECONSIDERACION').select_related('pm', 'sim')
        titulo_tipo = 'RECONSIDERACIONES (RR)'
    elif tipo_doc == 'autotpe':
        qs = AUTOTPE.objects.all().select_related('pm', 'sim')
        titulo_tipo = 'AUTOS TPE'
    else:
        qs = Resolucion.objects.filter(instancia='PRIMERA').select_related('pm', 'sim')
        titulo_tipo = 'RESOLUCIONES (1RA INSTANCIA)'

    if gestion:
        try:
            qs = qs.filter(fecha__year=int(gestion))
        except (ValueError, TypeError):
            pass

    qs = qs.order_by('-fecha', 'pm__paterno', 'pm__nombre')
    items = list(qs)
    ids_all = [obj.pk for obj in items]

    if tipo_doc in ('resolucion', 'reconsideracion'):
        docs_con_pdf = set(
            DocumentoAdjunto.objects.filter(resolucion_id__in=ids_all)
            .values_list('resolucion_id', flat=True)
        )
        notifs_map = {
            n.resolucion_id: n
            for n in Notificacion.objects.filter(resolucion_id__in=ids_all)
        }
    else:
        docs_con_pdf = set(
            DocumentoAdjunto.objects.filter(autotpe_id__in=ids_all)
            .values_list('autotpe_id', flat=True)
        )
        notifs_map = {
            n.autotpe_id: n
            for n in Notificacion.objects.filter(autotpe_id__in=ids_all)
        }

    wb = Workbook()
    ws = wb.active
    ws.title = titulo_tipo[:31]

    hdr_fill = PatternFill('solid', fgColor='2C3E50')
    hdr_font = Font(bold=True, color='FFFFFF', size=10, name='Arial')
    data_font = Font(size=9, name='Arial')
    thin = Side(style='thin', color='DEE2E6')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal='center', vertical='center')

    headers = ['N°', 'GRADO', 'PATERNO', 'NOMBRE', 'MATERNO', 'FECHA DOC.',
               'NOTIFICACIÓN', 'PDF', 'FOTO', 'AÑO PROM.', 'ARMA', 'CI', 'ESTADO SIM']
    col_widths = [10, 15, 20, 20, 20, 14, 14, 8, 8, 12, 10, 14, 22]

    for col_idx, (hdr, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col_idx, value=hdr)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = center
        cell.border = border
        ws.column_dimensions[cell.column_letter].width = w

    def _fd(d):
        return d.strftime('%d/%m/%Y') if d else ''

    for row_idx, obj in enumerate(items, 2):
        notif = notifs_map.get(obj.pk)
        pm = obj.pm
        grado_str = pm.grado or '' if pm else ''
        paterno_str = pm.paterno or '' if pm else ''
        nombre_str = pm.nombre or '' if pm else ''
        materno_str = pm.materno or '' if pm else ''
        foto_str = 'SÍ' if (pm and pm.foto) else ''
        año_str = pm.anio_promocion if (pm and pm.anio_promocion) else ''
        arma_str = pm.arma or '' if pm else ''
        ci_str = int(pm.ci) if (pm and pm.ci) else ''
        notif_str = _fd(notif.fecha) if notif else 'PENDIENTE'
        pdf_str = 'SÍ' if obj.pk in docs_con_pdf else ''
        estado_str = obj.sim.get_estado_display() if obj.sim else ''

        row_data = [
            obj.numero or '', grado_str, paterno_str, nombre_str, materno_str,
            _fd(obj.fecha), notif_str, pdf_str, foto_str,
            año_str, arma_str, ci_str, estado_str,
        ]
        alt_fill = PatternFill('solid', fgColor='F8F9FA') if row_idx % 2 == 0 else None
        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = data_font
            cell.border = border
            cell.alignment = Alignment(vertical='center')
            if alt_fill:
                cell.fill = alt_fill

    ws.freeze_panes = 'A2'
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f'TPE_{tipo_doc.upper()}{"_" + gestion if gestion else ""}.xlsx'
    response = HttpResponse(
        buf.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{fname}"'
    return response


# ============================================================
# IMPORTACIÓN HISTÓRICA - Análisis automático de documentos
# ============================================================

def _analizar_documentos_historicos(sim):
    """
    Analiza documentos existentes de un SIM e identifica patrones de llenado.
    Retorna lista de propuestas por militar.
    """
    propuestas = []

    for pm_sim in sim.pm_sim_set.select_related('pm').all():
        pm = pm_sim.pm

        # Buscar documentos relacionados
        res_primera = Resolucion.objects.filter(
            sim=sim, pm=pm, instancia='PRIMERA'
        ).select_related('abogado').first()

        res_rr = Resolucion.objects.filter(
            sim=sim, pm=pm, instancia='RECONSIDERACION'
        ).first()

        autos_tpe = AUTOTPE.objects.filter(
            sim=sim, pm=pm
        ).select_related('abogado')

        recursos_tsp = RecursoTSP.objects.filter(
            sim=sim, pm=pm
        )

        # Análisis de patrones

        # Patrón 1: Solo RES PRIMERA (sin RR, sin autos)
        if res_primera and not res_rr and not autos_tpe.exists():
            propuesta = {
                'pm': pm,
                'patron': 'RES_SOLO',
                'titulo': 'Solo Resolución PRIMERA (sin apelación)',
                'documentos_encontrados': ['RES (PRIMERA)'],
                'falta': ['RR', 'Autos'],
                'estado_propuesto': 'PROCESO_CONCLUIDO_TPE',
                'fase_propuesta': 'NOTIFICADO_1RA',
                'acciones': 'Cambiar estado y fase (caso concluido sin apelación)'
            }
            propuestas.append(propuesta)

        # Patrón 2: RES PRIMERA + RR (sin autos)
        elif res_primera and res_rr and not autos_tpe.exists():
            propuesta = {
                'pm': pm,
                'patron': 'RES_RR_SOLO',
                'titulo': 'Resolución + Reconsideración (sin Ejecutoria)',
                'documentos_encontrados': ['RES (PRIMERA)', 'RES (RECONSIDERACIÓN)'],
                'falta': ['Autos'],
                'estado_propuesto': 'PROCESO_CONCLUIDO_TPE',
                'fase_propuesta': 'NOTIFICADO_RR',
                'acciones': 'Cambiar estado y fase (caso concluido en 2da instancia)'
            }
            propuestas.append(propuesta)

        # Patrón 3: RES PRIMERA + Autos TPE (sin RR explícito)
        elif res_primera and autos_tpe.exists() and not res_rr:
            auto_ejecutoria = autos_tpe.filter(tipo='AUTO_EJECUTORIA').first()
            if auto_ejecutoria:
                propuesta = {
                    'pm': pm,
                    'patron': 'RES_AUTO_EJECUTORIA',
                    'titulo': 'Resolución + Auto de Ejecutoria (ejecución directa)',
                    'documentos_encontrados': ['RES (PRIMERA)', 'AUTO EJECUTORIA'],
                    'falta': ['RR'],
                    'estado_propuesto': 'PROCESO_CONCLUIDO_TPE',
                    'fase_propuesta': 'MEMORANDUM_RETORNADO' if auto_ejecutoria.memo_numero else 'EJECUTORIA_NOTIFICADA',
                    'acciones': 'Cambiar estado y fase (caso ejecutoriado)'
                }
                propuestas.append(propuesta)

        # Patrón 4: Todos los documentos (RES + RR + Autos)
        elif res_primera and res_rr and autos_tpe.exists():
            propuesta = {
                'pm': pm,
                'patron': 'COMPLETO',
                'titulo': 'Ciclo completo registrado',
                'documentos_encontrados': [
                    'RES (PRIMERA)',
                    'RES (RECONSIDERACIÓN)',
                    f'AUTOS ({autos_tpe.count()})'
                ],
                'falta': [],
                'estado_propuesto': 'PROCESO_CONCLUIDO_TPE',
                'fase_propuesta': 'MEMORANDUM_RETORNADO',
                'acciones': 'Estado y fase ya correctos (revisar si falta algo)'
            }
            propuestas.append(propuesta)

        # Patrón 5: Apelación al TSP
        if recursos_tsp.filter(instancia='APELACION').exists():
            propuesta = {
                'pm': pm,
                'patron': 'APELACION_TSP',
                'titulo': 'Caso en apelación al TSP',
                'documentos_encontrados': ['RAP (APELACIÓN TSP)'],
                'falta': ['Respuesta TSP'],
                'estado_propuesto': 'PROCESO_EN_EL_TSP',
                'fase_propuesta': 'ELEVADO_TSP',
                'acciones': 'Cambiar a estado TSP (caso elevado al Tribunal Supremo)'
            }
            propuestas.append(propuesta)

    return propuestas


@rol_requerido('AYUDANTE')
def ayudante_importar_historico(request):
    """
    Vista de importación histórica con edición de estado, fase y memo.
    GET: Muestra formulario de búsqueda
    POST (búsqueda): Busca SIM, analiza documentos, muestra formulario de edición
    POST (guardar): Guarda cambios de estado, fase y memo
    """
    busqueda_form = BuscarSIMHistoricoForm()
    edicion_form = None
    sim = None
    propuestas = None

    if request.method == 'POST':
        # Detectar si es búsqueda o guardado
        if 'codigo' in request.POST:
            # Búsqueda de SIM
            busqueda_form = BuscarSIMHistoricoForm(request.POST)
            if busqueda_form.is_valid():
                codigo = busqueda_form.cleaned_data['codigo']

                try:
                    sim = SIM.objects.get(codigo=codigo)
                    propuestas = _analizar_documentos_historicos(sim)
                    edicion_form = EditarSIMHistoricoForm(instance=sim)

                except SIM.DoesNotExist:
                    messages.error(request, f'No se encontró SIM con código: {codigo}')

        elif 'estado' in request.POST or 'fase' in request.POST:
            # Guardado de cambios
            # Necesitamos obtener el SIM del request (pasarlo como parámetro o buscarlo por código)
            # Por ahora, buscaremos el código en los datos enviados

            # Obtener el SIM que se está editando
            sim_id = request.POST.get('sim_id')
            if sim_id:
                try:
                    sim = SIM.objects.get(id=sim_id)
                    edicion_form = EditarSIMHistoricoForm(request.POST, instance=sim)

                    if edicion_form.is_valid():
                        # Guardar cambios de estado y fase (sin cambios automáticos)
                        sim = edicion_form.save(commit=False)

                        # Manejar datos de memorándum
                        memo_numero = request.POST.get('memo_numero', '').strip()
                        memo_fecha = request.POST.get('memo_fecha', '').strip()
                        memo_fecha_entrega = request.POST.get('memo_fecha_entrega', '').strip()

                        if any([memo_numero, memo_fecha, memo_fecha_entrega]):
                            # Crear Memorandum ligado a la primera Resolución del sumario (si existe)
                            resolucion = Resolucion.objects.filter(sim=sim).first()

                            if memo_numero and memo_fecha:  # Requerimos al menos número y fecha
                                memorandum, created = Memorandum.objects.get_or_create(
                                    numero=memo_numero,
                                    defaults={
                                        'resolucion': resolucion,
                                        'autotpe': None,
                                        'fecha': memo_fecha,
                                        'fecha_entrega': memo_fecha_entrega or None,
                                    }
                                )

                                if not created:
                                    # Actualizar si ya existe
                                    memorandum.resolucion = resolucion
                                    memorandum.fecha = memo_fecha or memorandum.fecha
                                    memorandum.fecha_entrega = memo_fecha_entrega or memorandum.fecha_entrega
                                    memorandum.save()

                        sim.save()

                        # Mensaje con detalles del cambio
                        msg = f'✅ SIM {sim.codigo} actualizado. Estado: {sim.get_estado_display()}, Fase: {sim.get_fase_display()}'
                        if memo_numero:
                            msg += f' | Memorándum creado: {memo_numero}'
                        messages.success(request, msg)

                        # Mostrar de nuevo el formulario con los cambios
                        propuestas = _analizar_documentos_historicos(sim)
                        edicion_form = EditarSIMHistoricoForm(instance=sim)
                    else:
                        messages.error(request, 'Error al guardar los cambios. Verifica los datos.')

                except SIM.DoesNotExist:
                    messages.error(request, 'No se encontró el sumario a editar.')

    # Si es GET o no hay SIM seleccionado, mostrar solo formulario de búsqueda
    if not sim:
        edicion_form = None

    context = {
        'form': busqueda_form,
        'edicion_form': edicion_form,
        'sim': sim,
        'propuestas': propuestas or [],
        'total_propuestas': len(propuestas) if propuestas else 0,
    }

    return render(request, 'tpe_app/ayudante/importar_historico.html', context)
