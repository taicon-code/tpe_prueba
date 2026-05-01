# ============================================================
#  VISTAS DE EXPORTACIÓN - PDF (ZIP) y EXCEL
#  Archivo: tpe_app/views/export_views.py
# ============================================================

from io import BytesIO
from datetime import datetime, date
from zipfile import ZipFile
from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak, HRFlowable
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT, TA_JUSTIFY
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from django.contrib.auth.decorators import login_required
from tpe_app.models import PM, SIM, AUTOTPE, AUTOTSP, Resolucion, RecursoTSP, PerfilUsuario


def _format_date(date_obj):
    """Formatea una fecha como DD/MM/YYYY"""
    if not date_obj:
        return "N/A"
    return date_obj.strftime("%d/%m/%Y")


def _sanitize_filename(text):
    """Elimina caracteres inválidos para nombres de archivo"""
    if not text:
        return "SIN_NOMBRE"
    return "".join(c if c.isalnum() or c in "-_" else "" for c in text.upper())


def _compilar_documentos(sim, historial, pm=None):
    """Compila lista coordinada de documentos (RES, RR, AUTOTPE, RAP, RAEE, AUTOTSP) con columnas estándar.
    Si pm se indica, filtra solo los documentos de ese militar (AUTOTSP se excluye pues no tiene FK a pm).
    Retorna: (documentos, info_tsp) donde info_tsp es dict con numero_oficio y fecha_oficio si hay RAP
    """
    documentos = []
    info_tsp = None

    def _filt(qs):
        qs = qs.filter(sim=sim)
        if pm is not None:
            qs = qs.filter(pm=pm)
        return qs

    # Resoluciones (PRIMERA instancia)
    for res in _filt(historial['resoluciones']):
        notif_info = None
        _notif = getattr(res, 'notificacion', None)
        if _notif:
            notif_info = {
                'tipo': _notif.get_tipo_display(),
                'fecha': _notif.fecha,
            }
        documentos.append({
            'tipo': 'RESOLUCIÓN',
            'numero': res.numero or 'S/N',
            'fecha_doc': res.fecha,
            'resolutiva': (res.texto or 'N/A').upper(),
            'notificacion': notif_info,
            'memo': None,
        })

    # Recurso de Reconsideración (RR)
    for rr in _filt(historial.get('segundas_resoluciones', Resolucion.objects.none())):
        notif_info = None
        _notif = getattr(rr, 'notificacion', None)
        if _notif:
            notif_info = {
                'tipo': _notif.get_tipo_display(),
                'fecha': _notif.fecha,
            }
        documentos.append({
            'tipo': 'REC. RECONSIDERACIÓN',
            'numero': rr.numero or 'S/N',
            'fecha_doc': rr.fecha,
            'resolutiva': (rr.texto or 'N/A').upper(),
            'notificacion': notif_info,
            'memo': None,
        })

    # Auto TPE
    for auto in _filt(historial['autos_tpe']):
        memo_info = None
        memo = getattr(auto, 'memorandum', None)
        if memo:
            memo_info = {
                'numero': memo.numero or 'S/N',
                'fecha': memo.fecha,
                'fecha_recepcion': memo.fecha_entrega,
            }
        notif_info = None
        _notif = getattr(auto, 'notificacion', None)
        if _notif:
            notif_info = {
                'tipo': _notif.get_tipo_display(),
                'fecha': _notif.fecha,
            }
        documentos.append({
            'tipo': 'AUTO TPE',
            'numero': auto.numero or 'S/N',
            'fecha_doc': auto.fecha,
            'resolutiva': (auto.texto or (auto.get_tipo_display() if auto.tipo else 'N/A')).upper(),
            'notificacion': notif_info,
            'memo': memo_info,
        })

    # Recurso de Apelación (RAP) - NO se incluye en tabla (no es documento físico del TPE)
    # Solo se extrae la información para mostrar estado en TSP
    for rap in _filt(historial['recursos_apelacion']):
        if not info_tsp and rap.numero_oficio:
            info_tsp = {
                'numero_oficio': rap.numero_oficio,
                'fecha_oficio': rap.fecha_oficio,
            }

    # RAEE
    for raee in _filt(historial['raees']):
        notif_info = None
        _notif = getattr(raee, 'notificacion', None)
        if _notif:
            notif_info = {
                'tipo': _notif.get_tipo_display(),
                'fecha': _notif.fecha,
            }
        documentos.append({
            'tipo': 'REC. ACLARACIÓN Y ENMIENDA (RAEE)',
            'numero': raee.numero or 'S/N',
            'fecha_doc': raee.fecha,
            'resolutiva': (raee.get_instancia_display() if raee.instancia else 'ACLARACIÓN Y ENMIENDA').upper(),
            'notificacion': notif_info,
            'memo': None,
        })

    # Auto TSP — sin FK a pm, solo se incluye cuando no hay filtro por militar
    if pm is None:
        for autotsp in historial['autos_tsp'].filter(sim=sim):
            notif_info = None
            _notif = getattr(autotsp, 'notificacion', None)
            if _notif:
                notif_info = {
                    'tipo': _notif.get_tipo_display(),
                    'fecha': _notif.fecha,
                }
            documentos.append({
                'tipo': 'AUTO TSP',
                'numero': autotsp.numero or 'S/N',
                'fecha_doc': autotsp.fecha,
                'resolutiva': (autotsp.texto or (autotsp.get_tipo_display() if autotsp.tipo else 'N/A')).upper(),
                'notificacion': notif_info,
                'memo': None,
            })

    documentos.sort(key=lambda x: x['fecha_doc'] or date.min)
    return documentos, info_tsp


def _obtener_historial(personal_id):
    """Obtiene historial completo de una persona"""
    try:
        personal = PM.objects.get(id=personal_id)
    except PM.DoesNotExist:
        return None, None

    sims = SIM.objects.filter(militares__id=personal_id).distinct()
    sim_ids = list(sims.values_list('id', flat=True))

    historial = {
        'personal': personal,
        'sumarios': sims,
        'resoluciones': Resolucion.objects.filter(sim__in=sim_ids, instancia='PRIMERA', pm=personal),
        'segundas_resoluciones': Resolucion.objects.filter(sim__in=sim_ids, instancia='RECONSIDERACION', pm=personal),
        'recursos_apelacion': RecursoTSP.objects.filter(sim__in=sim_ids, instancia='APELACION', pm=personal),
        'raees': RecursoTSP.objects.filter(sim__in=sim_ids, instancia='ACLARACION_ENMIENDA', pm=personal),
        'autos_tpe': AUTOTPE.objects.filter(sim__in=sim_ids, pm=personal),
        'autos_tsp': AUTOTSP.objects.filter(sim__in=sim_ids),
    }

    return personal, historial


@login_required
def export_person_historial_pdf(request, personal_id):
    """Genera PDF formal del historial disciplinario de un militar."""
    personal, historial = _obtener_historial(personal_id)
    if not personal or not historial:
        get_object_or_404(PM, id=personal_id)
        return HttpResponse("Personal no encontrado", status=404)

    # Registrar fuente Arial desde Windows
    try:
        pdfmetrics.getFont('Arial-Bold')
    except KeyError:
        try:
            pdfmetrics.registerFont(TTFont('Arial',      'C:/Windows/Fonts/arial.ttf'))
            pdfmetrics.registerFont(TTFont('Arial-Bold', 'C:/Windows/Fonts/arialbd.ttf'))
        except Exception:
            # Si las fuentes no existen, usar Helvetica por defecto
            pass

    buffer = BytesIO()
    page_w, _ = letter
    margin = 0.65 * inch


    # Estilos de párrafo
    def _estilo(nombre, fuente='Helvetica', tamaño=9, alineacion=TA_LEFT,
                negrita=False, color=colors.black, espacio_antes=0, espacio_despues=2,
                interlinea=11, sangria=0):
        fn = (fuente + '-Bold') if negrita else fuente
        return ParagraphStyle(nombre, fontName=fn, fontSize=tamaño,
                              alignment=alineacion, textColor=color,
                              spaceBefore=espacio_antes, spaceAfter=espacio_despues,
                              leading=interlinea, leftIndent=sangria)

    s_inst1   = _estilo('inst1',  fuente='Arial', tamaño=10, alineacion=TA_LEFT, negrita=True, espacio_despues=1, interlinea=9)
    s_inst2   = _estilo('inst2',  fuente='Arial', tamaño=10, alineacion=TA_LEFT, negrita=True, espacio_despues=1, interlinea=9, sangria=15)
    s_pais    = _estilo('pais',   fuente='Arial', tamaño=10, alineacion=TA_LEFT, negrita=True, espacio_despues=6, interlinea=13, sangria=70)
    s_seccion = _estilo('secc',   fuente='Arial', tamaño=14, alineacion=TA_CENTER, negrita=True, espacio_antes=6, espacio_despues=4, interlinea=17)
    s_dato    = _estilo('dato',   tamaño=8.5)
    s_objeto  = _estilo('obj',    tamaño=8,  alineacion=TA_JUSTIFY, interlinea=10)
    s_sim_tit = _estilo('simtit', tamaño=9,  negrita=True, color=colors.black, espacio_antes=4, espacio_despues=2)
    s_th      = _estilo('th',     tamaño=7.5, alineacion=TA_CENTER, negrita=True, color=colors.black)
    s_td      = _estilo('td',     tamaño=7,  interlinea=9)
    s_td_c    = _estilo('tdc',    tamaño=7,  alineacion=TA_CENTER, interlinea=9)
    s_stat_h  = _estilo('sth',    tamaño=8,  alineacion=TA_CENTER, negrita=True, color=colors.black)
    s_stat_v  = _estilo('stv',    tamaño=18, alineacion=TA_CENTER, negrita=True, color=colors.black, interlinea=22)
    s_memo    = _estilo('memo',   tamaño=6.5, interlinea=8, color=colors.Color(0.35, 0.35, 0.35), sangria=8)


    # Obtener datos del usuario que imprime
    grado_usuario = ""
    arma_usuario = ""
    especialidad_usuario = ""
    nombre_completo = request.user.get_full_name() or request.user.username
    if request.user.is_authenticated:
        try:
            perfil = request.user.perfilusuario
            pm_pie = perfil.pm if perfil.pm else (perfil.vocal.pm if perfil.vocal and perfil.vocal.pm else None)
            if pm_pie:
                grado_usuario  = pm_pie.get_grado_display() or ""
                arma_usuario  = pm_pie.get_arma_display() or ""
                especialidad_usuario = pm_pie.especialidad or ""
                nombre_completo = f"{pm_pie.nombre or ''} {pm_pie.paterno or ''} {pm_pie.materno or ''}".strip()
        except Exception:
            pass

    partes_pie = [p for p in [grado_usuario, arma_usuario, especialidad_usuario, nombre_completo.upper()] if p]
    texto_impreso = "  ".join(partes_pie)

    fecha_hoy = datetime.now().strftime("%d/%m/%Y")
    hora_hoy  = datetime.now().strftime("%H:%M")

    def _pie_pagina(canv, doc):
        canv.saveState()
        canv.setStrokeColor(colors.lightgrey)
        canv.setLineWidth(0.5)
        canv.line(margin, 0.52 * inch, page_w - margin, 0.52 * inch)
        canv.setFont('Helvetica', 6.5)
        canv.setFillColor(colors.grey)
        texto_pie = (f"Impreso por: {texto_impreso}   |   "
                     f"{fecha_hoy}  {hora_hoy}   |   Pág. {doc.page}")
        canv.drawCentredString(page_w / 2, 0.33 * inch, texto_pie)
        canv.restoreState()

    doc = SimpleDocTemplate(
        buffer, pagesize=letter,
        leftMargin=margin, rightMargin=margin,
        topMargin=0.55 * inch, bottomMargin=0.75 * inch,
    )
    usable_w = page_w - 2 * margin
    story = []

    # ── ENCABEZADO INSTITUCIONAL ─────────────────────────────────────────────
    story.append(Paragraph("COMANDO GENERAL DEL EJÉRCITO", s_inst1))
    story.append(Paragraph("DEPARTAMENTO I - PERSONAL", s_inst2))
    story.append(Paragraph("<u>BOLIVIA</u>", s_pais))

    # ── DATOS DEL PERSONAL ───────────────────────────────────────────────────
    story.append(Paragraph("<u>DATOS DEL PERSONAL</u>", s_seccion))
    story.append(Spacer(1, 4))

    grado        = personal.get_grado_display() or 'N/A'
    especialidad = personal.get_arma_display() or 'N/A'
    nombre_comp  = f"{personal.nombre} {personal.paterno} {personal.materno or ''}".strip()
    ci           = str(personal.ci) if personal.ci else 'N/A'
    escalafon    = personal.get_escalafon_display() or 'N/A'
    estado_pm    = personal.get_estado_display()
    anio_egreso  = str(personal.anio_promocion) if personal.anio_promocion else 'N/A'
    años_serv    = personal.años_servicio
    grado_esp    = personal.grado_esperado
    no_asc_txt   = '  ⚠ No ascendió al grado correspondiente' if personal.no_ascendio else ''

    grado_line = grado
    if grado_esp and grado_esp != grado:
        grado_line = f"{grado}  (esperado: {grado_esp}{no_asc_txt})"
    elif personal.no_ascendio:
        grado_line = f"{grado}{no_asc_txt}"

    col2 = usable_w / 2
    tabla_datos = Table([
        [Paragraph(f"<b>Grado:</b>  {grado_line}",    s_dato), Paragraph(f"<b>Especialidad:</b>  {especialidad}", s_dato)],
        [Paragraph(f"<b>Nombre:</b>  {nombre_comp}",  s_dato), ''],
        [Paragraph(f"<b>C.I.:</b>  {ci}",             s_dato), Paragraph(f"<b>Escalafón:</b>  {escalafon}",     s_dato)],
        [Paragraph(f"<b>Estado:</b>  {estado_pm}",    s_dato), Paragraph(f"<b>Año Egreso:</b>  {anio_egreso}  ({años_serv} años de servicio)" if años_serv else f"<b>Año Egreso:</b>  {anio_egreso}", s_dato)],
    ], colWidths=[col2, col2])
    tabla_datos.setStyle(TableStyle([
        ('SPAN',          (0, 1), (1, 1)),
        ('TOPPADDING',    (0, 0), (-1, -1), 3),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
        ('LEFTPADDING',   (0, 0), (-1, -1), 4),
        ('LINEBELOW',     (0, -1), (-1, -1), 0.5, colors.lightgrey),
    ]))
    story.append(tabla_datos)
    story.append(Spacer(1, 10))

    # ── RESUMEN ESTADÍSTICO ──────────────────────────────────────────────────
    story.append(Paragraph("<u>RESUMEN ESTADÍSTICO</u>", s_seccion))
    story.append(Spacer(1, 4))

    n_sim  = historial['sumarios'].count()
    n_res  = (historial['resoluciones'].count() +
              historial['segundas_resoluciones'].count())
    n_auto = historial['autos_tpe'].count()
    col3 = usable_w / 3
    tabla_stats = Table([
        [Paragraph('SUMARIOS', s_stat_h), Paragraph('RESOLUCIONES', s_stat_h), Paragraph('AUTOS TPE', s_stat_h)],
        [Paragraph(str(n_sim),  s_stat_v), Paragraph(str(n_res),   s_stat_v), Paragraph(str(n_auto), s_stat_v)],
    ], colWidths=[col3, col3, col3])
    tabla_stats.setStyle(TableStyle([
        ('BOX',           (0, 0), (-1, -1), 0.8, colors.black),
        ('INNERGRID',     (0, 0), (-1, -1), 0.5, colors.black),
        ('ALIGN',         (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN',        (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING',    (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
    ]))
    story.append(tabla_stats)
    story.append(Spacer(1, 10))

    # ── ACTUADOS POR SUMARIO ─────────────────────────────────────────────────
    story.append(Paragraph("<u>ACTUADOS POR SUMARIO</u>", s_seccion))
    story.append(Spacer(1, 6))

    sumarios = historial['sumarios']
    if not sumarios.exists():
        story.append(Paragraph("No se registran sumarios para este personal.", s_dato))
    else:
        # Anchos de columna: TIPO 20% | N° 9% | FECHA 12% | RESOLUTIVA 44% | NOTIF. 15%
        cw = [usable_w * p for p in (0.20, 0.09, 0.12, 0.44, 0.15)]

        for idx, sim in enumerate(sumarios, 1):
            story.append(Paragraph(f"SUMARIO N.° {sim.codigo}", s_sim_tit))

            objeto = sim.objeto or 'N/A'
            story.append(Paragraph(f"<b>Objeto:</b> {objeto}", s_objeto))

            estado_display = (sim.get_estado_display()
                              if hasattr(sim, 'get_estado_display') else sim.estado)
            story.append(Paragraph(f"<b>Estado:</b> {estado_display}", s_dato))

            # Grado que tenía el militar cuando se tramitó este sumario
            from tpe_app.models import PM_SIM as PM_SIM_model
            pm_sim_reg = PM_SIM_model.objects.filter(sim=sim, pm=personal).first()
            if pm_sim_reg and pm_sim_reg.grado_en_fecha:
                story.append(Paragraph(
                    f"<b>Grado al momento del sumario:</b> {pm_sim_reg.grado_en_fecha}",
                    s_dato
                ))
            story.append(Spacer(1, 5))

            # Recopilar actuados con función coordinada — filtrar solo documentos de este militar
            documentos, info_tsp = _compilar_documentos(sim, historial, pm=personal)

            if documentos:
                headers = ['TIPO DE DOCUMENTO', 'N°', 'FECHA', 'RESOLUTIVA', 'NOTIF.']
                filas = [[Paragraph(h, s_th) for h in headers]]
                ts = [
                    ('BOX',           (0, 0), (-1, -1), 0.5, colors.black),
                    ('INNERGRID',     (0, 0), (-1, -1), 0.3, colors.black),
                    ('LINEBELOW',     (0, 0), (-1, 0), 0.8, colors.black),
                    ('VALIGN',        (0, 0), (-1, -1), 'MIDDLE'),
                    ('TOPPADDING',    (0, 0), (-1, -1), 4),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
                    ('LEFTPADDING',   (0, 0), (-1, -1), 4),
                    ('RIGHTPADDING',  (0, 0), (-1, -1), 4),
                ]
                for actuado in documentos:
                    notif_str = 'NO'
                    if actuado['notificacion']:
                        notif_tipo = actuado['notificacion']['tipo']
                        notif_fecha = _format_date(actuado['notificacion']['fecha'])
                        notif_str = f"{notif_tipo}\n({notif_fecha})"
                    filas.append([
                        Paragraph(actuado['tipo'],                    s_td),
                        Paragraph(str(actuado['numero']),             s_td_c),
                        Paragraph(_format_date(actuado['fecha_doc']), s_td_c),
                        Paragraph(actuado['resolutiva'],              s_td),
                        Paragraph(notif_str,                          s_td_c),
                    ])
                    if actuado['memo']:
                        memo = actuado['memo']
                        entrega = _format_date(memo['fecha_recepcion']) if memo['fecha_recepcion'] else 'PENDIENTE'
                        memo_txt = f"<i>Memorándum N° {memo['numero']}  —  Fecha: {_format_date(memo['fecha'])}  —  Entrega: {entrega}</i>"
                        ri = len(filas)
                        filas.append([Paragraph(memo_txt, s_memo), '', '', '', ''])
                        ts += [
                            ('SPAN',          (0, ri), (-1, ri)),
                            ('TOPPADDING',    (0, ri), (-1, ri), 2),
                            ('BOTTOMPADDING', (0, ri), (-1, ri), 3),
                            ('LEFTPADDING',   (0, ri), (-1, ri), 12),
                        ]

                t = Table(filas, colWidths=cw, repeatRows=1)
                t.setStyle(TableStyle(ts))
                story.append(t)
            else:
                story.append(Paragraph("Sin actuados registrados.", s_dato))

            # Mostrar estado en TSP si aplica (DESPUÉS de la tabla)
            if info_tsp:
                story.append(Spacer(1, 8))
                story.append(Paragraph(
                    f"<b>PROCESO EN EL TRIBUNAL SUPERIOR DEL PERSONAL DE LAS FF.AA.</b><br/>Oficio Nº {info_tsp['numero_oficio']} de {_format_date(info_tsp['fecha_oficio'])}",
                    s_dato
                ))
                story.append(Spacer(1, 8))
            else:
                story.append(Spacer(1, 8))
            if idx < sumarios.count():
                story.append(HRFlowable(width="100%", thickness=0.4, color=colors.lightgrey))
                story.append(Spacer(1, 4))

    doc.build(story, onFirstPage=_pie_pagina, onLaterPages=_pie_pagina)
    buffer.seek(0)

    fecha_export = datetime.now().strftime("%d-%m-%Y")
    filename = f"HISTORIAL_{personal.ci}_{fecha_export}.pdf"
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


@login_required
def export_person_pdfs_zip(request, personal_id):
    """
    DEPRECATED: Mantener por compatibilidad. Redirige a export_person_historial_pdf.
    """
    return export_person_historial_pdf(request, personal_id)




@login_required
def export_person_excel(request, personal_id):
    """
    Genera un Excel con 4 hojas: Personal, Sumarios, Resoluciones, Cronología.
    """
    personal, historial = _obtener_historial(personal_id)

    if not personal or not historial:
        personal = get_object_or_404(PM, id=personal_id)
        return HttpResponse("Personal no encontrado", status=404)

    if not historial:
        return HttpResponse("Personal no encontrado", status=404)

    # Crear workbook
    wb = Workbook()
    wb.remove(wb.active)  # Eliminar hoja por defecto

    # ===== HOJA 1: PERSONAL =====
    ws_personal = wb.create_sheet("PERSONAL")
    ws_personal['A1'] = "INFORMACIÓN PERSONAL"
    ws_personal['A1'].font = Font(bold=True, size=12)

    personal_data = [
        ("CI:", str(personal.ci)),
        ("Nombre:", f"{personal.nombre} {personal.paterno} {personal.materno or ''}"),
        ("Grado:", personal.get_grado_display() or "N/A"),
        ("Escalafón:", personal.get_escalafon_display() or "N/A"),
        ("Arma:", personal.get_arma_display() or "N/A"),
        ("Especialidad:", personal.especialidad or "N/A"),
        ("Estado:", personal.get_estado_display()),
        ("Fecha Promoción:", _format_date(personal.anio_promocion)),
    ]

    for idx, (label, value) in enumerate(personal_data, start=2):
        ws_personal[f'A{idx}'] = label
        ws_personal[f'A{idx}'].font = Font(bold=True)
        ws_personal[f'B{idx}'] = value

    ws_personal.column_dimensions['A'].width = 20
    ws_personal.column_dimensions['B'].width = 40

    # ===== HOJA 2: SUMARIOS =====
    ws_sumarios = wb.create_sheet("SUMARIOS")
    headers_sim = ["DJE", "Tipo", "Objeto", "Resumen", "Fecha Ingreso", "Estado"]
    for col, header in enumerate(headers_sim, 1):
        cell = ws_sumarios.cell(row=1, column=col)
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="185FA5", end_color="185FA5", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")

    for row_idx, sim in enumerate(historial['sumarios'], 2):
        ws_sumarios.cell(row=row_idx, column=1, value=sim.codigo)
        ws_sumarios.cell(row=row_idx, column=2, value=sim.get_tipo_display() if sim.tipo else "N/A")
        ws_sumarios.cell(row=row_idx, column=3, value=sim.objeto[:50] + "..." if len(sim.objeto or '') > 50 else sim.objeto)
        ws_sumarios.cell(row=row_idx, column=4, value=sim.resumen or "N/A")
        ws_sumarios.cell(row=row_idx, column=5, value=_format_date(sim.fecha_ingreso))
        ws_sumarios.cell(row=row_idx, column=6, value=sim.estado)

    for col in range(1, len(headers_sim) + 1):
        ws_sumarios.column_dimensions[chr(64 + col)].width = 25

    # ===== HOJA 3: DOCUMENTOS (Coordinado) =====
    ws_docs = wb.create_sheet("DOCUMENTOS")
    headers_docs = ["SIM", "DOCUMENTO", "NÚMERO", "FECHA", "RESOLUTIVA", "NOTIFICACIÓN", "MEMO N°", "MEMO FECHA", "MEMO ENTREGA", "ESTADO TSP", "OFICIO N°", "FECHA OFICIO"]
    for col, header in enumerate(headers_docs, 1):
        cell = ws_docs.cell(row=1, column=col)
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="185FA5", end_color="185FA5", fill_type="solid")

    row_idx = 2
    # Compilar todos los documentos por SIM
    for sim in historial['sumarios']:
        hist_sim = {
            'resoluciones': historial['resoluciones'].filter(sim=sim),
            'segundas_resoluciones': historial['segundas_resoluciones'].filter(sim=sim),
            'autos_tpe': historial['autos_tpe'].filter(sim=sim),
            'recursos_apelacion': historial['recursos_apelacion'].filter(sim=sim),
            'raees': historial['raees'].filter(sim=sim),
            'autos_tsp': historial['autos_tsp'].filter(sim=sim),
        }
        documentos, info_tsp = _compilar_documentos(sim, hist_sim)

        for doc in documentos:
            ws_docs.cell(row=row_idx, column=1, value=sim.codigo)
            ws_docs.cell(row=row_idx, column=2, value=doc['tipo'])
            ws_docs.cell(row=row_idx, column=3, value=doc['numero'])
            ws_docs.cell(row=row_idx, column=4, value=_format_date(doc['fecha_doc']))
            ws_docs.cell(row=row_idx, column=5, value=doc['resolutiva'])
            notif = doc['notificacion']
            notif_str = f"{notif['tipo']} ({_format_date(notif['fecha'])})" if notif else ""
            ws_docs.cell(row=row_idx, column=6, value=notif_str)
            if doc['memo']:
                ws_docs.cell(row=row_idx, column=7, value=doc['memo']['numero'])
                ws_docs.cell(row=row_idx, column=8, value=_format_date(doc['memo']['fecha']))
                ws_docs.cell(row=row_idx, column=9, value=_format_date(doc['memo']['fecha_recepcion']))
            row_idx += 1

        # Agregar fila de estado TSP si hay
        if info_tsp:
            ws_docs.cell(row=row_idx, column=1, value=sim.codigo)
            ws_docs.cell(row=row_idx, column=10, value="EN TRIBUNAL SUPREMO POLICIAL")
            ws_docs.cell(row=row_idx, column=11, value=info_tsp['numero_oficio'])
            ws_docs.cell(row=row_idx, column=12, value=_format_date(info_tsp['fecha_oficio']))
            # Aplicar negrita para destacar
            for col_num in [10, 11, 12]:
                ws_docs.cell(row=row_idx, column=col_num).font = Font(bold=True)
            row_idx += 1

    # Ajustar anchos de columna
    ws_docs.column_dimensions['A'].width = 12  # SIM
    ws_docs.column_dimensions['B'].width = 20  # DOCUMENTO
    ws_docs.column_dimensions['C'].width = 12  # NÚMERO
    ws_docs.column_dimensions['D'].width = 14  # FECHA
    ws_docs.column_dimensions['E'].width = 30  # RESOLUTIVA
    ws_docs.column_dimensions['F'].width = 14  # NOTIFICACIÓN
    ws_docs.column_dimensions['G'].width = 12  # MEMO N°
    ws_docs.column_dimensions['H'].width = 14  # MEMO FECHA
    ws_docs.column_dimensions['I'].width = 14  # MEMO ENTREGA
    ws_docs.column_dimensions['J'].width = 28  # ESTADO TSP
    ws_docs.column_dimensions['K'].width = 14  # OFICIO N°
    ws_docs.column_dimensions['L'].width = 14  # FECHA OFICIO

    # ===== HOJA 4: CRONOLOGÍA =====
    ws_crono = wb.create_sheet("CRONOLOGÍA")
    headers_crono = ["Fecha", "Tipo Evento", "Descripción"]
    for col, header in enumerate(headers_crono, 1):
        cell = ws_crono.cell(row=1, column=col)
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="185FA5", end_color="185FA5", fill_type="solid")

    # Crear timeline ordenada
    eventos = []
    for sim in historial['sumarios']:
        eventos.append((sim.fecha_ingreso, "Sumario Ingreso", f"SIM {sim.codigo}: {sim.resumen}"))

    for res in historial['resoluciones']:
        eventos.append((res.fecha, "Resolución TPE", f"RES {res.numero}: {res.get_tipo_display()}"))

    for rap in historial['recursos_apelacion']:
        eventos.append((rap.fecha, "Apelación TSP", f"RAP {rap.numero or 'Pendiente'}: {rap.tipo or 'N/A'}"))

    # Ordenar por fecha
    eventos = sorted([e for e in eventos if e[0]], key=lambda x: x[0])

    for row_idx, (fecha, tipo, desc) in enumerate(eventos, 2):
        ws_crono.cell(row=row_idx, column=1, value=_format_date(fecha))
        ws_crono.cell(row=row_idx, column=2, value=tipo)
        ws_crono.cell(row=row_idx, column=3, value=desc)

    for col in range(1, len(headers_crono) + 1):
        ws_crono.column_dimensions[chr(64 + col)].width = 25

    # Guardar
    fecha_export = datetime.now().strftime("%Y-%m-%d")
    excel_filename = f"HISTORIAL_{personal.ci}_{fecha_export}.xlsx"

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(buffer, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{excel_filename}"'
    return response


# ============================================================
# EXPORTACIÓN DE SIM COMPLETO
# ============================================================

@login_required
def export_sim_pdf(request, sim_id):
    """Exporta un SIM completo a PDF con militares y actuados — Formato Platypus moderno"""
    sim = get_object_or_404(SIM, id=sim_id)
    _orden_grado = {g: i for i, (g, _) in enumerate(PM.GRADO_CHOICES)}
    militares = sorted(
        sim.militares.all(),
        key=lambda m: _orden_grado.get(m.grado, 999)
    )
    resoluciones = Resolucion.objects.filter(sim=sim)
    autos_tpe = AUTOTPE.objects.filter(sim=sim)
    recursos_tsp = RecursoTSP.objects.filter(sim=sim)

    # Registrar fuente Arial desde Windows
    try:
        pdfmetrics.getFont('Arial-Bold')
    except KeyError:
        try:
            pdfmetrics.registerFont(TTFont('Arial',      'C:/Windows/Fonts/arial.ttf'))
            pdfmetrics.registerFont(TTFont('Arial-Bold', 'C:/Windows/Fonts/arialbd.ttf'))
        except Exception:
            # Si las fuentes no existen, usar Helvetica por defecto
            pass

    buffer = BytesIO()
    page_w, _ = letter
    margin = 0.65 * inch

    # Estilos de párrafo
    def _estilo(nombre, fuente='Helvetica', tamaño=9, alineacion=TA_LEFT,
                negrita=False, color=colors.black, espacio_antes=0, espacio_despues=2,
                interlinea=11, sangria=0):
        fn = (fuente + '-Bold') if negrita else fuente
        return ParagraphStyle(nombre, fontName=fn, fontSize=tamaño,
                              alignment=alineacion, textColor=color,
                              spaceBefore=espacio_antes, spaceAfter=espacio_despues,
                              leading=interlinea, leftIndent=sangria)

    s_inst1   = _estilo('inst1',  fuente='Arial', tamaño=10, alineacion=TA_LEFT, negrita=True, espacio_despues=1, interlinea=9)
    s_inst2   = _estilo('inst2',  fuente='Arial', tamaño=10, alineacion=TA_LEFT, negrita=True, espacio_despues=1, interlinea=9, sangria=15)
    s_pais    = _estilo('pais',   fuente='Arial', tamaño=10, alineacion=TA_LEFT, negrita=True, espacio_despues=6, interlinea=13, sangria=70)
    s_seccion = _estilo('secc',   fuente='Arial', tamaño=14, alineacion=TA_CENTER, negrita=True, espacio_antes=6, espacio_despues=4, interlinea=17)
    s_dato    = _estilo('dato',   tamaño=8.5)
    s_dato_j  = _estilo('datoj',  tamaño=8.5, alineacion=TA_JUSTIFY)
    s_sim_tit = _estilo('simtit', tamaño=10, negrita=True, espacio_antes=4, espacio_despues=2)
    s_th      = _estilo('th',     tamaño=7.5, alineacion=TA_CENTER, negrita=True, color=colors.black)
    s_td      = _estilo('td',     tamaño=7,  interlinea=9)
    s_td_c    = _estilo('tdc',    tamaño=7,  alineacion=TA_CENTER, interlinea=9)
    s_memo    = _estilo('memo2',  tamaño=6.5, interlinea=8, color=colors.Color(0.35, 0.35, 0.35), sangria=8)
    s_pm_sub  = _estilo('pmsub',  tamaño=8.5, negrita=True, espacio_antes=6, espacio_despues=3, color=colors.black)

    # Obtener datos del usuario que imprime
    grado_usuario = ""
    arma_usuario = ""
    especialidad_usuario = ""
    nombre_completo = request.user.get_full_name() or request.user.username
    if request.user.is_authenticated:
        try:
            perfil = request.user.perfilusuario
            pm_pie = perfil.pm if perfil.pm else (perfil.vocal.pm if perfil.vocal and perfil.vocal.pm else None)
            if pm_pie:
                grado_usuario  = pm_pie.get_grado_display() or ""
                arma_usuario  = pm_pie.get_arma_display() or ""
                especialidad_usuario = pm_pie.especialidad or ""
                nombre_completo = f"{pm_pie.nombre or ''} {pm_pie.paterno or ''} {pm_pie.materno or ''}".strip()
        except Exception:
            pass

    partes_pie = [p for p in [grado_usuario, arma_usuario, especialidad_usuario, nombre_completo.upper()] if p]
    texto_impreso = "  ".join(partes_pie)

    fecha_hoy = datetime.now().strftime("%d/%m/%Y")
    hora_hoy  = datetime.now().strftime("%H:%M")

    def _pie_pagina(canv, doc):
        canv.saveState()
        canv.setStrokeColor(colors.lightgrey)
        canv.setLineWidth(0.5)
        canv.line(margin, 0.52 * inch, page_w - margin, 0.52 * inch)
        canv.setFont('Helvetica', 6.5)
        canv.setFillColor(colors.grey)
        texto_pie = (f"Impreso por: {texto_impreso}   |   "
                     f"{fecha_hoy}  {hora_hoy}   |   Pág. {doc.page}")
        canv.drawCentredString(page_w / 2, 0.33 * inch, texto_pie)
        canv.restoreState()

    doc = SimpleDocTemplate(
        buffer, pagesize=letter,
        leftMargin=margin, rightMargin=margin,
        topMargin=0.55 * inch, bottomMargin=0.75 * inch,
    )
    usable_w = page_w - 2 * margin
    story = []

    # ── ENCABEZADO INSTITUCIONAL ─────────────────────────────────────────────
    story.append(Paragraph("COMANDO GENERAL DEL EJÉRCITO", s_inst1))
    story.append(Paragraph("DEPARTAMENTO I - PERSONAL", s_inst2))
    story.append(Paragraph("<u>BOLIVIA</u>", s_pais))

    # ── DATOS DEL SUMARIO ───────────────────────────────────────────────────
    story.append(Paragraph(f"<u>REPORTE DE SUMARIO N.° {sim.codigo}</u>", s_seccion))
    story.append(Spacer(1, 4))

    col2 = usable_w / 2
    tabla_sim = Table([
        [Paragraph(f"<b>Código:</b>  {sim.codigo}", s_dato), Paragraph(f"<b>Tipo:</b>  {sim.get_tipo_display()}", s_dato)],
        [Paragraph(f"<b>Objeto:</b>  {sim.objeto or 'N/A'}", s_dato_j), ''],
        [Paragraph(f"<b>Resumen:</b>  {sim.resumen or 'N/A'}", s_dato), ''],
        [Paragraph(f"<b>Fecha Ingreso:</b>  {_format_date(sim.fecha_ingreso)}", s_dato), Paragraph(f"<b>Estado:</b>  {sim.get_estado_display()}", s_dato)],
    ], colWidths=[col2, col2])
    tabla_sim.setStyle(TableStyle([
        ('SPAN',          (0, 1), (1, 1)),
        ('SPAN',          (0, 2), (1, 2)),
        ('TOPPADDING',    (0, 0), (-1, -1), 3),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
        ('LEFTPADDING',   (0, 0), (-1, -1), 4),
        ('LINEBELOW',     (0, -1), (-1, -1), 0.5, colors.lightgrey),
    ]))
    story.append(tabla_sim)
    story.append(Spacer(1, 10))

    # ── MILITARES INVESTIGADOS ──────────────────────────────────────────────
    if militares:
        story.append(Paragraph("<u>MILITARES INVESTIGADOS</u>", s_seccion))
        story.append(Spacer(1, 4))

        col_militares = [usable_w * 0.16, usable_w * 0.18, usable_w * 0.24, usable_w * 0.24, usable_w * 0.18]

        filas_mil = [[
            Paragraph('GRADO', s_th),
            Paragraph('NOMBRE', s_th),
            Paragraph('AP. PATERNO', s_th),
            Paragraph('AP. MATERNO', s_th),
            Paragraph('C.I.', s_th),
        ]]

        for militar in militares:
            grado_txt = militar.get_grado_display() or 'N/A'
            if militar.especialidad:
                grado_txt += f'\n{militar.especialidad}'
            filas_mil.append([
                Paragraph(grado_txt, s_td_c),
                Paragraph(militar.nombre or 'N/A', s_td),
                Paragraph(militar.paterno or 'N/A', s_td),
                Paragraph(militar.materno or 'N/A', s_td),
                Paragraph(str(militar.ci) if militar.ci else 'N/A', s_td_c),
            ])

        tabla_mil = Table(filas_mil, colWidths=col_militares, repeatRows=1)
        tabla_mil.setStyle(TableStyle([
            ('BOX',           (0, 0), (-1, -1), 0.5, colors.black),
            ('INNERGRID',     (0, 0), (-1, -1), 0.3, colors.black),
            ('LINEBELOW',     (0, 0), (-1, 0), 0.8, colors.black),
            ('VALIGN',        (0, 0), (-1, -1), 'MIDDLE'),
            ('TOPPADDING',    (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
            ('LEFTPADDING',   (0, 0), (-1, -1), 4),
            ('RIGHTPADDING',  (0, 0), (-1, -1), 4),
        ]))
        story.append(tabla_mil)
        story.append(Spacer(1, 10))

    # ── ACTUADOS POR MILITAR ─────────────────────────────────────────────────
    story.append(Paragraph("<u>ACTUADOS</u>", s_seccion))
    story.append(Spacer(1, 6))

    cw_actuados = [usable_w * p for p in (0.20, 0.09, 0.12, 0.44, 0.15)]

    hist_simple = {
        'resoluciones':         resoluciones.filter(instancia='PRIMERA'),
        'segundas_resoluciones': resoluciones.filter(instancia='RECONSIDERACION'),
        'autos_tpe':            autos_tpe,
        'recursos_apelacion':   recursos_tsp.filter(instancia='APELACION'),
        'raees':                recursos_tsp.filter(instancia='ACLARACION_ENMIENDA'),
        'autos_tsp':            AUTOTSP.objects.filter(sim=sim),
    }

    def _tabla_docs(documentos):
        """Construye la tabla de actuados y su TableStyle dinámico."""
        filas = [[
            Paragraph('TIPO DE DOCUMENTO', s_th),
            Paragraph('N°', s_th),
            Paragraph('FECHA', s_th),
            Paragraph('RESOLUTIVA', s_th),
            Paragraph('NOTIF.', s_th),
        ]]
        ts = [
            ('BOX',           (0, 0), (-1, -1), 0.5, colors.black),
            ('INNERGRID',     (0, 0), (-1, -1), 0.3, colors.black),
            ('LINEBELOW',     (0, 0), (-1, 0),  0.8, colors.black),
            ('VALIGN',        (0, 0), (-1, -1), 'MIDDLE'),
            ('TOPPADDING',    (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
            ('LEFTPADDING',   (0, 0), (-1, -1), 4),
            ('RIGHTPADDING',  (0, 0), (-1, -1), 4),
        ]
        for d in documentos:
            # Sub-nota RAP (no es fila completa, se muestra como el memorándum)
            if d.get('tipo') == '_RAP_NOTA':
                rap_txt = (f"<i>Elevado al T.S.P. — Oficio N° {d['numero_oficio'] or 'S/N'}"
                           f"  —  Fecha: {_format_date(d['fecha_oficio'])}</i>")
                ri = len(filas)
                filas.append([Paragraph(rap_txt, s_memo), '', '', '', ''])
                ts += [
                    ('SPAN',          (0, ri), (-1, ri)),
                    ('TOPPADDING',    (0, ri), (-1, ri), 2),
                    ('BOTTOMPADDING', (0, ri), (-1, ri), 3),
                    ('LEFTPADDING',   (0, ri), (-1, ri), 12),
                ]
                continue

            # Formatear notificación: dict con 'tipo' y 'fecha', o None
            notif_str = 'NO'
            if d['notificacion'] and isinstance(d['notificacion'], dict):
                notif_tipo = d['notificacion'].get('tipo', 'N/A')
                notif_fecha = d['notificacion'].get('fecha')
                notif_fecha_str = _format_date(notif_fecha) if notif_fecha else 'N/A'
                notif_str = f"{notif_tipo}\n({notif_fecha_str})"

            filas.append([
                Paragraph(d['tipo'],                    s_td),
                Paragraph(str(d['numero']),             s_td_c),
                Paragraph(_format_date(d['fecha_doc']), s_td_c),
                Paragraph(d['resolutiva'],              s_td),
                Paragraph(notif_str,                    s_td_c),
            ])
            if d['memo']:
                memo = d['memo']
                entrega = _format_date(memo['fecha_recepcion']) if memo['fecha_recepcion'] else 'PENDIENTE'
                memo_txt = f"<i>Memorándum N° {memo['numero']}  —  Fecha: {_format_date(memo['fecha'])}  —  Entrega: {entrega}</i>"
                ri = len(filas)
                filas.append([Paragraph(memo_txt, s_memo), '', '', '', ''])
                ts += [
                    ('SPAN',          (0, ri), (-1, ri)),
                    ('TOPPADDING',    (0, ri), (-1, ri), 2),
                    ('BOTTOMPADDING', (0, ri), (-1, ri), 3),
                    ('LEFTPADDING',   (0, ri), (-1, ri), 12),
                ]
        t = Table(filas, colWidths=cw_actuados, repeatRows=1)
        t.setStyle(TableStyle(ts))
        return t

    hay_actuados = False

    for pm_obj in militares:
        docs_pm, _ = _compilar_documentos(sim, hist_simple, pm=pm_obj)

        # RAP: agregar como sub-nota (tipo '_RAP_NOTA'), se renderiza igual que memorándum
        for rap in hist_simple['recursos_apelacion'].filter(pm=pm_obj).order_by('fecha_oficio'):
            docs_pm.append({
                'tipo': '_RAP_NOTA',
                'numero_oficio': rap.numero_oficio,
                'fecha_oficio': rap.fecha_oficio,
                'fecha_doc': rap.fecha_oficio or rap.fecha_presentacion,
            })
        docs_pm.sort(key=lambda x: x['fecha_doc'] or date.min)

        if not docs_pm:
            continue
        hay_actuados = True
        nombre_pm = f"{pm_obj.get_grado_display() or ''} {pm_obj.nombre or ''} {pm_obj.paterno or ''} {pm_obj.materno or ''}".strip().upper()
        story.append(Paragraph(nombre_pm, s_pm_sub))
        story.append(Spacer(1, 3))
        story.append(_tabla_docs(docs_pm))
        story.append(Spacer(1, 10))

    # Autos TSP — sin FK a militar, se muestran al final como sección propia
    autos_tsp_qs = AUTOTSP.objects.filter(sim=sim)
    if autos_tsp_qs.exists():
        hay_actuados = True
        story.append(Paragraph("AUTOS TSP", s_pm_sub))
        story.append(Spacer(1, 3))
        docs_tsp = []
        for autotsp in autos_tsp_qs.order_by('fecha'):
            notif_info = None
            _notif = getattr(autotsp, 'notificacion', None)
            if _notif:
                notif_info = {
                    'tipo': _notif.get_tipo_display(),
                    'fecha': _notif.fecha,
                }
            docs_tsp.append({
                'tipo': 'AUTO TSP',
                'numero': autotsp.numero or 'S/N',
                'fecha_doc': autotsp.fecha,
                'resolutiva': (autotsp.texto or (autotsp.get_tipo_display() if autotsp.tipo else 'N/A')).upper(),
                'notificacion': notif_info,
                'memo': None,
            })
        story.append(_tabla_docs(docs_tsp))
        story.append(Spacer(1, 10))

    if not hay_actuados:
        story.append(Paragraph("Sin actuados registrados.", s_dato))

    doc.build(story, onFirstPage=_pie_pagina, onLaterPages=_pie_pagina)
    buffer.seek(0)

    fecha_export = datetime.now().strftime("%d-%m-%Y")
    filename = f"SIM_{_sanitize_filename(sim.codigo)}_{fecha_export}.pdf"

    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


@login_required
def export_sim_excel(request, sim_id):
    """Exporta un SIM completo a Excel con 3 hojas: SIM, Militares, Actuados"""
    sim = get_object_or_404(SIM, id=sim_id)
    militares = sim.militares.all()
    resoluciones = Resolucion.objects.filter(sim=sim)
    autos_tpe = AUTOTPE.objects.filter(sim=sim)
    autos_tsp = AUTOTSP.objects.filter(sim=sim)
    recursos_tsp = RecursoTSP.objects.filter(sim=sim)

    wb = Workbook()
    wb.remove(wb.active)

    # ===== HOJA 1: INFORMACIÓN DEL SIM =====
    ws_sim = wb.create_sheet("SIM")
    ws_sim['A1'] = "INFORMACIÓN DEL SUMARIO"
    ws_sim['A1'].font = Font(bold=True, size=12)

    sim_data = [
        ("Código:", sim.codigo),
        ("Tipo:", sim.get_tipo_display()),
        ("Objeto:", sim.objeto),
        ("Resumen:", sim.resumen),
        ("Fecha Ingreso:", _format_date(sim.fecha_ingreso)),
        ("Estado:", sim.get_estado_display()),
    ]

    for idx, (label, value) in enumerate(sim_data, start=2):
        ws_sim[f'A{idx}'] = label
        ws_sim[f'A{idx}'].font = Font(bold=True)
        ws_sim[f'B{idx}'] = value

    ws_sim.column_dimensions['A'].width = 20
    ws_sim.column_dimensions['B'].width = 50

    # ===== HOJA 2: MILITARES =====
    ws_militares = wb.create_sheet("MILITARES")
    headers_militares = ["Grado", "Nombre", "Apellido Paterno", "Apellido Materno", "CI", "Arma", "Escalafón"]
    for col, header in enumerate(headers_militares, 1):
        cell = ws_militares.cell(row=1, column=col)
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="185FA5", end_color="185FA5", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")

    for idx, militar in enumerate(militares, start=2):
        ws_militares.cell(row=idx, column=1, value=militar.grado)
        ws_militares.cell(row=idx, column=2, value=militar.nombre)
        ws_militares.cell(row=idx, column=3, value=militar.paterno)
        ws_militares.cell(row=idx, column=4, value=militar.materno or "")
        ws_militares.cell(row=idx, column=5, value=str(militar.ci) if militar.ci else "")
        ws_militares.cell(row=idx, column=6, value=militar.get_arma_display() or "")
        ws_militares.cell(row=idx, column=7, value=militar.get_escalafon_display() or "")

    for col in range(1, len(headers_militares) + 1):
        ws_militares.column_dimensions[chr(64 + col)].width = 18

    # ===== HOJA 3: ACTUADOS (Coordinado) =====
    ws_actuados = wb.create_sheet("ACTUADOS")
    headers_actuados = ["DOCUMENTO", "NÚMERO", "FECHA", "RESOLUTIVA", "NOTIFICACIÓN", "MEMO N°", "MEMO FECHA", "MEMO ENTREGA"]
    for col, header in enumerate(headers_actuados, 1):
        cell = ws_actuados.cell(row=1, column=col)
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="185FA5", end_color="185FA5", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")

    row_idx = 2

    # Compilar documentos coordinados
    hist_sim = {
        'resoluciones': resoluciones.filter(instancia='PRIMERA'),
        'segundas_resoluciones': resoluciones.filter(instancia='RECONSIDERACION'),
        'autos_tpe': autos_tpe,
        'recursos_apelacion': recursos_tsp.filter(instancia='APELACION'),
        'raees': recursos_tsp.filter(instancia='ACLARACION_ENMIENDA'),
        'autos_tsp': autos_tsp,
    }
    documentos, info_tsp = _compilar_documentos(sim, hist_sim)

    for doc in documentos:
        ws_actuados.cell(row=row_idx, column=1, value=doc['tipo'])
        ws_actuados.cell(row=row_idx, column=2, value=doc['numero'])
        ws_actuados.cell(row=row_idx, column=3, value=_format_date(doc['fecha_doc']))
        ws_actuados.cell(row=row_idx, column=4, value=doc['resolutiva'])
        notif = doc['notificacion']
        notif_str = f"{notif['tipo']} ({_format_date(notif['fecha'])})" if notif else ""
        ws_actuados.cell(row=row_idx, column=5, value=notif_str)
        if doc['memo']:
            ws_actuados.cell(row=row_idx, column=6, value=doc['memo']['numero'])
            ws_actuados.cell(row=row_idx, column=7, value=_format_date(doc['memo']['fecha']))
            ws_actuados.cell(row=row_idx, column=8, value=_format_date(doc['memo']['fecha_recepcion']))
        row_idx += 1

    # Agregar fila de estado TSP si hay
    if info_tsp:
        ws_actuados.cell(row=row_idx, column=1, value="EN TRIBUNAL SUPREMO POLICIAL")
        ws_actuados.cell(row=row_idx, column=2, value=info_tsp['numero_oficio'])
        ws_actuados.cell(row=row_idx, column=3, value=_format_date(info_tsp['fecha_oficio']))
        # Aplicar negrita para destacar
        for col_num in range(1, 4):
            ws_actuados.cell(row=row_idx, column=col_num).font = Font(bold=True)
        row_idx += 1

    # Ajustar anchos de columna
    ws_actuados.column_dimensions['A'].width = 25  # DOCUMENTO
    ws_actuados.column_dimensions['B'].width = 12  # NÚMERO
    ws_actuados.column_dimensions['C'].width = 14  # FECHA
    ws_actuados.column_dimensions['D'].width = 35  # RESOLUTIVA
    ws_actuados.column_dimensions['E'].width = 14  # NOTIFICACIÓN
    ws_actuados.column_dimensions['F'].width = 12  # MEMO N°
    ws_actuados.column_dimensions['G'].width = 14  # MEMO FECHA
    ws_actuados.column_dimensions['H'].width = 14  # MEMO ENTREGA

    # Guardar
    fecha_export = datetime.now().strftime("%Y-%m-%d")
    excel_filename = f"SIM_{_sanitize_filename(sim.codigo)}_{fecha_export}.xlsx"

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(buffer, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{excel_filename}"'
    return response
