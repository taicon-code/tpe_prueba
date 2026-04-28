# tpe_app/views/buscador_views.py
import unicodedata
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db.models import Q, Value
from django.db.models.functions import Replace, Collate
from ..decorators import rol_requerido
from ..models import SIM, PM, AUTOTPE, AUTOTSP, Resolucion, RecursoTSP, DocumentoAdjunto, CustodiaSIM


def _normalizar(texto):
    """Quita acentos y ñ para búsqueda flexible. 'alarcón' → 'ALARCON', 'siñani' → 'SINANI'."""
    nfkd = unicodedata.normalize('NFKD', texto)
    sin_acentos = ''.join(c for c in nfkd if not unicodedata.combining(c))
    return sin_acentos.upper()


def _campo_sin_n(campo):
    """Reemplaza Ñ→N en el campo de BD para comparación. Los datos están en mayúsculas."""
    return Replace(campo, Value('Ñ'), Value('N'))


def _obtener_historial_completo(personal_id):
    """Obtiene el historial completo de un personal"""
    try:
        personal = PM.objects.get(id=personal_id)
    except PM.DoesNotExist:
        return None

    # Obtener todos los SIM donde participa este personal
    sims = SIM.objects.filter(militares__id=personal_id).distinct()
    sim_ids = list(sims.values_list('id', flat=True))

    historial = {
        'personal': personal,
        'sumarios': sims,
        'resoluciones': Resolucion.objects.filter(sim__in=sim_ids, instancia='PRIMERA'),
        'segundas_resoluciones': Resolucion.objects.filter(sim__in=sim_ids, instancia='RECONSIDERACION'),
        'recursos_apelacion': RecursoTSP.objects.filter(sim__in=sim_ids, instancia='APELACION'),
        'raees': RecursoTSP.objects.filter(sim__in=sim_ids, instancia='ACLARACION_ENMIENDA'),
        'autos_tpe': AUTOTPE.objects.filter(sim__in=sim_ids),
        'autos_tsp': AUTOTSP.objects.filter(sim__in=sim_ids),
    }

    return historial


def _compilar_documentos_lotes(sim, historial):
    """Compila documentos coordinados para reportes por lote
    Retorna string formateado con tipo, numero, fecha y resolutiva"""
    documentos = []

    # Resoluciones
    for res in historial['resoluciones'].filter(sim=sim):
        fecha_str = res.fecha.strftime('%d/%m/%y') if res.fecha else 'S/F'
        resolutiva = (res.texto or 'N/A').upper() if res.texto else 'N/A'
        documentos.append(('RES', res.numero or 'S/N', fecha_str, resolutiva, None))

    # Segundas Resoluciones
    for rr in historial.get('segundas_resoluciones', Resolucion.objects.none()).filter(sim=sim):
        fecha_str = rr.fecha.strftime('%d/%m/%y') if rr.fecha else 'S/F'
        resolutiva = (rr.texto or 'N/A').upper() if rr.texto else 'N/A'
        documentos.append(('RR', rr.numero or 'S/N', fecha_str, resolutiva, None))

    # Autos TPE
    for auto in historial['autos_tpe'].filter(sim=sim):
        fecha_str = auto.fecha.strftime('%d/%m/%y') if auto.fecha else 'S/F'
        resolutiva = (auto.texto or (auto.get_tipo_display() if auto.tipo else 'N/A')).upper()
        memo = getattr(auto, 'memorandum', None)
        if memo:
            entrega = memo.fecha_entrega.strftime('%d/%m/%y') if memo.fecha_entrega else 'PENDIENTE'
            memo_str = f"MEMO N° {memo.numero}  |  FECHA: {memo.fecha.strftime('%d/%m/%y') if memo.fecha else 'S/F'}  |  ENTREGA: {entrega}"
        else:
            memo_str = None
        documentos.append(('AUTO TPE', auto.numero or 'S/N', fecha_str, resolutiva, memo_str))

    # Recursos Apelación
    for rap in historial['recursos_apelacion'].filter(sim=sim):
        fecha_str = rap.fecha.strftime('%d/%m/%y') if rap.fecha else 'S/F'
        resolutiva = 'RECURSO DE APELACIÓN'
        documentos.append(('RAP', rap.numero or 'S/N', fecha_str, resolutiva, None))

    # RAEE
    for raee in historial['raees'].filter(sim=sim):
        fecha_str = raee.fecha.strftime('%d/%m/%y') if raee.fecha else 'S/F'
        resolutiva = 'ACLARACIÓN Y ENMIENDA'
        documentos.append(('RAEE', raee.numero or 'S/N', fecha_str, resolutiva, None))

    # Autos TSP
    for autotsp in historial['autos_tsp'].filter(sim=sim):
        fecha_str = autotsp.fecha.strftime('%d/%m/%y') if autotsp.fecha else 'S/F'
        resolutiva = (autotsp.texto or (autotsp.get_tipo_display() if autotsp.tipo else 'N/A')).upper()
        documentos.append(('AUTO TSP', autotsp.numero or 'S/N', fecha_str, resolutiva, None))

    # Ordenar por fecha
    return documentos


def _obtener_estado_actual(personal_id):
    """Obtiene el estado actual del personal con estadísticas simplificadas.
    Solo cuenta documentos emitidos por el TPE: RES, RR, AUTOTPE"""
    historial = _obtener_historial_completo(personal_id)
    if not historial:
        return None

    return {
        'total_sumarios': historial['sumarios'].count(),
        'total_resoluciones': (historial['resoluciones'].count() +
                               historial['segundas_resoluciones'].count() +
                               historial['autos_tpe'].count()),
        'total_autos_tpe': historial['autos_tpe'].count(),
        'estado_actual': 'Historial disponible'
    }


@login_required
def buscador_dashboard(request):
    """Dashboard para búsqueda unificada - búsqueda por CI, código SIM, nombre, apellidos"""

    query     = request.GET.get('q', '').strip()
    promocion = request.GET.get('promocion', '').strip()
    personal_seleccionado = None
    historial = None
    estado = None
    resultados_pm = []
    resultados_sim = []

    # Búsqueda por año de promoción (lista todos los militares de esa promoción)
    if promocion and promocion.isdigit():
        resultados_pm = list(
            PM.objects.filter(anio_promocion=int(promocion))
            .order_by('paterno', 'nombre')
        )

    if query:
        # Normalizamos el query: quitamos tildes y ñ → 'alarcón'→'ALARCON', 'siñani'→'SINANI'
        q_norm = _normalizar(query)

        # 1. Intentar búsqueda por CI exacto (es muy específico)
        resultados_pm = list(
            PM.objects.filter(ci__iexact=query).distinct()[:20]
        )

        # 2. Si no hay resultados por CI, verificar si es búsqueda por "apellido_paterno, apellido_materno"
        if not resultados_pm and ',' in query:
            partes = [p.strip() for p in query.split(',', 1)]
            if len(partes) == 2:
                ap, am = partes
                ap_norm = _normalizar(ap) if ap else ''
                am_norm = _normalizar(am) if am else ''

                filtro = PM.objects
                if ap_norm:
                    filtro = filtro.annotate(
                        pat_norm=Collate(_campo_sin_n('paterno'), 'utf8mb4_general_ci')
                    ).filter(pat_norm__icontains=ap_norm)
                if am_norm:
                    filtro = filtro.annotate(
                        mat_norm=Collate(_campo_sin_n('materno'), 'utf8mb4_general_ci')
                    ).filter(mat_norm__icontains=am_norm)

                resultados_pm = list(filtro.distinct()[:20])

        # 3. Si aún no hay resultados, buscar por nombre/apellido normalizando (búsqueda general)
        if not resultados_pm:
            resultados_pm = list(
                PM.objects.annotate(
                    pat_norm=Collate(_campo_sin_n('paterno'), 'utf8mb4_general_ci'),
                    nom_norm=Collate(_campo_sin_n('nombre'),  'utf8mb4_general_ci'),
                    mat_norm=Collate(_campo_sin_n('materno'), 'utf8mb4_general_ci'),
                ).filter(
                    Q(pat_norm__icontains=q_norm) |
                    Q(nom_norm__icontains=q_norm) |
                    Q(mat_norm__icontains=q_norm)
                ).distinct()[:20]
            )

        resultados_sim = list(
            SIM.objects.filter(
                Q(codigo__icontains=query) |
                Q(resumen__icontains=query) |
                Q(objeto__icontains=query)
            ).prefetch_related('abogados', 'militares').distinct()[:20]
        )

        # Si hay exactamente 1 PM, mostrar su historial completo
        if len(resultados_pm) == 1:
            personal_seleccionado = resultados_pm[0]
            historial = _obtener_historial_completo(personal_seleccionado.id)
            estado = _obtener_estado_actual(personal_seleccionado.id)

    context = {
        'query': query,
        'promocion': promocion,
        'resultados_pm': resultados_pm,
        'resultados_sim': resultados_sim,
        'total_pm': len(resultados_pm),
        'total_sim': len(resultados_sim),
        'personal_seleccionado': personal_seleccionado,
        'historial': historial,
        'estado': estado,
    }
    return render(request, 'tpe_app/buscador/dashboard_buscador.html', context)


@login_required
def detalles_sim(request, sim_id):
    """Vista detallada de un SIM: militares, resoluciones, autos, custodia (solo Admin2), etc."""

    sim = get_object_or_404(SIM, id=sim_id)

    # Obtener todos los militares del SIM
    militares = sim.militares.all()

    # Obtener todos los actuados del SIM
    resoluciones = Resolucion.objects.filter(sim=sim).select_related('abogado', 'pm')
    autos_tpe = AUTOTPE.objects.filter(sim=sim).select_related('abogado', 'pm')
    autos_tsp = AUTOTSP.objects.filter(sim=sim)
    recursos_tsp = RecursoTSP.objects.filter(sim=sim).select_related('abogado', 'pm')

    # Agrupar actuados por militar en orden cronológico del flujo
    militares_con_docs = []
    for pm_obj in militares:
        militares_con_docs.append({
            'pm': pm_obj,
            'resoluciones': resoluciones.filter(pm=pm_obj, instancia='PRIMERA').order_by('fecha'),
            'rrs':          resoluciones.filter(pm=pm_obj, instancia='RECONSIDERACION').order_by('fecha'),
            'autos_tpe':    autos_tpe.filter(pm=pm_obj).order_by('fecha'),
            'recursos_tsp': recursos_tsp.filter(pm=pm_obj).order_by('fecha'),
        })

    # Obtener historial de custodia (trazabilidad) - SOLO para Admin2
    custodia_historial = None
    custodia_actual = None
    es_admin2 = hasattr(request.user, 'perfilusuario') and request.user.perfilusuario.rol == 'ADMIN2_ARCHIVO'

    if es_admin2:
        custodia_historial = CustodiaSIM.objects.filter(sim=sim).select_related('abogado').order_by('fecha_recepcion')
        custodia_actual = CustodiaSIM.objects.filter(sim=sim, estado='RECIBIDA_CONFORME').select_related('abogado').first()

    context = {
        'sim': sim,
        'militares': militares,
        'militares_con_docs': militares_con_docs,
        'autos_tsp': autos_tsp,
        'custodia_historial': custodia_historial,
        'custodia_actual': custodia_actual,
        'es_admin2': es_admin2,
    }
    return render(request, 'tpe_app/buscador/detalles_sim.html', context)


@login_required
def busqueda_por_lotes(request):
    """Vista para búsqueda y reporte por lotes de múltiples militares por AP + AM"""
    militares_encontrados = []

    if request.method == 'POST':
        lista_apellidos = request.POST.get('lista_apellidos', '').strip()

        if lista_apellidos:
            # Procesar cada línea como "APELLIDO_PATERNO, APELLIDO_MATERNO"
            lineas = [l.strip() for l in lista_apellidos.split('\n') if l.strip()]

            for linea in lineas:
                partes = [p.strip() for p in linea.split(',')]
                if len(partes) >= 2:
                    ap = partes[0].upper()
                    am = partes[1].upper()
                    nb = partes[2].upper() if len(partes) >= 3 else None

                    filtro = PM.objects.filter(materno__iexact=am)
                    if ap:
                        filtro = filtro.filter(paterno__iexact=ap)
                    else:
                        filtro = filtro.filter(Q(paterno='') | Q(paterno__isnull=True))
                    if nb:
                        filtro = filtro.filter(nombre__icontains=nb)

                    for pm in filtro:
                        historial = _obtener_historial_completo(pm.id)
                        if historial:
                            militares_encontrados.append({
                                'personal': pm,
                                'historial': historial,
                            })

    context = {
        'militares_encontrados': militares_encontrados,
    }
    return render(request, 'tpe_app/buscador/busqueda_lotes.html', context)


@login_required
def export_batch_pdf(request):
    """Genera PDF con tabla compacta de múltiples militares"""
    from django.http import HttpResponse
    from io import BytesIO
    from reportlab.lib.pagesizes import letter
    from reportlab.lib.units import inch
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.enums import TA_CENTER, TA_LEFT
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    from datetime import datetime

    # Registrar fuente Arial desde Windows
    try:
        pdfmetrics.getFont('Arial-Bold')
    except KeyError:
        try:
            pdfmetrics.registerFont(TTFont('Arial',      'C:/Windows/Fonts/arial.ttf'))
            pdfmetrics.registerFont(TTFont('Arial-Bold', 'C:/Windows/Fonts/arialbd.ttf'))
        except Exception:
            # Si las fuentes no existen, usar Helvetica por defecto
            pass

    lista_apellidos = request.POST.get('lista_apellidos', '').strip()

    if not lista_apellidos:
        return HttpResponse("No se proporcionó lista de militares", status=400)

    militares = []
    lineas = [l.strip() for l in lista_apellidos.split('\n') if l.strip()]

    for linea in lineas:
        partes = [p.strip() for p in linea.split(',')]
        if len(partes) >= 2:
            ap = partes[0].upper()
            am = partes[1].upper()
            nb = partes[2].upper() if len(partes) >= 3 else None

            filtro = PM.objects.filter(materno__iexact=am)
            if ap:
                filtro = filtro.filter(paterno__iexact=ap)
            else:
                filtro = filtro.filter(Q(paterno='') | Q(paterno__isnull=True))
            if nb:
                filtro = filtro.filter(nombre__icontains=nb)

            for pm in filtro:
                historial = _obtener_historial_completo(pm.id)
                if historial:
                    militares.append({
                        'personal': pm,
                        'historial': historial,
                    })

    if not militares:
        return HttpResponse("No se encontraron militares con los datos proporcionados", status=404)

    buffer = BytesIO()
    page_w, _ = letter
    margin = 0.5 * inch

    def _estilo(nombre, fuente='Helvetica', tamaño=9, alineacion=TA_LEFT,
                negrita=False, color=colors.black, espacio_antes=0, espacio_despues=2,
                interlinea=11, sangria=0):
        fn = (fuente + '-Bold') if negrita else fuente
        return ParagraphStyle(nombre, fontName=fn, fontSize=tamaño,
                              alignment=alineacion, textColor=color,
                              spaceBefore=espacio_antes, spaceAfter=espacio_despues,
                              leading=interlinea, leftIndent=sangria)

    s_inst1   = _estilo('inst1',  fuente='Arial', tamaño=10, alineacion=TA_LEFT, negrita=True, espacio_despues=1, interlinea=9)
    s_inst2   = _estilo('inst2',  fuente='Arial', tamaño=10, alineacion=TA_LEFT, negrita=True, espacio_despues=1, interlinea=9, sangria=15)
    s_pais    = _estilo('pais',   fuente='Arial', tamaño=10, alineacion=TA_LEFT, negrita=True, espacio_despues=6, interlinea=9, sangria=70)
    s_seccion = _estilo('secc',   fuente='Arial', tamaño=14, alineacion=TA_CENTER, negrita=True, espacio_antes=6, espacio_despues=4, interlinea=17)
    s_th = _estilo('th', tamaño=7, alineacion=TA_CENTER, negrita=True, color=colors.black)
    s_td = _estilo('td', tamaño=7, interlinea=9)
    s_td_c = _estilo('tdc', tamaño=7, alineacion=TA_CENTER, interlinea=9)

    doc = SimpleDocTemplate(
        buffer, pagesize=letter,
        leftMargin=margin, rightMargin=margin,
        topMargin=0.55 * inch, bottomMargin=0.75 * inch,
    )
    usable_w = page_w - 2 * margin
    story = []

    # ── ENCABEZADO INSTITUCIONAL ─────────────────────────────────────────────
    story.append(Paragraph("COMANDO GENERAL DEL EJÉRCITO", s_inst1))
    story.append(Paragraph("DEPARTAMENTO I - PERSONAL", s_inst2))
    story.append(Paragraph("<u>BOLIVIA</u>", s_pais))

    # ── REPORTE POR LOTES ─────────────────────────────────────────────────
    story.append(Paragraph("<u>ANTECEDENTES DISCIPLINARIOS TRATADOS POR EL TRIBUNAL DE PERSONAL DEL EJÉRCITO</u>", s_seccion))
    story.append(Spacer(1, 6))

    # Pie de página
    grado_usuario = ""
    espec_usuario = ""
    nombre_completo = request.user.get_full_name() or request.user.username
    if request.user.is_authenticated:
        try:
            perfil = request.user.perfilusuario
            pm_pie = perfil.pm if perfil.pm else (perfil.vocal.pm if perfil.vocal and perfil.vocal.pm else None)
            if pm_pie:
                grado_usuario  = pm_pie.get_grado_display() or ""
                espec_usuario  = pm_pie.get_arma_display() or ""
                nombre_completo = f"{pm_pie.nombre or ''} {pm_pie.paterno or ''} {pm_pie.materno or ''}".strip()
        except Exception:
            pass

    partes_pie = [p for p in [grado_usuario, espec_usuario, nombre_completo.upper()] if p]
    texto_impreso = "  ".join(partes_pie)

    fecha_hoy = datetime.now().strftime("%d/%m/%Y")
    hora_hoy  = datetime.now().strftime("%H:%M")

    def _pie_pagina(canv, doc):
        canv.saveState()
        canv.setStrokeColor(colors.lightgrey)
        canv.setLineWidth(0.5)
        canv.line(margin, 0.52 * inch, page_w - margin, 0.52 * inch)
        canv.setFont('Helvetica', 6.5)
        canv.setFillColor(colors.grey)
        texto_pie = (f"Impreso por: {texto_impreso}   |   "
                     f"{fecha_hoy}  {hora_hoy}   |   Pág. {doc.page}")
        canv.drawCentredString(page_w / 2, 0.33 * inch, texto_pie)
        canv.restoreState()

    # Construir tabla con todos los militares
    # Ancho distribuido: GRADO=8% | NOMBRES=15% | AP=15% | AM=15% | SIM=10% | OBJETO=18% | ACTUADOS=12% | ESTADO=7%
    col_widths = [usable_w * p for p in (0.08, 0.15, 0.15, 0.15, 0.10, 0.18, 0.12, 0.07)]

    filas = [[
        Paragraph('GRADO', s_th),
        Paragraph('NOMBRES', s_th),
        Paragraph('APELLIDO PATERNO', s_th),
        Paragraph('APELLIDO MATERNO', s_th),
        Paragraph('SIM', s_th),
        Paragraph('OBJETO', s_th),
        Paragraph('ACTUADOS', s_th),
        Paragraph('ESTADO', s_th),
    ]]

    for mil in militares:
        pm = mil['personal']
        hist = mil['historial']

        for sim in hist['sumarios']:
            # Obtener actuados coordinados para este SIM
            documentos = _compilar_documentos_lotes(sim, hist)
            actuados_list = []
            contador = 1
            numeros_circulos = ['①', '②', '③', '④', '⑤', '⑥', '⑦', '⑧', '⑨', '⑩']

            for tipo, numero, fecha, resolutiva, memo in documentos:
                num_circulo = numeros_circulos[min(contador - 1, 9)]
                linea = f"{num_circulo} {tipo} {numero} ({fecha})<br/>   {resolutiva}"
                if memo:
                    linea += f"<br/>   <i>└─ {memo}</i>"
                actuados_list.append(linea + "<br/><br/>")
                contador += 1

            # Unir todo con HTML y remover último <br/><br/>
            actuados_str = ''.join(actuados_list) if actuados_list else 'PENDIENTE'
            if actuados_str.endswith('<br/><br/>'):
                actuados_str = actuados_str[:-10]  # Remover último <br/><br/>

            objeto_completo = (sim.objeto or 'N/A').upper()

            filas.append([
                Paragraph((pm.get_grado_display() or 'N/A').upper(), s_td_c),
                Paragraph((pm.nombre or 'N/A').upper(), s_td),
                Paragraph((pm.paterno or 'N/A').upper(), s_td),
                Paragraph((pm.materno or 'N/A').upper(), s_td),
                Paragraph(sim.codigo or 'N/A', s_td_c),
                Paragraph(objeto_completo, s_td),
                Paragraph(actuados_str, s_td),
                Paragraph((sim.get_estado_display() or 'N/A').upper(), s_td_c),
            ])

    tabla = Table(filas, colWidths=col_widths, repeatRows=1)
    tabla.setStyle(TableStyle([
        ('BOX', (0, 0), (-1, -1), 0.5, colors.black),
        ('INNERGRID', (0, 0), (-1, -1), 0.3, colors.black),
        ('BACKGROUND', (0, 0), (-1, 0), colors.white),
        ('LINEBELOW', (0, 0), (-1, 0), 0.8, colors.black),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('TOPPADDING', (0, 0), (-1, -1), 3),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
        ('LEFTPADDING', (0, 0), (-1, -1), 3),
        ('RIGHTPADDING', (0, 0), (-1, -1), 3),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 8),
        ('FONTSIZE', (0, 1), (-1, -1), 7),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
    ]))

    story.append(tabla)

    doc.build(story, onFirstPage=_pie_pagina, onLaterPages=_pie_pagina)
    buffer.seek(0)

    fecha_export = datetime.now().strftime("%d-%m-%Y")
    filename = f"ANTECEDENTES_LOTE_{fecha_export}.pdf"
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


@login_required
def export_batch_excel(request):
    """Genera Excel con tabla de múltiples militares"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from io import BytesIO
    from django.http import HttpResponse
    from datetime import datetime

    lista_apellidos = request.POST.get('lista_apellidos', '').strip()

    if not lista_apellidos:
        return HttpResponse("No se proporcionó lista de militares", status=400)

    militares = []
    lineas = [l.strip() for l in lista_apellidos.split('\n') if l.strip()]

    for linea in lineas:
        partes = [p.strip() for p in linea.split(',')]
        if len(partes) >= 2:
            ap = partes[0].upper()
            am = partes[1].upper()
            nb = partes[2].upper() if len(partes) >= 3 else None

            filtro = PM.objects.filter(materno__iexact=am)
            if ap:
                filtro = filtro.filter(paterno__iexact=ap)
            else:
                filtro = filtro.filter(Q(paterno='') | Q(paterno__isnull=True))
            if nb:
                filtro = filtro.filter(nombre__icontains=nb)

            for pm in filtro:
                historial = _obtener_historial_completo(pm.id)
                if historial:
                    militares.append({
                        'personal': pm,
                        'historial': historial,
                    })

    if not militares:
        return HttpResponse("No se encontraron militares con los datos proporcionados", status=404)

    wb = Workbook()
    ws = wb.active
    ws.title = "ANTECEDENTES"

    # Encabezados
    headers = ['GRADO', 'NOMBRES', 'APELLIDO PATERNO', 'APELLIDO MATERNO', 'SIM', 'OBJETO', 'ACTUADOS', 'ESTADO']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF", size=10)
        cell.fill = PatternFill(start_color="185FA5", end_color="185FA5", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    row_idx = 2

    for mil in militares:
        pm = mil['personal']
        hist = mil['historial']

        for sim in hist['sumarios']:
            # Obtener actuados coordinados para este SIM
            documentos = _compilar_documentos_lotes(sim, hist)
            actuados_list = []
            contador = 1
            numeros_circulos = ['①', '②', '③', '④', '⑤', '⑥', '⑦', '⑧', '⑨', '⑩']

            for tipo, numero, fecha, resolutiva, memo in documentos:
                num_circulo = numeros_circulos[min(contador - 1, 9)]
                linea = f"{num_circulo} {tipo} {numero} ({fecha})\n   {resolutiva}"
                if memo:
                    linea += f"\n   └─ {memo}"
                actuados_list.append(linea + "\n")
                contador += 1

            # Unir todo y remover último salto de línea
            actuados_str = ''.join(actuados_list) if actuados_list else 'PENDIENTE'
            actuados_str = actuados_str.rstrip('\n')

            ws.cell(row=row_idx, column=1, value=(pm.get_grado_display() or 'N/A').upper())
            ws.cell(row=row_idx, column=2, value=(pm.nombre or 'N/A').upper())
            ws.cell(row=row_idx, column=3, value=(pm.paterno or 'N/A').upper())
            ws.cell(row=row_idx, column=4, value=(pm.materno or 'N/A').upper())
            ws.cell(row=row_idx, column=5, value=sim.codigo or 'N/A')
            ws.cell(row=row_idx, column=6, value=(sim.objeto or 'N/A').upper())
            ws.cell(row=row_idx, column=7, value=actuados_str)
            ws.cell(row=row_idx, column=8, value=(sim.get_estado_display() or 'N/A').upper())

            row_idx += 1

    # Ajustar anchos
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 28
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 12

    fecha_export = datetime.now().strftime("%Y-%m-%d")
    excel_filename = f"ANTECEDENTES_LOTE_{fecha_export}.xlsx"

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(buffer, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{excel_filename}"'
    return response


@login_required
def upload_foto_pm(request, pm_id):
    """Subir o reemplazar la foto de un Personal Militar"""
    pm = get_object_or_404(PM, pk=pm_id)

    if request.method == 'POST':
        foto = request.FILES.get('foto')
        if foto:
            # Validar que sea imagen
            content_type = foto.content_type or ''
            if not content_type.startswith('image/'):
                messages.error(request, '❌ El archivo debe ser una imagen (JPG, PNG, etc.)')
            else:
                # Eliminar foto anterior si existe
                if pm.foto:
                    pm.foto.delete(save=False)
                pm.foto = foto
                pm.save(update_fields=['foto'])
                messages.success(request, f'✅ Foto actualizada para {pm.nombre} {pm.paterno}')
        else:
            messages.error(request, '❌ No se seleccionó ningún archivo')

    # Volver al buscador con la misma búsqueda si venía de ahí
    referer = request.POST.get('next') or request.META.get('HTTP_REFERER', '')
    if 'buscador' in referer or 'q=' in referer:
        return redirect(referer)
    return redirect('buscador_dashboard')


@login_required
def export_custodia_pdf(request, sim_id):
    """Descargar PDF del historial de custodia de un SIM (Solo Admin2)"""
    from django.http import HttpResponse
    from django.utils import timezone as tz
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch, cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib import colors
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    import os

    # Registrar fuentes TrueType para soporte completo de acentos y ñ
    _fonts_dir = r'C:\Windows\Fonts'
    _arial      = os.path.join(_fonts_dir, 'arial.ttf')
    _arial_bold = os.path.join(_fonts_dir, 'arialbd.ttf')
    if os.path.exists(_arial):
        pdfmetrics.registerFont(TTFont('Arial', _arial))
        pdfmetrics.registerFont(TTFont('Arial-Bold', _arial_bold))
        FONT_NORMAL = 'Arial'
        FONT_BOLD   = 'Arial-Bold'
    else:
        # Fallback si no está Arial (Linux/Mac en producción)
        FONT_NORMAL = 'Helvetica'
        FONT_BOLD   = 'Helvetica-Bold'

    # Verificar que sea Admin2
    if not (hasattr(request.user, 'perfilusuario') and request.user.perfilusuario.rol == 'ADMIN2_ARCHIVO'):
        messages.error(request, '❌ No tienes permiso para descargar este archivo')
        return redirect('admin2_dashboard')

    sim = get_object_or_404(SIM, id=sim_id)
    custodia_historial = CustodiaSIM.objects.filter(sim=sim).select_related('abog').order_by('fecha_recepcion')

    # Crear PDF en orientación vertical (portrait)
    response = HttpResponse(content_type='application/pdf')
    now_local = tz.localtime(tz.now())
    response['Content-Disposition'] = f'attachment; filename="custodia_{sim.codigo}_{now_local.strftime("%d%m%Y")}.pdf"'

    doc = SimpleDocTemplate(response, pagesize=A4, topMargin=0.5*inch, bottomMargin=0.5*inch,
                            leftMargin=0.5*inch, rightMargin=0.5*inch)
    story = []
    styles = getSampleStyleSheet()

    # Título (sin emoji para evitar caracteres mezclados en reportlab)
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=13,
        textColor=colors.HexColor('#185FA5'),
        spaceAfter=10,
        alignment=1
    )
    story.append(Paragraph(f'Historial de Custodia de Carpeta - {sim.codigo}', title_style))
    story.append(Spacer(1, 0.15*inch))

    # Información del SIM — ancho ajustado a portrait (A4 usable ≈ 7.17 in)
    info_data = [
        ['Codigo', sim.codigo],
        ['Tipo', sim.get_tipo_display()],
        ['Estado', sim.get_estado_display()],
        ['Ingreso', sim.fecha_ingreso.strftime('%d/%m/%Y') if sim.fecha_ingreso else '-'],
    ]
    info_table = Table(info_data, colWidths=[1.4*inch, 5.77*inch])
    info_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#E8F4F8')),
        ('BACKGROUND', (1, 0), (1, -1), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('FONTNAME', (0, 0), (0, -1), FONT_BOLD),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#CCCCCC')),
        ('LEFTPADDING', (0, 0), (-1, -1), 8),
        ('RIGHTPADDING', (0, 0), (-1, -1), 8),
        ('TOPPADDING', (0, 0), (-1, -1), 5),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
    ]))
    story.append(info_table)
    story.append(Spacer(1, 0.2*inch))

    # Historial de custodia
    heading_style = ParagraphStyle(
        'SectionHead',
        parent=styles['Heading2'],
        fontSize=10,
        textColor=colors.HexColor('#185FA5'),
        spaceAfter=6,
    )
    story.append(Paragraph('Movimientos de Custodia', heading_style))

    # Estilo para celdas con texto largo (permite salto de línea automático)
    cell_style = ParagraphStyle(
        'CellText',
        parent=styles['Normal'],
        fontName=FONT_NORMAL,
        fontSize=8,
        leading=10,
        wordWrap='CJK',
    )

    # Anchos para portrait A4 (usable ≈ 18.46 cm)
    # Fecha Recep | Custodio | Abogado | Estado | Fecha Entrega | Observacion
    col_widths = [2.2*cm, 3.7*cm, 2.5*cm, 3.2*cm, 2.2*cm, 4.66*cm]

    if custodia_historial:
        custodia_data = [
            ['Fecha Recep.', 'Custodio', 'Abogado', 'Estado', 'Fecha Entrega', 'Observacion']
        ]

        for custodia in custodia_historial:
            observacion = custodia.observacion if custodia.observacion else '-'

            custodia_data.append([
                tz.localtime(custodia.fecha_recepcion).strftime('%d/%m/%Y\n%H:%M'),
                custodia.get_tipo_custodio_display(),
                custodia.abogado.paterno if custodia.abogado else '-',
                custodia.get_estado_display(),
                tz.localtime(custodia.fecha_entrega).strftime('%d/%m/%Y\n%H:%M') if custodia.fecha_entrega else 'Activa',
                Paragraph(observacion, cell_style),
            ])

        custodia_table = Table(custodia_data, colWidths=col_widths)

        custodia_table.setStyle(TableStyle([
            # Cabecera: fondo blanco, texto negro en negrita (sin relleno azul)
            ('BACKGROUND', (0, 0), (-1, 0), colors.white),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
            ('FONTNAME', (0, 0), (-1, 0), FONT_BOLD),
            ('FONTSIZE', (0, 0), (-1, 0), 8),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('VALIGN', (0, 0), (-1, 0), 'MIDDLE'),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 7),
            ('TOPPADDING', (0, 0), (-1, 0), 7),
            # Línea inferior de la cabecera más gruesa para separar visualmente
            ('LINEBELOW', (0, 0), (-1, 0), 1.2, colors.HexColor('#185FA5')),
            # Filas de datos
            ('ALIGN', (0, 1), (4, -1), 'CENTER'),
            ('ALIGN', (5, 1), (5, -1), 'LEFT'),
            ('VALIGN', (0, 1), (-1, -1), 'TOP'),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('FONTNAME', (0, 1), (-1, -1), FONT_NORMAL),
            ('GRID', (0, 0), (-1, -1), 0.4, colors.HexColor('#CCCCCC')),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F5F5F5')]),
            ('LEFTPADDING', (0, 0), (-1, -1), 5),
            ('RIGHTPADDING', (0, 0), (-1, -1), 5),
            ('TOPPADDING', (0, 1), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 4),
        ]))

        story.append(custodia_table)
    else:
        story.append(Paragraph('<i>No hay movimientos de custodia registrados.</i>', styles['Normal']))

    # Pie de página
    story.append(Spacer(1, 0.25*inch))
    pie_style = ParagraphStyle(
        'Footer',
        parent=styles['Normal'],
        fontSize=7,
        textColor=colors.HexColor('#999999'),
        alignment=0,
    )
    generated_time = now_local.strftime('%d/%m/%Y %H:%M:%S')
    story.append(Paragraph(f'Generado: {generated_time}', pie_style))

    doc.build(story)
    return response
