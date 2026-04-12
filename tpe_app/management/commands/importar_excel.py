"""
Comando: importar_excel
Importa datos reales desde docs/tpe2026.xlsx a la base de datos.

Hojas procesadas:
  SIM.                >> modelos SIM + PM + AGENDA
  CASOS ESPECIALES    >> modelos SIM + PM (tipo SOLICITUD_*)
  RESOLUCIONES        >> modelo RES  (vinculado a SIM)
  AUTOS               >> modelo AUTOTPE (vinculado a SIM)
  REC. RECONSIDERACION>> modelo RR   (vinculado a RES + SIM)
  REC. APELACION      >> modelo RAP  (vinculado a SIM)

Uso:
  python manage.py importar_excel
  python manage.py importar_excel --archivo docs/tpe2026.xlsx
  python manage.py importar_excel --limpiar          # borra datos de prueba primero
  python manage.py importar_excel --limpiar --dry-run # solo muestra qué haría
"""

import re
from datetime import date, datetime
from django.core.management.base import BaseCommand
from django.db import transaction
from openpyxl import load_workbook

from tpe_app.models import (
    PM, SIM, PM_SIM, AGENDA, RES, RR, RAP, AUTOTPE,
    RAEE, AUTOTSP, DICTAMEN, ABOG_SIM,
)


# ─── Meses en español ────────────────────────────────────────────────────────
MESES = {
    'ENE': 1, 'FEB': 2, 'MAR': 3, 'ABR': 4,
    'MAY': 5, 'JUN': 6, 'JUL': 7, 'AGO': 8,
    'SEP': 9, 'OCT': 10, 'NOV': 11, 'DIC': 12,
}

# ─── Mapping grado Excel >> clave del modelo ──────────────────────────────────
GRADO_SIMPLE = {
    'CNL.':   ('CORONEL',    'OFICIAL_SUPERIOR'),
    'TCNL.':  ('TCNEL',      'OFICIAL_SUPERIOR'),
    'MY.':    ('MAYOR',      'OFICIAL_SUPERIOR'),
    'CAP.':   ('CAPITAN',    'OFICIAL_SUBALTERNO'),
    'TTE.':   ('TENIENTE',   'OFICIAL_SUBALTERNO'),
    'SBTTE.': ('SUBTENIENTE','OFICIAL_SUBALTERNO'),
    'SLDO.':  ('SOLDADO',    'TROPA'),
    'CABO':   ('CABO',       'TROPA'),
}

# ─── Mapping arma abreviatura >> clave del modelo ─────────────────────────────
ARMA_MAP = {
    'INF.':    'INFANTERIA',
    'CAB.':    'CABALLERIA',
    'ART.':    'ARTILLERIA',
    'ING.':    'INGENIERIA',
    'COM.':    'COMUNICACIONES',
    'INT.':    'INTENDENCIA',
    'SAN.':    'SANIDAD',
    'TGRAFO.': 'TOPOGRAFIA',
    'AV.':     'AVIACION',
    'MUS.':    'MÚSICA',
}

# Tokens que siguen al grado pero NO son arma (se descartan silenciosamente)
TOKENS_SKIP = {'(R.O.)', '(+)', 'INCL.', 'M.B.', 'MOT.', 'LOG.',
               'DEM.', 'DAEN.', 'DEPSS.', 'GCOE.', 'GCOSE.'}

# Todos los posibles primeros tokens de grado (para dividir texto multi-persona)
GRADO_STARTERS = re.compile(
    r'(?<!\w)(GRAL\.|CNL\.|TCNL\.|MY\.|CAP\.|TTE\.|SBTTE\.'
    r'|SOF\.|SGTO\.|SLDO\.|CABO)(?!\w)'
)


# ─────────────────────────────────────────────────────────────────────────────
def parsear_fecha(valor) -> date | None:
    """Parsea ' DD-MMM-YY', datetime, date o None."""
    if valor is None:
        return None
    if isinstance(valor, datetime):
        return valor.date()
    if isinstance(valor, date):
        return valor
    texto = str(valor).strip()
    texto = re.sub(r'^FN:\s*', '', texto).strip()
    m = re.match(r'(\d{1,2})-([A-Z]{3})-(\d{2,4})$', texto)
    if m:
        dia, mes_str, anio = m.groups()
        mes = MESES.get(mes_str)
        if mes:
            anio = int(anio)
            if anio < 100:
                anio += 2000
            try:
                return date(anio, mes, int(dia))
            except ValueError:
                pass
    return None


def parsear_res_num(texto) -> str | None:
    """Extrae '004/26' de 'RES. TPE. N° 004/26'."""
    if not texto:
        return None
    m = re.search(r'N[°º\.]\s*(\d+/\d+)', str(texto))
    return m.group(1) if m else None


def es_vacio(valor) -> bool:
    if valor is None:
        return True
    return str(valor).strip() in ('', '\xa0', '_________', 'N/A')


def truncar(texto, largo) -> str:
    if not texto:
        return ''
    return str(texto).strip()[:largo]


# ─── Parser de persona: "GRADO [ARMA] NOMBRE PATERNO MATERNO" ────────────────
def parsear_una_persona(texto: str) -> dict | None:
    """
    Recibe una cadena como 'TCNL. INF. JUAN PEREZ FLORES' y retorna
    {grado, arma, escalafon, espec, nombre, paterno, materno}.
    """
    tokens = texto.split()
    if not tokens:
        return None

    grado = escalafon = arma = espec = None
    idx = 0
    t = tokens[idx]

    # Generales: GRAL. BRIG. / GRAL. DIV. / GRAL. EJ.
    if t == 'GRAL.':
        idx += 1
        sub = tokens[idx] if idx < len(tokens) else ''
        if sub in ('BRIG.', 'DIV.', 'EJ.', 'EJERCITO'):
            idx += 1
            grado = {'BRIG.': 'GENERAL_BRIGADA',
                     'DIV.': 'GENERAL_DIVISION'}.get(sub, 'GENERAL_EJERCITO')
        else:
            grado = 'GENERAL_BRIGADA'
        escalafon = 'GENERAL'

    # Suboficiales: SOF. MY. / 1RO. / 2DO. / INCL. / MTRE.
    elif t == 'SOF.':
        idx += 1
        sub = tokens[idx] if idx < len(tokens) else ''
        mapa = {'MY.': 'SUBOFICIAL_MAYOR', '1RO.': 'SUBOFICIAL_1RO',
                '2DO.': 'SUBOFICIAL_2DO', 'INCL.': 'SUBOFICIAL_INICIAL',
                'MTRE.': 'SUBOFICIAL_MAESTRE'}
        if sub in mapa:
            grado = mapa[sub]
            idx += 1
        else:
            grado = 'SUBOFICIAL_1RO'
        escalafon = 'SUBOFICIAL'

    # Sargentos: SGTO. 1RO. / 2DO. / INCL.
    elif t == 'SGTO.':
        idx += 1
        sub = tokens[idx] if idx < len(tokens) else ''
        mapa = {'1RO.': 'SARGENTO_1RO', '2DO.': 'SARGENTO_2DO',
                'INCL.': 'SARGENTO_INICIAL'}
        if sub in mapa:
            grado = mapa[sub]
            idx += 1
        else:
            grado = 'SARGENTO_INICIAL'
        escalafon = 'SARGENTO'

    elif t in GRADO_SIMPLE:
        grado, escalafon = GRADO_SIMPLE[t]
        idx += 1

    else:
        return None  # texto sin grado reconocible

    # Consumir tokens de arma / especialidad / skip
    while idx < len(tokens):
        tok = tokens[idx]
        if tok in TOKENS_SKIP:
            idx += 1
            continue
        if tok in ARMA_MAP:
            arma = ARMA_MAP[tok]
            idx += 1
            continue
        break  # ya llegamos al nombre

    # Resto = nombre(s) + apellidos
    resto = tokens[idx:]
    if not resto:
        return None

    if len(resto) == 1:
        nombre, paterno, materno = resto[0], '', None
    elif len(resto) == 2:
        nombre, paterno, materno = resto[0], resto[1], None
    else:
        materno = resto[-1]
        paterno = resto[-2]
        nombre = ' '.join(resto[:-2])

    return {
        'grado': grado,
        'arma': arma,
        'escalafon': escalafon,
        'espec': espec,
        'nombre': truncar(nombre, 25),
        'paterno': truncar(paterno, 25),
        'materno': truncar(materno, 25) if materno else None,
    }


def parsear_personas(texto: str) -> list[dict]:
    """Divide texto con múltiples militares y parsea cada uno."""
    if not texto or es_vacio(texto):
        return []
    texto = texto.replace('"', '').strip()

    # Dividir en bloques por aparición de grado
    segmentos = GRADO_STARTERS.split(texto)
    resultado = []
    i = 1
    while i < len(segmentos):
        grado_tok = segmentos[i].strip()
        resto = segmentos[i + 1].strip() if i + 1 < len(segmentos) else ''
        texto_persona = grado_tok + ' ' + ' '.join(resto.split())
        persona = parsear_una_persona(texto_persona)
        if persona and persona['paterno']:
            resultado.append(persona)
        i += 2

    if not resultado:
        persona = parsear_una_persona(' '.join(texto.split()))
        if persona and persona['paterno']:
            resultado.append(persona)

    return resultado


# ─── Determinar tipo de SIM / RES ─────────────────────────────────────────────
def tipo_sim(objeto: str, desc: str = '') -> str:
    t = (objeto + ' ' + desc).upper()
    if 'ASCENSO POSTUMO' in t or 'POSTUMO' in t:
        return 'ASCENSO POSTUMO'
    if 'LETRA D' in t or 'LETRA "D"' in t:
        return 'SOLICITUD_LETRA_D'
    if 'LICENCIA MAXIMA' in t:
        return 'SOLICITUD_LICENCIA_MAXIMA'
    if 'RESTITUCION' in t and 'ANTIGUEDAD' in t:
        return 'SOLICITUD_RESTITUCION_ANTIGUEDAD'
    if 'DERECHOS PROFESIONALES' in t:
        return 'SOLICITUD_DE_RESTITUCION_DE_DERECHOS_PROFESIONALES'
    if 'ASCENSO' in t:
        return 'SOLICITUD_ASCENSO_AL_GRADO_INMEDIATO_SUPERIOR'
    if 'ADMINISTRATIVO' in t:
        return 'ADMINISTRATIVO'
    return 'DISCIPLINARIO'


def tipo_res(resol: str, desc: str = '') -> str:
    t = (resol + ' ' + desc).upper()
    if 'RETIRO OBLIGATORIO' in t:
        return 'SANCION_RETIRO_OBLIGATORIO'
    if 'LETRA "B"' in t or "LETRA B" in t:
        return 'SANCION_LETRA_B'
    if 'ARRESTO' in t:
        return 'SANCION_ARRESTO'
    if 'BAJA' in t:
        return 'SANCION_BAJA'
    if 'ARCHIVO' in t:
        return 'ARCHIVO_OBRADOS'
    if 'ASCENSO' in t:
        return 'SOLICITUD_ASCENSO'
    if 'POSESION' in t or 'ADMINISTRATIVO' in t:
        return 'ADMINISTRATIVO'
    if 'DISCIPLINARIA' in t or 'SANCION' in t:
        return 'SANCIONES_DISCIPLINARIAS'
    return 'OTRO'


def tipo_autotpe(resol: str) -> str:
    t = resol.upper()
    if 'SOBRESEIDO' in t:
        return 'SOBRESEIDO'
    if 'NULIDAD' in t:
        return 'NULIDAD_OBRADOS'
    if 'ARRESTO' in t or 'EJECUTIVA' in t or 'EJECUTORIA' in t:
        return 'AUTO_EJECUTORIA'
    if 'LETRA "B"' in t or 'LETRA B' in t:
        return 'AUTO_CUMPLIMIENTO'
    if 'EXCUSA' in t:
        return 'AUTO_EXCUSA'
    if 'RECHAZO' in t:
        return 'AUTO_RECHAZO_RECURSO'
    return 'AUTO_CUMPLIMIENTO'


# ─── Comando principal ────────────────────────────────────────────────────────
class Command(BaseCommand):
    help = 'Importa datos reales desde el archivo Excel tpe2026.xlsx'

    def add_arguments(self, parser):
        parser.add_argument(
            '--archivo', type=str, default='docs/tpe2026.xlsx',
            help='Ruta al archivo Excel (relativa al directorio del proyecto)',
        )
        parser.add_argument(
            '--limpiar', action='store_true',
            help='Limpia los datos existentes antes de importar',
        )
        parser.add_argument(
            '--dry-run', action='store_true',
            help='Simula la importación sin guardar nada en la base de datos',
        )

    def handle(self, *args, **options):
        self.dry = options['dry_run']
        self.stats = {k: 0 for k in ('pm', 'sim', 'agenda', 'res', 'auto', 'rr', 'rap', 'skip')}

        if self.dry:
            self.stdout.write(self.style.WARNING('[DRY-RUN] No se guardan cambios'))

        wb = load_workbook(options['archivo'], data_only=True)

        if options['limpiar'] and not self.dry:
            self._limpiar()

        try:
            with transaction.atomic():
                self._importar_sim(wb['SIM.'])
                self._importar_casos_especiales(wb['CASOS ESPECIALES'])
                self._importar_resoluciones(wb['RESOLUCIONES'])
                self._importar_autos(wb['AUTOS'])
                self._importar_rr(wb['REC. RECONSIDERACION'])
                self._importar_rap(wb['REC. APELACION'])
                if self.dry:
                    raise RuntimeError('dry-run: rollback intencional')
        except RuntimeError as e:
            if 'dry-run' in str(e):
                self.stdout.write(self.style.WARNING('Rollback aplicado (dry-run)'))
            else:
                raise

        wb.close()
        self._resumen()

    # ── Limpieza ──────────────────────────────────────────────────────────────
    def _limpiar(self):
        self.stdout.write(self.style.WARNING('Limpiando datos existentes...'))
        RAEE.objects.all().delete()
        AUTOTSP.objects.all().delete()
        RAP.objects.all().delete()
        RR.objects.all().delete()
        AUTOTPE.objects.all().delete()
        DICTAMEN.objects.all().delete()
        RES.objects.all().delete()
        AGENDA.objects.all().delete()
        PM_SIM.objects.all().delete()
        ABOG_SIM.objects.all().delete()
        SIM.objects.all().delete()
        PM.objects.all().delete()
        self.stdout.write('  Datos borrados.')

    # ── Helpers ───────────────────────────────────────────────────────────────
    def _pm(self, persona: dict) -> PM | None:
        """Obtiene o crea un PM a partir de los datos parseados."""
        if not persona or not persona.get('paterno'):
            return None
        nombre  = persona['nombre'].upper()
        paterno = persona['paterno'].upper()
        materno = persona['materno'].upper() if persona.get('materno') else None

        qs = PM.objects.filter(PM_NOMBRE=nombre, PM_PATERNO=paterno)
        if materno:
            qs = qs.filter(PM_MATERNO=materno)
        if qs.exists():
            return qs.first()

        if self.dry:
            self.stats['pm'] += 1
            return None

        pm = PM(
            PM_NOMBRE=nombre, PM_PATERNO=paterno, PM_MATERNO=materno,
            PM_GRADO=persona.get('grado'),
            PM_ARMA=persona.get('arma'),
            PM_ESCALAFON=persona.get('escalafon'),
            PM_ESPEC=persona.get('espec'),
            PM_ESTADO='ACTIVO',
        )
        pm.save()
        self.stats['pm'] += 1
        return pm

    def _vincular_personas(self, sim: SIM, texto: str):
        for p in parsear_personas(texto):
            pm = self._pm(p)
            if pm and not self.dry:
                PM_SIM.objects.get_or_create(sim=sim, pm=pm)

    def _agenda(self, nombre: str, fecha: date | None = None) -> AGENDA | None:
        if not nombre or es_vacio(nombre):
            return None
        num = truncar(nombre, 50).upper()
        if self.dry:
            self.stats['agenda'] += 1
            return None
        ag, creada = AGENDA.objects.get_or_create(
            AG_NUM=num,
            defaults={'AG_FECPROG': fecha, 'AG_TIPO': 'ORDINARIA'},
        )
        if creada:
            self.stats['agenda'] += 1
        elif fecha and not ag.AG_FECREAL:
            ag.AG_FECREAL = fecha
            ag.save(update_fields=['AG_FECREAL'])
        return ag

    def _sim_placeholder(self, cod: str, objeto: str, estado: str = 'CONCLUIDO',
                         nombres_texto: str = '') -> SIM:
        """Crea un SIM mínimo para satisfacer FK cuando el original es histórico."""
        cod = truncar(cod.upper(), 10)
        if not self.dry:
            sim, creada = SIM.objects.get_or_create(
                SIM_COD=cod,
                defaults={
                    'SIM_OBJETO': truncar(objeto, 500) or 'REFERENCIA HISTÓRICA',
                    'SIM_RESUM': truncar(objeto, 200) or 'REFERENCIA HISTÓRICA',
                    'SIM_TIPO': tipo_sim(objeto),
                    'SIM_ESTADO': estado,
                },
            )
            if creada:
                self.stats['sim'] += 1
                if nombres_texto:
                    self._vincular_personas(sim, nombres_texto)
            return sim
        self.stats['sim'] += 1
        return None

    # ── Hoja SIM. ────────────────────────────────────────────────────────────
    def _importar_sim(self, ws):
        self.stdout.write('>> Importando SIM...')
        # Cabecera en filas 4-5, datos desde fila 6
        # Cols: A=N°, B=FECHA, C=COD, D=GRADO/NOMBRES, E=OBJETO, F=AGENDA, G=FEC_AGENDA, H=RESP, I=DESC
        for row in ws.iter_rows(min_row=6, values_only=True):
            if not row[0] or not isinstance(row[0], (int, float)):
                continue
            cod = truncar(row[2], 10).upper() if row[2] else None
            if not cod:
                self.stats['skip'] += 1
                continue

            objeto  = truncar(row[4], 2000) if row[4] else 'SIN OBJETO'
            desc    = truncar(row[8], 200)  if row[8] else ''
            resum   = desc or truncar(objeto, 200)
            fecing  = parsear_fecha(row[1])
            ag_nom  = str(row[5]).strip() if row[5] and not es_vacio(row[5]) else None
            ag_fec  = parsear_fecha(row[6])
            estado  = 'PROCESO_EN_EL_TPE' if ag_nom else 'PARA_AGENDA'

            if self.dry:
                self.stats['sim'] += 1
                self._vincular_personas(None, str(row[3]) if row[3] else '')
                if ag_nom:
                    self._agenda(ag_nom, ag_fec)
                continue

            sim, creada = SIM.objects.get_or_create(
                SIM_COD=cod,
                defaults={
                    'SIM_FECING': fecing, 'SIM_ESTADO': estado,
                    'SIM_OBJETO': objeto, 'SIM_RESUM': resum,
                    'SIM_TIPO': tipo_sim(objeto, desc),
                },
            )
            if creada:
                self.stats['sim'] += 1
            self._vincular_personas(sim, str(row[3]) if row[3] else '')
            if ag_nom:
                self._agenda(ag_nom, ag_fec)

    # ── Hoja CASOS ESPECIALES ─────────────────────────────────────────────────
    def _importar_casos_especiales(self, ws):
        self.stdout.write('>> Importando CASOS ESPECIALES...')
        # Cols: A=N°, B=FECHA, C=GRADO/NOMBRES, D=SOLICITUD, E=AGENDA, F=FEC, G=RESP, H=OBS
        contador = 0
        for row in ws.iter_rows(min_row=5, values_only=True):
            if not row[0] or not isinstance(row[0], (int, float)):
                continue
            contador += 1
            objeto = truncar(row[3], 2000) if row[3] else 'SIN DESCRIPCION'
            fecing = parsear_fecha(row[1])
            cod    = f'CE-{contador:03d}/26'  # max 10 chars ✓

            if self.dry:
                self.stats['sim'] += 1
                self._vincular_personas(None, str(row[2]) if row[2] else '')
                continue

            sim, creada = SIM.objects.get_or_create(
                SIM_COD=cod,
                defaults={
                    'SIM_FECING': fecing, 'SIM_ESTADO': 'PARA_AGENDA',
                    'SIM_OBJETO': objeto, 'SIM_RESUM': truncar(objeto, 200),
                    'SIM_TIPO': tipo_sim(objeto),
                },
            )
            if creada:
                self.stats['sim'] += 1
            self._vincular_personas(sim, str(row[2]) if row[2] else '')

    # ── Hoja RESOLUCIONES ────────────────────────────────────────────────────
    def _importar_resoluciones(self, ws):
        self.stdout.write('>> Importando RESOLUCIONES...')
        # Cols: A=N°, B=SIM_COD, C=GRADO/NOMBRES, D=N°REUNION, E=N°RES, F=FECHA_RES,
        #       G=MOTIVO, H=PARTE RESOLUTIVA, I=FECHA_NOTIF, J=VENCE, K=ENVIO,
        #       L=N°OF/AUTO, M=ESTADO, N=DESC_SANCION
        _cont_esp = {'POSESION': 0, 'INFORME DE NECESIDAD': 0, 'SOLICITUD': 0}
        prefijos  = {'POSESION': 'POS', 'INFORME DE NECESIDAD': 'IN', 'SOLICITUD': 'SOL'}

        for row in ws.iter_rows(min_row=5, values_only=True):
            res_num_raw = str(row[4]).strip() if row[4] else ''
            if not res_num_raw:
                continue
            res_fec = parsear_fecha(row[5])
            if not res_fec:
                continue  # resolución sin fecha = fila vacía

            res_num = parsear_res_num(res_num_raw) or truncar(res_num_raw, 15)
            resol   = truncar(row[7], 5000) if row[7] else ''
            desc    = truncar(row[13], 500) if len(row) > 13 and row[13] else ''

            # ── Encontrar o crear SIM ────────────────────────────────────────
            sim_cod_raw = str(row[1]).strip() if row[1] else ''
            sim = None

            if re.match(r'DJE-', sim_cod_raw, re.I):
                # SIM disciplinario con código DJE
                try:
                    sim = SIM.objects.get(SIM_COD=sim_cod_raw.upper())
                except SIM.DoesNotExist:
                    # Caso histórico (2023-2025) no en hoja SIM.
                    objeto_ph = truncar(row[6], 500) if row[6] else 'SIM HISTÓRICO'
                    sim = self._sim_placeholder(
                        sim_cod_raw, objeto_ph, 'CONCLUIDO',
                        str(row[2]) if row[2] else '',
                    )
            elif sim_cod_raw:
                # Especiales: POSESION, INFORME DE NECESIDAD, SOLICITUD
                clave = sim_cod_raw.upper()
                _cont_esp[clave] = _cont_esp.get(clave, 0) + 1
                pref  = prefijos.get(clave, 'ESP')
                cod   = f'{pref}-{_cont_esp[clave]:02d}/26'  # max 10 ✓
                objeto_ph = truncar(row[6], 500) if row[6] else clave
                sim = self._sim_placeholder(cod, objeto_ph, 'CONCLUIDO',
                                            str(row[2]) if row[2] else '')
            else:
                self.stats['skip'] += 1
                continue

            if sim is None and not self.dry:
                self.stats['skip'] += 1
                continue

            # ── Agenda ──────────────────────────────────────────────────────
            ag = self._agenda(str(row[3]).strip() if row[3] else None, res_fec)

            # ── Actualizar estado del SIM según columna M ────────────────────
            estado_excel = str(row[12]).strip().upper() if row[12] else ''
            if estado_excel == 'CONCLUIDO' and sim and not self.dry:
                sim.SIM_ESTADO = 'CONCLUIDO'
                sim.save(update_fields=['SIM_ESTADO'])

            # ── Crear RES ────────────────────────────────────────────────────
            if self.dry:
                self.stats['res'] += 1
                continue

            fec_notif = parsear_fecha(row[8]) if not es_vacio(row[8]) else None
            _, creada = RES.objects.get_or_create(
                sim=sim,
                RES_NUM=truncar(res_num, 15),
                defaults={
                    'RES_FEC': res_fec,
                    'RES_RESOL': resol or desc or 'SIN TEXTO',
                    'RES_TIPO': tipo_res(resol, desc),
                    'agenda': ag,
                    'RES_FECNOT': fec_notif,
                },
            )
            if creada:
                self.stats['res'] += 1

    # ── Hoja AUTOS ───────────────────────────────────────────────────────────
    def _importar_autos(self, ws):
        self.stdout.write('>> Importando AUTOS TPE...')
        # Cols: A=N°, B=N°RES_EJEC, C=GRADO/NOMBRES, D=N°AUTO, E=FECHA,
        #       F=PARTE_RESOLUTIVA, G=DESC, H=FECHA_NOTIF, I=SITUACION, J=ESTADO
        for row in ws.iter_rows(min_row=3, values_only=True):
            if not row[0] or not isinstance(row[0], (int, float)):
                continue
            auto_num = truncar(row[3], 20) if row[3] else None
            if not auto_num:
                continue
            auto_fec = parsear_fecha(row[4])
            if not auto_fec:
                continue

            resol      = truncar(row[5], 3000) if row[5] else ''
            nombres    = str(row[2]).strip() if row[2] else ''
            res_ejec   = str(row[1]).strip() if row[1] else ''

            # ── Encontrar SIM ────────────────────────────────────────────────
            sim = None

            # 1) Por resolución ejecutada
            if re.match(r'RES\.', res_ejec, re.I):
                num = parsear_res_num(res_ejec)
                if num:
                    try:
                        sim = RES.objects.get(RES_NUM=num).sim
                    except RES.DoesNotExist:
                        pass

            # 2) Por nombre del militar
            if not sim and nombres:
                for p in parsear_personas(nombres):
                    if p and p.get('paterno'):
                        qs = PM_SIM.objects.filter(
                            pm__PM_PATERNO=p['paterno'].upper()
                        ).select_related('sim')
                        if qs.exists():
                            sim = qs.first().sim
                            break

            # 3) Placeholder
            if not sim:
                num_corto = parsear_res_num(auto_num) or auto_num[-6:]
                cod = 'AT-' + re.sub(r'[^0-9/]', '', num_corto)[-7:]  # max 10
                sim = self._sim_placeholder(
                    cod, nombres or auto_num, 'CONCLUIDO', nombres,
                )

            if sim is None and not self.dry:
                self.stats['skip'] += 1
                continue

            num_corto = parsear_res_num(auto_num) or truncar(auto_num, 15)

            if self.dry:
                self.stats['auto'] += 1
                continue

            _, creada = AUTOTPE.objects.get_or_create(
                sim=sim,
                TPE_NUM=truncar(num_corto, 15),
                defaults={
                    'TPE_FEC': auto_fec,
                    'TPE_RESOL': resol,
                    'TPE_TIPO': tipo_autotpe(resol),
                },
            )
            if creada:
                self.stats['auto'] += 1

    # ── Hoja REC. RECONSIDERACION ─────────────────────────────────────────────
    def _importar_rr(self, ws):
        self.stdout.write('>> Importando REC. RECONSIDERACION...')
        # Cols: A=N°, B=GRADO/NOMBRES, C=(CONTRA)N°RES, D=MOTIVO, E=PARTE_RESOL,
        #       F=FECHA_NOTIF_RES, G=FECHA_PRES_RR, H=FECHA_LIMITE, I=TRATADO_AGENDA, J=ENCARGADO
        for row in ws.iter_rows(min_row=3, values_only=True):
            if not row[0] or not isinstance(row[0], (int, float)):
                continue
            nombres   = str(row[1]).strip() if row[1] else ''
            res_ref   = str(row[2]).strip() if row[2] else ''
            resol     = truncar(row[4], 3000) if row[4] else ''
            fec_pres  = parsear_fecha(row[6])
            fec_lim   = parsear_fecha(row[7])

            if not fec_pres:
                self.stats['skip'] += 1
                continue

            # ── Encontrar RES ────────────────────────────────────────────────
            res_obj = None
            num = parsear_res_num(res_ref)
            if num:
                try:
                    res_obj = RES.objects.get(RES_NUM=num)
                except RES.DoesNotExist:
                    pass

            if not res_obj:
                # Crear SIM + RES placeholder para referencia histórica 2025
                cod_sim = 'RP-' + re.sub(r'[^0-9/]', '', num or '')[-7:] if num else f'RP-{row[0]:03d}/25'
                sim_ph  = self._sim_placeholder(cod_sim, nombres or res_ref, 'CONCLUIDO', nombres)
                if sim_ph is None and not self.dry:
                    self.stats['skip'] += 1
                    continue
                if not self.dry:
                    res_obj, _ = RES.objects.get_or_create(
                        sim=sim_ph,
                        RES_NUM=truncar(num or res_ref, 15),
                        defaults={
                            'RES_FEC': fec_pres,
                            'RES_RESOL': resol or 'RESOLUCION HISTORICA 2025',
                            'RES_TIPO': tipo_res(resol),
                        },
                    )

            if self.dry:
                self.stats['rr'] += 1
                continue

            sim = res_obj.sim
            ag  = self._agenda(str(row[8]).strip() if row[8] else None)

            _, creada = RR.objects.get_or_create(
                res=res_obj,
                sim=sim,
                defaults={
                    'RR_FECPRESEN': fec_pres,
                    'RR_FECLIMITE': fec_lim,
                    'RR_RESOL': resol,
                    'RR_RESUM': truncar(resol, 200),
                    'agenda': ag,
                },
            )
            if creada:
                self.stats['rr'] += 1

    # ── Hoja REC. APELACION ───────────────────────────────────────────────────
    def _importar_rap(self, ws):
        self.stdout.write('>> Importando REC. APELACION...')
        # Cols: A=N°, B=GRADO/NOMBRES, C=(CONTRA)N°RES, D=MOTIVO, E=PARTE_RESOL,
        #       F=FECHA_NOTIF, G=FECHA_PRES, H=N°OFICIO, I=FECHA_OFICIO, J=RESPONSABLE
        for row in ws.iter_rows(min_row=3, values_only=True):
            if not row[0] or not isinstance(row[0], (int, float)):
                continue
            nombres   = str(row[1]).strip() if row[1] else ''
            res_ref   = str(row[2]).strip() if row[2] else ''
            resol     = truncar(row[4], 3000) if row[4] else ''
            fec_pres  = parsear_fecha(row[6])
            ofi_num   = truncar(row[7], 25) if row[7] else None
            ofi_fec   = parsear_fecha(row[8])

            if not fec_pres:
                self.stats['skip'] += 1
                continue

            # ── Encontrar SIM (via RR -> RES -> SIM o placeholder) ───────────
            sim = None
            rr_obj = None
            num = parsear_res_num(res_ref)
            if num:
                try:
                    rr_obj = RR.objects.get(res__RES_NUM=num)
                    sim    = rr_obj.sim
                except RR.DoesNotExist:
                    try:
                        sim = RES.objects.get(RES_NUM=num).sim
                    except RES.DoesNotExist:
                        pass

            if not sim:
                cod_sim = 'RA-' + re.sub(r'[^0-9/]', '', num or '')[-7:] if num else f'RA-{row[0]:03d}/25'
                sim = self._sim_placeholder(cod_sim, nombres or res_ref, 'CONCLUIDO', nombres)

            if sim is None and not self.dry:
                self.stats['skip'] += 1
                continue

            if self.dry:
                self.stats['rap'] += 1
                continue

            _, creada = RAP.objects.get_or_create(
                sim=sim,
                RAP_FECPRESEN=fec_pres,
                defaults={
                    'rr': rr_obj,
                    'RAP_OFI': ofi_num,
                    'RAP_FECOFI': ofi_fec,
                    'RAP_RESOL': resol,
                },
            )
            if creada:
                self.stats['rap'] += 1

    # ── Resumen ───────────────────────────────────────────────────────────────
    def _resumen(self):
        s = self.stats
        self.stdout.write(self.style.SUCCESS(
            '\nImportacion completada:\n'
            f'  PM creados          : {s["pm"]}\n'
            f'  SIM importados      : {s["sim"]}\n'
            f'  Agendas             : {s["agenda"]}\n'
            f'  Resoluciones (RES)  : {s["res"]}\n'
            f'  Autos TPE           : {s["auto"]}\n'
            f'  Rec. Reconsideracion: {s["rr"]}\n'
            f'  Rec. Apelacion      : {s["rap"]}\n'
            f'  Filas saltadas      : {s["skip"]}\n'
        ))
