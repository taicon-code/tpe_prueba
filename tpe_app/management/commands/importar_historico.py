"""
Management command para importar el histórico de sumarios desde Excel.

Uso:
    python manage.py importar_historico plantilla_historico_sim.xlsx
    python manage.py importar_historico plantilla_historico_sim.xlsx --dry-run
    python manage.py importar_historico plantilla_historico_sim.xlsx --desde-hoja 4_RES

Opciones:
    --dry-run        Valida el archivo sin insertar nada en la base de datos
    --desde-hoja     Nombre de la hoja desde donde continuar (ej: 4_RES)
    --log            Ruta del archivo de log (default: importacion_historico.log)

Nombres de columna esperados en el Excel (primera línea del encabezado):
  Hoja 1_SIM:    codigo, fecha_ingreso, estado, tipo, objeto, resumen, auto_final, numero_carpeta
  Hoja 2_PM:     ci, escalafon, grado, arma, especialidad, nombre, paterno, materno, estado, anio_promocion
  Hoja 3_PM_SIM: codigo, ci, grado_en_fecha
  Hoja 4_RES:    codigo, numero, fecha, tipo, texto, tipo_notif, notif_a, fecha_notif, hora_notif
  Hoja 5_RR:     codigo, res_numero, fecha_presentacion, numero, fecha, texto, tipo,
                 tipo_notif, notif_a, fecha_notif, hora_notif
  Hoja 6_RAP:    codigo, fecha_presentacion, numero_oficio, fecha_oficio, numero, fecha, tipo, texto,
                 tipo_notif, notif_a, fecha_notif, hora_notif
  Hoja 7_AUTOTPE: codigo, numero, fecha, tipo, texto, tipo_notif, notif_a, fecha_notif, hora_notif,
                  memo_numero, memo_fecha, memo_fecha_entrega
  Hoja 8_RAEE:   codigo, numero, fecha, texto, tipo_notif, notif_a, fecha_notif, hora_notif
"""

import logging
from datetime import date, time, datetime
from pathlib import Path

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from openpyxl import load_workbook

from tpe_app.models import (
    PM, SIM, PM_SIM,
    AUTOTPE, Resolucion, RecursoTSP,
    Notificacion, Memorandum,
)


# ─── Helpers ─────────────────────────────────────────────────────────────────

def _parse_fecha(valor) -> date | None:
    if valor is None or str(valor).strip() == "":
        return None
    if isinstance(valor, (date, datetime)):
        return valor.date() if isinstance(valor, datetime) else valor
    texto = str(valor).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(texto, fmt).date()
        except ValueError:
            continue
    raise ValueError(f"Fecha no reconocida: '{texto}'")


def _parse_hora(valor) -> time | None:
    if valor is None or str(valor).strip() == "":
        return None
    if isinstance(valor, time):
        return valor
    if isinstance(valor, datetime):
        return valor.time()
    texto = str(valor).strip()
    for fmt in ("%H:%M", "%H:%M:%S"):
        try:
            return datetime.strptime(texto, fmt).time()
        except ValueError:
            continue
    raise ValueError(f"Hora no reconocida: '{texto}'")


def _str(valor) -> str | None:
    if valor is None:
        return None
    s = str(valor).strip()
    return s if s else None


def _ci(valor):
    if valor is None or str(valor).strip() == "":
        return None
    try:
        return int(float(str(valor).strip()))
    except (ValueError, TypeError):
        return None


def _anio(valor) -> int | None:
    """Convierte un año a entero. Acepta '1998', '1998-01-01', etc."""
    if valor is None or str(valor).strip() == "":
        return None
    texto = str(valor).strip()
    try:
        # Si viene como fecha completa, tomar solo el año
        if len(texto) >= 4 and texto[4] in ("-", "/"):
            return int(texto[:4])
        return int(float(texto))
    except (ValueError, TypeError):
        return None


def _iter_filas(ws, encabezado_fila=2, datos_desde=3):
    """
    Genera dicts {nombre_columna: valor} para cada fila de datos.
    El nombre de columna es la primera línea de cada celda del encabezado.
    Salta filas completamente vacías.
    """
    headers = []
    for cell in ws[encabezado_fila]:
        h = str(cell.value or "").split("\n")[0].strip()
        headers.append(h)

    for row in ws.iter_rows(min_row=datos_desde, values_only=True):
        if all(v is None or str(v).strip() == "" for v in row):
            continue
        yield dict(zip(headers, row))


# ─── Command ─────────────────────────────────────────────────────────────────

class Command(BaseCommand):
    help = "Importa el histórico de sumarios SIM desde la plantilla Excel v4.1."

    def add_arguments(self, parser):
        parser.add_argument(
            "archivo",
            type=str,
            help="Ruta al archivo Excel (plantilla_historico_sim.xlsx)",
        )
        parser.add_argument(
            "--dry-run",
            action="store_true",
            default=False,
            help="Valida sin insertar datos.",
        )
        parser.add_argument(
            "--desde-hoja",
            type=str,
            default=None,
            help="Nombre de la hoja desde donde continuar (ej: 4_RES).",
        )
        parser.add_argument(
            "--log",
            type=str,
            default="importacion_historico.log",
            help="Ruta del archivo de log (default: importacion_historico.log).",
        )

    def handle(self, *args, **options):
        archivo  = options["archivo"]
        dry_run  = options["dry_run"]
        desde    = options["desde_hoja"]
        log_path = options["log"]

        logging.basicConfig(
            filename=log_path,
            filemode="a",
            encoding="utf-8",
            format="%(asctime)s [%(levelname)s] %(message)s",
            level=logging.DEBUG,
        )
        self.log = logging.getLogger(__name__)

        if not Path(archivo).exists():
            raise CommandError(f"Archivo no encontrado: {archivo}")

        self.stdout.write(f"\nCargando: {archivo}")
        if dry_run:
            self.stdout.write(self.style.WARNING("  MODO DRY-RUN — no se insertará nada\n"))

        wb = load_workbook(archivo, data_only=True)

        self.ok      = {}
        self.skip    = {}
        self.errores = []

        pasos = [
            ("1_SIM",     self._importar_sim),
            ("2_PM",      self._importar_pm),
            ("3_PM_SIM",  self._importar_pm_sim),
            ("4_RES",     self._importar_res),
            ("5_RR",      self._importar_rr),
            ("6_RAP",     self._importar_rap),
            ("7_AUTOTPE", self._importar_autotpe),
            ("8_RAEE",    self._importar_raee),
        ]

        activo = desde is None
        for nombre_hoja, fn in pasos:
            if not activo:
                if nombre_hoja == desde:
                    activo = True
                else:
                    self.stdout.write(f"  [SALTADA] {nombre_hoja}")
                    continue

            if nombre_hoja not in wb.sheetnames:
                self.stdout.write(self.style.WARNING(f"  [NO ENCONTRADA] {nombre_hoja}"))
                continue

            ws = wb[nombre_hoja]
            self.stdout.write(f"\n  Procesando hoja: {nombre_hoja} ...")

            try:
                if dry_run:
                    with transaction.atomic():
                        fn(ws, nombre_hoja)
                        transaction.set_rollback(True)
                else:
                    with transaction.atomic():
                        fn(ws, nombre_hoja)
            except Exception as exc:
                self.stdout.write(self.style.ERROR(f"    ERROR CRÍTICO en {nombre_hoja}: {exc}"))
                self.log.exception(f"ERROR CRÍTICO en {nombre_hoja}")

        self._imprimir_resumen(dry_run, log_path)

    def _imprimir_resumen(self, dry_run, log_path):
        self.stdout.write("\n" + "=" * 60)
        self.stdout.write("RESUMEN DE IMPORTACIÓN")
        if dry_run:
            self.stdout.write(self.style.WARNING("  (DRY-RUN — ningún dato fue guardado)"))
        self.stdout.write("=" * 60)

        total_ok = total_skip = 0
        for hoja in self.ok:
            ok   = self.ok.get(hoja, 0)
            skip = self.skip.get(hoja, 0)
            total_ok   += ok
            total_skip += skip
            self.stdout.write(f"  {hoja:<14}  insertados: {ok:>5}   omitidos: {skip:>5}")

        self.stdout.write("-" * 60)
        self.stdout.write(f"  TOTAL           insertados: {total_ok:>5}   omitidos: {total_skip:>5}")

        if self.errores:
            self.stdout.write(self.style.ERROR(f"\n  ERRORES ({len(self.errores)}):"))
            for hoja, num_fila, msg in self.errores[:20]:
                self.stdout.write(self.style.ERROR(f"    [{hoja} fila {num_fila}] {msg}"))
            if len(self.errores) > 20:
                self.stdout.write(self.style.ERROR(
                    f"    ... y {len(self.errores)-20} más. Ver: {log_path}"
                ))
        else:
            self.stdout.write(self.style.SUCCESS("\n  Sin errores."))

        self.stdout.write(f"\n  Log completo: {log_path}\n")

    def _registrar_ok(self, hoja):
        self.ok[hoja] = self.ok.get(hoja, 0) + 1

    def _registrar_skip(self, hoja, num_fila, motivo):
        self.skip[hoja] = self.skip.get(hoja, 0) + 1
        self.log.warning(f"[{hoja} fila {num_fila}] OMITIDA: {motivo}")

    def _registrar_error(self, hoja, num_fila, exc):
        self.errores.append((hoja, num_fila, str(exc)))
        self.log.error(f"[{hoja} fila {num_fila}] {exc}", exc_info=True)

    # ═══════════════════════════════════════════════════════════════════
    # IMPORTADORES POR HOJA
    # ═══════════════════════════════════════════════════════════════════

    # ── Hoja 1: SIM ──────────────────────────────────────────────────

    def _importar_sim(self, ws, hoja):
        for num_fila, fila in enumerate(_iter_filas(ws), start=3):
            try:
                cod = _str(fila.get("codigo"))
                if not cod:
                    self._registrar_skip(hoja, num_fila, "codigo vacío"); continue

                objeto = _str(fila.get("objeto"))
                resum  = _str(fila.get("resumen"))
                tipo   = _str(fila.get("tipo"))
                if not objeto:
                    self._registrar_skip(hoja, num_fila, f"{cod}: objeto vacío"); continue
                if not resum:
                    self._registrar_skip(hoja, num_fila, f"{cod}: resumen vacío"); continue
                if not tipo:
                    self._registrar_skip(hoja, num_fila, f"{cod}: tipo vacío"); continue

                numero_carpeta_raw = fila.get("numero_carpeta")
                numero_carpeta = None
                if numero_carpeta_raw is not None and str(numero_carpeta_raw).strip():
                    try:
                        numero_carpeta = int(float(str(numero_carpeta_raw).strip()))
                    except (ValueError, TypeError):
                        pass

                _, creado = SIM.objects.get_or_create(
                    codigo=cod.upper(),
                    version=1,
                    defaults=dict(
                        fecha_ingreso=_parse_fecha(fila.get("fecha_ingreso")),
                        estado=_str(fila.get("estado")) or "PARA_AGENDA",
                        objeto=objeto.upper(),
                        resumen=resum.upper()[:200],
                        auto_final=(_str(fila.get("auto_final")) or "").upper() or None,
                        tipo=tipo.upper(),
                        numero_carpeta=numero_carpeta,
                    ),
                )
                if creado:
                    self._registrar_ok(hoja)
                else:
                    self._registrar_skip(hoja, num_fila, f"{cod}: ya existe, omitido")
            except Exception as exc:
                self._registrar_error(hoja, num_fila, exc)

    # ── Hoja 2: PM ───────────────────────────────────────────────────

    def _importar_pm(self, ws, hoja):
        for num_fila, fila in enumerate(_iter_filas(ws), start=3):
            try:
                nombre  = _str(fila.get("nombre"))
                paterno = _str(fila.get("paterno"))
                ci      = _ci(fila.get("ci"))

                if not nombre or not paterno:
                    self._registrar_skip(hoja, num_fila, "nombre o paterno vacíos"); continue

                defaults = dict(
                    nombre=nombre.upper(),
                    paterno=paterno.upper(),
                    materno=(_str(fila.get("materno")) or "").upper() or None,
                    escalafon=_str(fila.get("escalafon")),
                    grado=_str(fila.get("grado")),
                    arma=_str(fila.get("arma")),
                    especialidad=(_str(fila.get("especialidad")) or "").upper() or None,
                    estado=_str(fila.get("estado")) or "ACTIVO",
                    anio_promocion=_anio(fila.get("anio_promocion")),
                )

                if ci:
                    pm, creado = PM.objects.get_or_create(ci=ci, defaults=defaults)
                else:
                    pm, creado = PM.objects.get_or_create(
                        nombre=nombre.upper(),
                        paterno=paterno.upper(),
                        defaults=defaults,
                    )

                if creado:
                    self._registrar_ok(hoja)
                else:
                    self._registrar_skip(hoja, num_fila, f"{nombre} {paterno}: ya existe, omitido")
            except Exception as exc:
                self._registrar_error(hoja, num_fila, exc)

    # ── Hoja 3: PM_SIM ───────────────────────────────────────────────

    def _importar_pm_sim(self, ws, hoja):
        for num_fila, fila in enumerate(_iter_filas(ws), start=3):
            try:
                cod = _str(fila.get("codigo"))
                ci  = _ci(fila.get("ci"))

                if not cod or not ci:
                    self._registrar_skip(hoja, num_fila, "codigo o ci vacíos"); continue

                sim = SIM.objects.filter(codigo=cod.upper(), version=1).first()
                if not sim:
                    self._registrar_skip(hoja, num_fila, f"SIM no encontrado: {cod}"); continue

                try:
                    pm = PM.objects.get(ci=ci)
                except PM.DoesNotExist:
                    self._registrar_skip(hoja, num_fila, f"PM no encontrado con CI: {ci}"); continue

                grado_en_fecha = _str(fila.get("grado_en_fecha"))

                pm_sim, creado = PM_SIM.objects.get_or_create(
                    sim=sim,
                    pm=pm,
                    defaults=dict(grado_en_fecha=grado_en_fecha),
                )
                if creado:
                    self._registrar_ok(hoja)
                else:
                    self._registrar_skip(hoja, num_fila, f"{cod}+{ci}: relación ya existe")
            except Exception as exc:
                self._registrar_error(hoja, num_fila, exc)

    # ── Hoja 4: RES ──────────────────────────────────────────────────

    def _importar_res(self, ws, hoja):
        for num_fila, fila in enumerate(_iter_filas(ws), start=3):
            try:
                cod      = _str(fila.get("codigo"))
                res_num  = _str(fila.get("numero"))
                res_fec  = _parse_fecha(fila.get("fecha"))
                res_tipo = _str(fila.get("tipo"))
                res_texto = _str(fila.get("texto"))

                if not cod:
                    self._registrar_skip(hoja, num_fila, "codigo vacío"); continue
                if not res_num:
                    self._registrar_skip(hoja, num_fila, f"{cod}: numero vacío"); continue
                if not res_fec:
                    self._registrar_skip(hoja, num_fila, f"{cod}: fecha vacía"); continue
                if not res_tipo:
                    self._registrar_skip(hoja, num_fila, f"{cod}: tipo vacío"); continue
                if not res_texto:
                    self._registrar_skip(hoja, num_fila, f"{cod}: texto vacío"); continue

                sim = SIM.objects.filter(codigo=cod.upper(), version=1).first()
                if not sim:
                    self._registrar_skip(hoja, num_fila, f"SIM no encontrado: {cod}"); continue

                res, creado = Resolucion.objects.get_or_create(
                    sim=sim,
                    instancia="PRIMERA",
                    numero=res_num.upper(),
                    defaults=dict(
                        fecha=res_fec,
                        tipo=res_tipo,
                        texto=res_texto.upper(),
                    ),
                )
                if creado:
                    t = _str(fila.get("tipo_notif"))
                    if t:
                        Notificacion.objects.get_or_create(
                            resolucion=res,
                            defaults=dict(
                                tipo=t,
                                notificado_a=_str(fila.get("notif_a")) or "",
                                fecha=_parse_fecha(fila.get("fecha_notif")),
                                hora=_parse_hora(fila.get("hora_notif")),
                            ),
                        )
                    self._registrar_ok(hoja)
                else:
                    self._registrar_skip(hoja, num_fila, f"{cod}/{res_num}: ya existe")
            except Exception as exc:
                self._registrar_error(hoja, num_fila, exc)

    # ── Hoja 5: RR ───────────────────────────────────────────────────

    def _importar_rr(self, ws, hoja):
        for num_fila, fila in enumerate(_iter_filas(ws), start=3):
            try:
                cod      = _str(fila.get("codigo"))
                res_num  = _str(fila.get("res_numero"))

                if not cod:
                    self._registrar_skip(hoja, num_fila, "codigo vacío"); continue
                if not res_num:
                    self._registrar_skip(hoja, num_fila, f"{cod}: res_numero vacío"); continue

                sim = SIM.objects.filter(codigo=cod.upper(), version=1).first()
                if not sim:
                    self._registrar_skip(hoja, num_fila, f"SIM no encontrado: {cod}"); continue

                try:
                    res = Resolucion.objects.get(
                        sim=sim, instancia="PRIMERA", numero=res_num.upper()
                    )
                except Resolucion.DoesNotExist:
                    self._registrar_skip(
                        hoja, num_fila, f"RES no encontrada: {res_num} en {cod}"
                    ); continue
                except Resolucion.MultipleObjectsReturned:
                    res = Resolucion.objects.filter(
                        sim=sim, instancia="PRIMERA", numero=res_num.upper()
                    ).first()

                if Resolucion.objects.filter(
                    sim=sim, instancia="RECONSIDERACION", resolucion_origen=res
                ).exists():
                    self._registrar_skip(hoja, num_fila, f"{cod}: RR ya existe para {res_num}")
                    continue

                numero_rr = _str(fila.get("numero"))
                if not numero_rr:
                    self._registrar_skip(hoja, num_fila, f"{cod}: numero del RR vacío"); continue

                tipo_raw = _str(fila.get("tipo"))
                tipo_val = tipo_raw[:100] if tipo_raw else None

                rr = Resolucion.objects.create(
                    sim=sim,
                    instancia="RECONSIDERACION",
                    resolucion_origen=res,
                    numero=numero_rr.upper(),
                    fecha_presentacion=_parse_fecha(fila.get("fecha_presentacion")),
                    fecha=_parse_fecha(fila.get("fecha")),
                    texto=(_str(fila.get("texto")) or "").upper() or None,
                    tipo=tipo_val,
                )
                t = _str(fila.get("tipo_notif"))
                if t:
                    Notificacion.objects.create(
                        resolucion=rr,
                        tipo=t,
                        notificado_a=_str(fila.get("notif_a")) or "",
                        fecha=_parse_fecha(fila.get("fecha_notif")),
                        hora=_parse_hora(fila.get("hora_notif")),
                    )
                self._registrar_ok(hoja)
            except Exception as exc:
                self._registrar_error(hoja, num_fila, exc)

    # ── Hoja 6: RAP ──────────────────────────────────────────────────

    def _importar_rap(self, ws, hoja):
        for num_fila, fila in enumerate(_iter_filas(ws), start=3):
            try:
                cod = _str(fila.get("codigo"))
                if not cod:
                    self._registrar_skip(hoja, num_fila, "codigo vacío"); continue

                sim = SIM.objects.filter(codigo=cod.upper(), version=1).first()
                if not sim:
                    self._registrar_skip(hoja, num_fila, f"SIM no encontrado: {cod}"); continue

                if RecursoTSP.objects.filter(sim=sim, instancia="APELACION").exists():
                    self._registrar_skip(hoja, num_fila, f"{cod}: RAP ya existe"); continue

                rap = RecursoTSP.objects.create(
                    sim=sim,
                    instancia="APELACION",
                    fecha_presentacion=_parse_fecha(fila.get("fecha_presentacion")),
                    numero_oficio=_str(fila.get("numero_oficio")),
                    fecha_oficio=_parse_fecha(fila.get("fecha_oficio")),
                    numero=_str(fila.get("numero")),
                    fecha=_parse_fecha(fila.get("fecha")),
                    tipo=_str(fila.get("tipo")),
                    texto=(_str(fila.get("texto")) or "").upper() or None,
                )
                t = _str(fila.get("tipo_notif"))
                if t:
                    Notificacion.objects.create(
                        recurso_tsp=rap,
                        tipo=t,
                        notificado_a=_str(fila.get("notif_a")) or "",
                        fecha=_parse_fecha(fila.get("fecha_notif")),
                        hora=_parse_hora(fila.get("hora_notif")),
                    )
                self._registrar_ok(hoja)
            except Exception as exc:
                self._registrar_error(hoja, num_fila, exc)

    # ── Hoja 7: AUTOTPE ──────────────────────────────────────────────

    def _importar_autotpe(self, ws, hoja):
        for num_fila, fila in enumerate(_iter_filas(ws), start=3):
            try:
                cod = _str(fila.get("codigo"))
                if not cod:
                    self._registrar_skip(hoja, num_fila, "codigo vacío"); continue

                sim = SIM.objects.filter(codigo=cod.upper(), version=1).first()
                if not sim:
                    self._registrar_skip(hoja, num_fila, f"SIM no encontrado: {cod}"); continue

                auto = AUTOTPE.objects.create(
                    sim=sim,
                    numero=_str(fila.get("numero")),
                    fecha=_parse_fecha(fila.get("fecha")),
                    tipo=_str(fila.get("tipo")),
                    texto=(_str(fila.get("texto")) or "").upper() or None,
                )
                t = _str(fila.get("tipo_notif"))
                if t:
                    Notificacion.objects.create(
                        autotpe=auto,
                        tipo=t,
                        notificado_a=_str(fila.get("notif_a")) or "",
                        fecha=_parse_fecha(fila.get("fecha_notif")),
                        hora=_parse_hora(fila.get("hora_notif")),
                    )
                memo_num = _str(fila.get("memo_numero"))
                if memo_num:
                    memo_fec = _parse_fecha(fila.get("memo_fecha"))
                    if memo_fec:
                        Memorandum.objects.create(
                            autotpe=auto,
                            numero=memo_num.upper(),
                            fecha=memo_fec,
                            fecha_entrega=_parse_fecha(fila.get("memo_fecha_entrega")),
                        )
                    else:
                        self.log.warning(
                            f"[{hoja} fila {num_fila}] memo_numero presente pero memo_fecha vacía — "
                            f"memorándum no creado para {cod}"
                        )
                self._registrar_ok(hoja)
            except Exception as exc:
                self._registrar_error(hoja, num_fila, exc)

    # ── Hoja 8: RAEE ─────────────────────────────────────────────────

    def _importar_raee(self, ws, hoja):
        for num_fila, fila in enumerate(_iter_filas(ws), start=3):
            try:
                cod = _str(fila.get("codigo"))
                if not cod:
                    self._registrar_skip(hoja, num_fila, "codigo vacío"); continue

                sim = SIM.objects.filter(codigo=cod.upper(), version=1).first()
                if not sim:
                    self._registrar_skip(hoja, num_fila, f"SIM no encontrado: {cod}"); continue

                raee = RecursoTSP.objects.create(
                    sim=sim,
                    instancia="ACLARACION_ENMIENDA",
                    numero=_str(fila.get("numero")),
                    fecha=_parse_fecha(fila.get("fecha")),
                    texto=(_str(fila.get("texto")) or "").upper() or None,
                )
                t = _str(fila.get("tipo_notif"))
                if t:
                    Notificacion.objects.create(
                        recurso_tsp=raee,
                        tipo=t,
                        notificado_a=_str(fila.get("notif_a")) or "",
                        fecha=_parse_fecha(fila.get("fecha_notif")),
                        hora=_parse_hora(fila.get("hora_notif")),
                    )
                self._registrar_ok(hoja)
            except Exception as exc:
                self._registrar_error(hoja, num_fila, exc)
