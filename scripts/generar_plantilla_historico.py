"""
Genera la plantilla Excel para carga masiva del histórico de sumarios (SIM).
Versión 4.1 — compatible con estructura snake_case v4.0 y modelos actuales.

Ejecutar desde la raíz del proyecto:
    python generar_plantilla_historico.py

Produce: plantilla_historico_sim.xlsx

IMPORTANTE: Los nombres de columna en cada hoja son exactamente los que lee
el comando `python manage.py importar_historico`. No cambiarlos.
"""

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# ─── Paleta de colores ───────────────────────────────────────────────────────
COLOR_REQUERIDO = "C00000"   # Rojo oscuro  — campo obligatorio
COLOR_OPCIONAL  = "203864"   # Azul oscuro  — campo opcional
COLOR_FK        = "833C00"   # Marrón       — clave foránea (referencia a otra hoja)
COLOR_HEADER_BG = "1F3864"
COLOR_EJEMPLO_BG = "E2EFDA"

FILL_REQUERIDO = PatternFill("solid", fgColor=COLOR_REQUERIDO)
FILL_OPCIONAL  = PatternFill("solid", fgColor=COLOR_OPCIONAL)
FILL_FK        = PatternFill("solid", fgColor=COLOR_FK)
FILL_HEADER    = PatternFill("solid", fgColor=COLOR_HEADER_BG)
FILL_EJEMPLO   = PatternFill("solid", fgColor=COLOR_EJEMPLO_BG)

FONT_HEADER  = Font(bold=True, color="FFFFFF", name="Calibri", size=12)
FONT_LABEL   = Font(bold=True, color="FFFFFF", name="Calibri", size=9)
FONT_EJEMPLO = Font(italic=True, name="Calibri", size=9)
FONT_NORMAL  = Font(name="Calibri", size=9)

THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin"),
)

ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)


# ─── Choices exactos según models.py v4.1 ────────────────────────────────────

# Listas cortas → fórmula inline en DataValidation (< 255 chars)
ESTADO_SIM = [
    "PARA_AGENDA", "PROCESO_EN_EL_TPE", "PROCESO_EN_EL_TSP",
    "PROCESO_CONCLUIDO_TPE", "PROCESO_EJECUTADO", "OBSERVADO",
]
ESCALAFON_PM = [
    "GENERAL", "OFICIAL SUPERIOR", "OFICIAL SUBALTERNO",
    "SUBOFICIAL", "SARGENTO", "TROPA", "EMPLEADO CIVIL",
]
ARMA_PM = [
    "INF.", "CAB.", "ART.", "ING.", "COM.", "LOG.",
    "M.B.", "INT.", "SAN.", "TGRAFO.", "AV.", "MÚS.",
]
ESTADO_PM = [
    "ACTIVO", "SERVICIO ACTIVO", "LETRA A", "SERVICIO PASIVO",
    "RETIRO OBLIGATORIO", "RESERVA ACTIVA", "BAJA", "FALLECIDO", "NO HABIDO", "OTRO",
]
NOTIF_CHOICES = ["FIRMA", "EDICTO", "CEDULON"]
TIPO_AUTOTPE  = [
    "SOBRESEIDO", "NULIDAD_OBRADOS", "SANCION_ARRESTO", "SANCION_LETRA_B",
    "SANCION_RETIRO_OBLIGATORIO", "AUTO_CUMPLIMIENTO", "AUTO_EJECUTORIA",
    "AUTO_EXCUSA", "AUTO_RECHAZO_RECURSO",
]
TIPO_RAP = [
    "REVOCAR", "CONFIRMAR", "MODIFICAR",
    "ANULAR HASTA EL VICIO MAS ANTIGUO", "OTRO",
]
RR_RESUM_CHOICES = ["PROCEDENCIA", "IMPROCEDENCIA"]

# Listas largas → en hoja _LISTAS (columnas A, B, C) para evitar límite 255 chars
GRADO_PM = [
    "GRAL. EJTO.", "GRAL. DIV.", "GRAL. BRIG.",
    "CNL.", "TCNL.", "MY.",
    "CAP.", "TTE.", "SBTTE.",
    "SOF. MTRE.", "SOF. MY.", "SOF. 1RO.", "SOF. 2DO.", "SOF. INCL.",
    "SGTO. 1RO.", "SGTO. 2DO.", "SGTO. INCL.",
    "CABO", "DGTE.", "SLDO.",
    "PROF. V", "PROF. IV", "PROF. III", "PROF. II", "PROF. I",
    "TEC. V", "TEC. IV", "TEC. III", "TEC. II", "TEC. I",
    "ADM. V", "ADM. IV", "ADM. III", "ADM. II", "ADM. I",
    "APAD. V", "APAD. IV", "APAD. III", "APAD. II", "APAD. I",
]
TIPO_SIM = [
    "DISCIPLINARIO", "ADMINISTRATIVO", "ASCENSO POSTUMO",
    "SOLICITUD DE RETIRO VOLUNTARIO",
    "SOLICITUD_LETRA_D", "SOLICITUD_LICENCIA_MAXIMA",
    "SOLICITUD_RESTITUCION_ANTIGUEDAD",
    "SOLICITUD_DE_RESTITUCION_DE_DERECHOS_PROFESIONALES",
    "SOLICITUD_ASCENSO_AL_GRADO_INMEDIATO_SUPERIOR",
    "SOLICITUD_ART_114_(Invalidez Instructor)",
    "SOLICITUD_ART_117_(Fallecimiento)",
    "SOLICITUD_ART_118_(Invalidez Sldo)",
]
TIPO_RES = [
    "ARCHIVO_OBRADOS", "ADMINISTRATIVO", "SANCIONES_DISCIPLINARIAS",
    "NO_HA_LUGAR_SANCION_DISCIPLINARIA", "SOLICITUD_DE_RETIRO_VOLUNTARIO",
    "SANCION_ARRESTO", "SANCION_LETRA_B", "SANCION_RETIRO_OBLIGATORIO", "SANCION_BAJA",
    "SOLICITUD_LETRA_D", "SOLICITUD_LICENCIA_MAXIMA", "SOLICITUD_ASCENSO",
    "SOLICITUD_RESTITUCION_ANTIGUEDAD",
    "SOLICITUD_RESTITUCION_DE_DERECHOS_PROFESIONALES",
    "SOLICITUD_ART_114_(Invalidez Instructor)",
    "SOLICITUD_ART_117_(Fallecimiento)",
    "SOLICITUD_ART_118_(Invalidez Sldo)",
    "OTRO",
]

# Referencias a _LISTAS (usadas como formula1 en DataValidation)
REF_GRADO   = f"'_LISTAS'!$A$2:$A${1 + len(GRADO_PM)}"
REF_TIPO_SIM = f"'_LISTAS'!$B$2:$B${1 + len(TIPO_SIM)}"
REF_TIPO_RES = f"'_LISTAS'!$C$2:$C${1 + len(TIPO_RES)}"


# ─── Utilidades ──────────────────────────────────────────────────────────────

def _header_cell(ws, row, col, text, fill):
    c = ws.cell(row=row, column=col, value=text)
    c.fill = fill
    c.font = FONT_LABEL
    c.alignment = ALIGN_CENTER
    c.border = THIN_BORDER
    return c


def _data_cell(ws, row, col, value=""):
    c = ws.cell(row=row, column=col, value=value)
    c.font = FONT_EJEMPLO
    c.fill = FILL_EJEMPLO
    c.alignment = ALIGN_LEFT
    c.border = THIN_BORDER
    return c


def _add_validation(ws, choices_or_formula, col_letter, start_row=3, end_row=5002):
    """
    Agrega validación desplegable a una columna.
    choices_or_formula puede ser:
      - lista de strings  → fórmula inline entre comillas
      - string            → referencia a rango (ej: "'_LISTAS'!$A$2:$A$41")
    """
    if isinstance(choices_or_formula, str):
        formula = choices_or_formula  # ya es una fórmula de rango
    else:
        formula = '"' + ",".join(choices_or_formula) + '"'

    dv = DataValidation(
        type="list",
        formula1=formula,
        allow_blank=True,
        showErrorMessage=True,
        errorTitle="Valor inválido",
        error="Seleccione un valor de la lista desplegable.",
    )
    ws.add_data_validation(dv)
    dv.sqref = f"{col_letter}{start_row}:{col_letter}{end_row}"


def _build_sheet(ws, columns, example_row, title=None):
    """
    columns = [(header_text, fill, choices_or_formula_or_None, width), ...]
    header_text PRIMERA LÍNEA = clave que leerá el importador.
    example_row = [val1, val2, ...]
    """
    ROW_TITLE   = 1
    ROW_HEADERS = 2
    ROW_EJEMPLO = 3

    if title:
        ws.merge_cells(start_row=1, start_column=1,
                       end_row=1,   end_column=len(columns))
        c = ws.cell(row=1, column=1, value=title)
        c.fill = FILL_HEADER
        c.font = FONT_HEADER
        c.alignment = ALIGN_CENTER
        ws.row_dimensions[1].height = 22

    for idx, (header, fill, _, _w) in enumerate(columns, start=1):
        _header_cell(ws, ROW_HEADERS, idx, header, fill)
    ws.row_dimensions[2].height = 40

    for idx, val in enumerate(example_row, start=1):
        _data_cell(ws, ROW_EJEMPLO, idx, val)
    ws.row_dimensions[3].height = 18

    for idx, (_, _, choices, width) in enumerate(columns, start=1):
        col_letter = get_column_letter(idx)
        ws.column_dimensions[col_letter].width = width
        if choices:
            _add_validation(ws, choices, col_letter)

    ws.freeze_panes = "A3"
    last_col = get_column_letter(len(columns))
    ws.auto_filter.ref = f"A2:{last_col}2"


# ─── MAIN ────────────────────────────────────────────────────────────────────

def generar_plantilla():
    wb = Workbook()

    # ──────────────────────────────────────────────────────────────────────────
    # HOJA _LISTAS (oculta) — listas largas para DataValidation
    # Columna A: GRADO_PM  |  B: TIPO_SIM  |  C: TIPO_RES
    # ──────────────────────────────────────────────────────────────────────────
    ws_listas = wb.active
    ws_listas.title = "_LISTAS"

    encabezados_listas = ["GRADO", "TIPO_SIM", "TIPO_RES"]
    for col, enc in enumerate(encabezados_listas, 1):
        c = ws_listas.cell(row=1, column=col, value=enc)
        c.font = Font(bold=True, name="Calibri", size=9)
        c.fill = FILL_HEADER
        c.font = FONT_LABEL

    for i, val in enumerate(GRADO_PM, 2):
        ws_listas.cell(row=i, column=1, value=val)
    for i, val in enumerate(TIPO_SIM, 2):
        ws_listas.cell(row=i, column=2, value=val)
    for i, val in enumerate(TIPO_RES, 2):
        ws_listas.cell(row=i, column=3, value=val)

    ws_listas.sheet_state = "hidden"

    # ──────────────────────────────────────────────────────────────────────────
    # HOJA INSTRUCCIONES
    # ──────────────────────────────────────────────────────────────────────────
    ws0 = wb.create_sheet("INSTRUCCIONES")

    instrucciones = [
        ("PLANTILLA DE CARGA MASIVA — HISTÓRICO SIM  v4.1", COLOR_REQUERIDO),
        ("", ""),
        ("LEYENDA DE COLORES EN ENCABEZADOS:", COLOR_HEADER_BG),
        ("  Rojo   → Campo OBLIGATORIO (no puede quedar vacío)", COLOR_REQUERIDO),
        ("  Azul   → Campo OPCIONAL (puede quedar vacío si no hay dato)", COLOR_OPCIONAL),
        ("  Marrón → Clave FORÁNEA — debe coincidir con un valor de otra hoja", COLOR_FK),
        ("", ""),
        ("ORDEN DE CARGA (seguir este orden exacto):", COLOR_HEADER_BG),
        ("  1. Hoja 1_SIM      → Sumarios (tabla central)", ""),
        ("  2. Hoja 2_PM       → Personal Militar", ""),
        ("  3. Hoja 3_PM_SIM   → Relación militar ↔ sumario", ""),
        ("  4. Hoja 4_RES      → Primera Resolución del TPE", ""),
        ("  5. Hoja 5_RR       → Recurso de Reconsideración", ""),
        ("  6. Hoja 6_RAP      → Recurso de Apelación al TSP", ""),
        ("  7. Hoja 7_AUTOTPE  → Autos del Tribunal TPE", ""),
        ("  8. Hoja 8_RAEE     → Aclaración, Explicación y Enmienda", ""),
        ("", ""),
        ("REGLAS GENERALES:", COLOR_HEADER_BG),
        ("  • Fechas en formato: AAAA-MM-DD  (ej: 2010-03-15)", ""),
        ("  • Horas en formato: HH:MM         (ej: 09:30)", ""),
        ("  • Año de promoción: solo el número del año (ej: 1998)", ""),
        ("  • 'codigo' es el código único del sumario (ej: DJE-095/25)", ""),
        ("  • 'ci' es la cédula de identidad del militar (solo números)", ""),
        ("  • La fila 3 de cada hoja es EJEMPLO — borrarla antes de importar", ""),
        ("  • Dropdowns disponibles: usar EXACTAMENTE el valor de la lista", ""),
        ("  • Si un caso no tiene RR, RAP o AUTOTPE: simplemente no llenar esa hoja", ""),
        ("", ""),
        ("COMANDO DE IMPORTACIÓN:", COLOR_HEADER_BG),
        ("  python manage.py importar_historico plantilla_historico_sim.xlsx", ""),
        ("  python manage.py importar_historico plantilla_historico_sim.xlsx --dry-run", ""),
        ("  python manage.py importar_historico plantilla_historico_sim.xlsx --desde-hoja 4_RES", ""),
        ("", ""),
        ("NOTA SOBRE NOMBRES DE COLUMNA:", COLOR_HEADER_BG),
        ("  Los encabezados usan nombres exactos de campos del modelo Django (snake_case).", ""),
        ("  NO MODIFICAR los encabezados — el importador los lee tal cual.", ""),
    ]

    ws0.column_dimensions["A"].width = 90
    for i, (texto, color) in enumerate(instrucciones, start=1):
        c = ws0.cell(row=i, column=1, value=texto)
        c.font = Font(
            bold=(color != ""),
            color=(color if color else "000000"),
            name="Calibri",
            size=10,
        )
        c.alignment = ALIGN_LEFT
        ws0.row_dimensions[i].height = 18


    # ──────────────────────────────────────────────────────────────────────────
    # HOJA 1: SIM
    # Importador lee: codigo, fecha_ingreso, estado, tipo, objeto, resumen,
    #                 auto_final, numero_carpeta
    # ──────────────────────────────────────────────────────────────────────────
    ws1 = wb.create_sheet("1_SIM")
    cols_sim = [
        # (primera línea = clave del importador, fill, choices, width)
        ("codigo\n(Código único SIM\nej: DJE-095/25)",         FILL_REQUERIDO, None,        22),
        ("fecha_ingreso\n(Fecha ingreso TPE\nAAAA-MM-DD)",     FILL_OPCIONAL,  None,        16),
        ("estado\n(Estado del sumario)",                        FILL_REQUERIDO, ESTADO_SIM,  24),
        ("tipo\n(Tipo de sumario)",                             FILL_REQUERIDO, REF_TIPO_SIM, 32),
        ("objeto\n(Objeto completo del sumario)",               FILL_REQUERIDO, None,        50),
        ("resumen\n(Resumen breve, máx 200 car.)",              FILL_REQUERIDO, None,        32),
        ("auto_final\n(Auto Final/Dictamen — opcional)",        FILL_OPCIONAL,  None,        35),
        ("numero_carpeta\n(N° carpeta física\nsolo número)",    FILL_OPCIONAL,  None,        16),
    ]
    ej_sim = [
        "DJE-095/25",
        "2005-03-15",
        "PROCESO_CONCLUIDO_TPE",
        "DISCIPLINARIO",
        "INVESTIGACION POR ABANDONO DE GUARDIA EN LA UNIDAD MILITAR N 4 DE INFANTERIA",
        "ABANDONO DE GUARDIA",
        "ARCHIVADO POR PRESCRIPCION",
        "132",
    ]
    _build_sheet(ws1, cols_sim, ej_sim,
                 title="HOJA 1 — SUMARIO INFORMATIVO MILITAR (SIM)  |  Una fila por sumario")


    # ──────────────────────────────────────────────────────────────────────────
    # HOJA 2: PM — Personal Militar
    # Importador lee: ci, escalafon, grado, arma, especialidad, nombre,
    #                 paterno, materno, estado, anio_promocion
    # ──────────────────────────────────────────────────────────────────────────
    ws2 = wb.create_sheet("2_PM")
    cols_pm = [
        ("ci\n(Cédula Identidad — único\nsolo números)",        FILL_REQUERIDO, None,        16),
        ("escalafon\n(Escalafón)",                              FILL_OPCIONAL,  ESCALAFON_PM, 22),
        ("grado\n(Grado militar actual)",                       FILL_OPCIONAL,  REF_GRADO,   20),
        ("arma\n(Arma o cuerpo)",                               FILL_OPCIONAL,  ARMA_PM,     14),
        ("especialidad\n(Especialidad — opcional)",             FILL_OPCIONAL,  None,        18),
        ("nombre\n(Nombre/s del militar)",                      FILL_REQUERIDO, None,        22),
        ("paterno\n(Apellido Paterno)",                         FILL_REQUERIDO, None,        22),
        ("materno\n(Apellido Materno — opcional)",              FILL_OPCIONAL,  None,        22),
        ("estado\n(Estado del militar)",                        FILL_OPCIONAL,  ESTADO_PM,   22),
        ("anio_promocion\n(Año de egreso/promoción\nsolo año ej: 1998)", FILL_OPCIONAL, None, 18),
    ]
    ej_pm = [
        "4521890",
        "OFICIAL SUPERIOR",
        "CNL.",
        "INF.",
        "",
        "JUAN CARLOS",
        "MAMANI",
        "QUISPE",
        "RETIRO OBLIGATORIO",
        "1998",
    ]
    _build_sheet(ws2, cols_pm, ej_pm,
                 title="HOJA 2 — PERSONAL MILITAR (PM)  |  Una fila por militar (no duplicar cédula)")


    # ──────────────────────────────────────────────────────────────────────────
    # HOJA 3: PM_SIM — Relación Militar ↔ Sumario
    # Importador lee: codigo, ci, grado_en_fecha
    # ──────────────────────────────────────────────────────────────────────────
    ws3 = wb.create_sheet("3_PM_SIM")
    cols_pm_sim = [
        ("codigo\n(FK → hoja 1_SIM  Código SIM)",        FILL_FK,       None,      22),
        ("ci\n(FK → hoja 2_PM  Cédula del militar)",     FILL_FK,       None,      16),
        ("grado_en_fecha\n(Grado al momento del sumario\nsi difiere del grado actual)", FILL_OPCIONAL, REF_GRADO, 22),
    ]
    ej_pm_sim = ["DJE-095/25", "4521890", "TCNL."]
    _build_sheet(ws3, cols_pm_sim, ej_pm_sim,
                 title="HOJA 3 — PM_SIM  |  Un militar puede aparecer en varios sumarios")


    # ──────────────────────────────────────────────────────────────────────────
    # HOJA 4: RES — Primera Resolución del TPE
    # Importador lee: codigo, numero, fecha, tipo, texto,
    #                 tipo_notif, notif_a, fecha_notif, hora_notif
    # ──────────────────────────────────────────────────────────────────────────
    ws4 = wb.create_sheet("4_RES")
    cols_res = [
        ("codigo\n(FK → hoja 1_SIM)",                    FILL_FK,        None,          22),
        ("numero\n(Número de resolución\nej: 045/26)",    FILL_REQUERIDO, None,          16),
        ("fecha\n(Fecha resolución\nAAAA-MM-DD)",         FILL_REQUERIDO, None,          14),
        ("tipo\n(Tipo de resolución)",                    FILL_REQUERIDO, REF_TIPO_RES,  32),
        ("texto\n(Texto completo de la resolución)",      FILL_REQUERIDO, None,          50),
        ("tipo_notif\n(Tipo notificación)",               FILL_OPCIONAL,  NOTIF_CHOICES, 16),
        ("notif_a\n(Notificado a / Dirección / Periódico)", FILL_OPCIONAL, None,         30),
        ("fecha_notif\n(Fecha notificación\nAAAA-MM-DD)", FILL_OPCIONAL,  None,          14),
        ("hora_notif\n(Hora notificación\nHH:MM)",        FILL_OPCIONAL,  None,          12),
    ]
    ej_res = [
        "DJE-095/25",
        "045/26",
        "2006-08-10",
        "SANCIONES_DISCIPLINARIAS",
        "SE IMPONE SANCION DE ARRESTO DE 15 DIAS AL TCnel. MAMANI POR ABANDONO DE GUARDIA",
        "FIRMA",
        "MAMANI QUISPE JUAN CARLOS",
        "2006-08-15",
        "09:30",
    ]
    _build_sheet(ws4, cols_res, ej_res,
                 title="HOJA 4 — PRIMERA RESOLUCIÓN (RES)  |  Un sumario tiene máximo una RES")


    # ──────────────────────────────────────────────────────────────────────────
    # HOJA 5: RR — Recurso de Reconsideración
    # Importador lee: codigo, res_numero (FK a RES), fecha_presentacion,
    #                 numero, fecha, texto, resumen,
    #                 tipo_notif, notif_a, fecha_notif, hora_notif
    # ──────────────────────────────────────────────────────────────────────────
    ws5 = wb.create_sheet("5_RR")
    cols_rr = [
        ("codigo\n(FK → hoja 1_SIM)",                         FILL_FK,        None,          22),
        ("res_numero\n(FK → hoja 4_RES\nN° de la 1ra resolución)", FILL_FK,   None,          18),
        ("fecha_presentacion\n(Fecha presentación RR\nAAAA-MM-DD)", FILL_OPCIONAL, None,     14),
        ("numero\n(Número del RR\nej: 012/26)",                FILL_OPCIONAL,  None,          16),
        ("fecha\n(Fecha resolución RR\nAAAA-MM-DD)",           FILL_OPCIONAL,  None,          14),
        ("texto\n(Texto resolución RR)",                       FILL_OPCIONAL,  None,          50),
        ("resumen\n(PROCEDENCIA o IMPROCEDENCIA)",             FILL_OPCIONAL,  RR_RESUM_CHOICES, 22),
        ("tipo_notif\n(Tipo notificación)",                    FILL_OPCIONAL,  NOTIF_CHOICES, 16),
        ("notif_a\n(Notificado a)",                            FILL_OPCIONAL,  None,          28),
        ("fecha_notif\n(Fecha notificación\nAAAA-MM-DD)",      FILL_OPCIONAL,  None,          14),
        ("hora_notif\n(Hora notificación\nHH:MM)",             FILL_OPCIONAL,  None,          12),
    ]
    ej_rr = [
        "DJE-095/25",
        "045/26",
        "2006-08-25",
        "012/26",
        "2006-09-20",
        "SE CONFIRMA LA SANCION DE ARRESTO IMPUESTA POR LA PRIMERA RESOLUCION",
        "PROCEDENCIA",
        "EDICTO",
        "El Diario — pág. 12",
        "2006-09-25",
        "10:00",
    ]
    _build_sheet(ws5, cols_rr, ej_rr,
                 title="HOJA 5 — RECURSO DE RECONSIDERACIÓN (RR)  |  Solo si el sumario tiene RR")


    # ──────────────────────────────────────────────────────────────────────────
    # HOJA 6: RAP — Recurso de Apelación al TSP
    # Importador lee: codigo, fecha_presentacion, numero_oficio, fecha_oficio,
    #                 numero, fecha, tipo, texto,
    #                 tipo_notif, notif_a, fecha_notif, hora_notif
    # ──────────────────────────────────────────────────────────────────────────
    ws6 = wb.create_sheet("6_RAP")
    cols_rap = [
        ("codigo\n(FK → hoja 1_SIM)",                           FILL_FK,        None,          22),
        ("fecha_presentacion\n(Fecha presentación RAP\nAAAA-MM-DD)", FILL_OPCIONAL, None,      14),
        ("numero_oficio\n(N° Oficio de Elevación al TSP)",      FILL_OPCIONAL,  None,          18),
        ("fecha_oficio\n(Fecha oficio elevación\nAAAA-MM-DD)",  FILL_OPCIONAL,  None,          14),
        ("numero\n(N° Resolución TSP\nej: 089/27)",             FILL_OPCIONAL,  None,          18),
        ("fecha\n(Fecha resolución TSP\nAAAA-MM-DD)",           FILL_OPCIONAL,  None,          14),
        ("tipo\n(Tipo resolución TSP)",                         FILL_OPCIONAL,  TIPO_RAP,      32),
        ("texto\n(Texto resolución TSP)",                       FILL_OPCIONAL,  None,          50),
        ("tipo_notif\n(Tipo notificación)",                     FILL_OPCIONAL,  NOTIF_CHOICES, 16),
        ("notif_a\n(Notificado a)",                             FILL_OPCIONAL,  None,          28),
        ("fecha_notif\n(Fecha notificación\nAAAA-MM-DD)",       FILL_OPCIONAL,  None,          14),
        ("hora_notif\n(Hora notificación\nHH:MM)",              FILL_OPCIONAL,  None,          12),
    ]
    ej_rap = [
        "DJE-095/25",
        "2006-10-05",
        "OFI-231/26",
        "2006-10-08",
        "089/27",
        "2007-03-15",
        "CONFIRMAR",
        "EL TSP CONFIRMA LA SEGUNDA RESOLUCION EN TODOS SUS TERMINOS",
        "FIRMA",
        "MAMANI QUISPE JUAN CARLOS",
        "2007-03-20",
        "08:45",
    ]
    _build_sheet(ws6, cols_rap, ej_rap,
                 title="HOJA 6 — RECURSO DE APELACIÓN TSP (RAP)  |  Solo si el sumario llegó al TSP")


    # ──────────────────────────────────────────────────────────────────────────
    # HOJA 7: AUTOTPE — Autos del Tribunal
    # Importador lee: codigo, numero, fecha, tipo, texto,
    #                 tipo_notif, notif_a, fecha_notif, hora_notif,
    #                 memo_numero, memo_fecha, memo_fecha_entrega
    # ──────────────────────────────────────────────────────────────────────────
    ws7 = wb.create_sheet("7_AUTOTPE")
    cols_tpe = [
        ("codigo\n(FK → hoja 1_SIM)",                          FILL_FK,        None,          22),
        ("numero\n(Número de Auto\nej: 015/27)",                FILL_OPCIONAL,  None,          16),
        ("fecha\n(Fecha del Auto\nAAAA-MM-DD)",                 FILL_OPCIONAL,  None,          14),
        ("tipo\n(Tipo de Auto)",                                FILL_OPCIONAL,  TIPO_AUTOTPE,  32),
        ("texto\n(Texto del Auto)",                             FILL_OPCIONAL,  None,          50),
        ("tipo_notif\n(Tipo notificación)",                     FILL_OPCIONAL,  NOTIF_CHOICES, 16),
        ("notif_a\n(Notificado a)",                             FILL_OPCIONAL,  None,          28),
        ("fecha_notif\n(Fecha notificación\nAAAA-MM-DD)",       FILL_OPCIONAL,  None,          14),
        ("hora_notif\n(Hora notificación\nHH:MM)",              FILL_OPCIONAL,  None,          12),
        ("memo_numero\n(N° Memorándum\nsolo si AUTO_EJECUTORIA)", FILL_OPCIONAL, None,         18),
        ("memo_fecha\n(Fecha Memorándum\nAAAA-MM-DD)",          FILL_OPCIONAL,  None,          14),
        ("memo_fecha_entrega\n(Fecha entrega Memo\nAAAA-MM-DD)", FILL_OPCIONAL, None,          14),
    ]
    ej_tpe = [
        "DJE-095/25",
        "015/27",
        "2007-04-10",
        "AUTO_EJECUTORIA",
        "SE DECLARA EJECUTORIADA LA RESOLUCION Y SE ORDENA SU CUMPLIMIENTO INMEDIATO",
        "FIRMA",
        "Unidad de Personal — Cmd. Ejército",
        "2007-04-12",
        "11:00",
        "MEMO-089/27",
        "2007-04-12",
        "2007-04-15",
    ]
    _build_sheet(ws7, cols_tpe, ej_tpe,
                 title="HOJA 7 — AUTOS TPE  |  Un sumario puede tener varios Autos")


    # ──────────────────────────────────────────────────────────────────────────
    # HOJA 8: RAEE — Aclaración, Explicación y Enmienda
    # Importador lee: codigo, numero, fecha, texto,
    #                 tipo_notif, notif_a, fecha_notif, hora_notif
    # ──────────────────────────────────────────────────────────────────────────
    ws8 = wb.create_sheet("8_RAEE")
    cols_raee = [
        ("codigo\n(FK → hoja 1_SIM)",                          FILL_FK,        None,          22),
        ("numero\n(Número resolución RAEE\nej: RAEE-003/27)",   FILL_OPCIONAL,  None,          18),
        ("fecha\n(Fecha resolución\nAAAA-MM-DD)",               FILL_OPCIONAL,  None,          14),
        ("texto\n(Texto de la resolución RAEE)",                FILL_OPCIONAL,  None,          50),
        ("tipo_notif\n(Tipo notificación)",                     FILL_OPCIONAL,  NOTIF_CHOICES, 16),
        ("notif_a\n(Notificado a)",                             FILL_OPCIONAL,  None,          28),
        ("fecha_notif\n(Fecha notificación\nAAAA-MM-DD)",       FILL_OPCIONAL,  None,          14),
        ("hora_notif\n(Hora notificación\nHH:MM)",              FILL_OPCIONAL,  None,          12),
    ]
    ej_raee = [
        "DJE-095/25",
        "RAEE-003/27",
        "2007-05-02",
        "SE ACLARA EL PUNTO 3 DE LA RESOLUCION TSP EN EL SENTIDO DE QUE EL PLAZO ES DE 30 DIAS",
        "CEDULON",
        "MAMANI QUISPE JUAN CARLOS — DOMICILIO CONOCIDO",
        "2007-05-06",
        "09:00",
    ]
    _build_sheet(ws8, cols_raee, ej_raee,
                 title="HOJA 8 — RAEE  |  Solo si el sumario tuvo recurso de aclaración")


    # ──────────────────────────────────────────────────────────────────────────
    # Guardar
    # ──────────────────────────────────────────────────────────────────────────
    nombre = "plantilla_historico_sim.xlsx"
    wb.save(nombre)
    print(f"\nPlantilla generada: {nombre}")
    print("  Hojas visibles:")
    for ws in wb.worksheets:
        if ws.sheet_state != "hidden":
            print(f"    • {ws.title}")
    print(f"\nHoja oculta '_LISTAS': listas de referencia para dropdowns")
    print("\nLeyenda de columnas:")
    print("  ROJO   = obligatorio")
    print("  AZUL   = opcional")
    print("  MARRÓN = clave foránea (referenciar otra hoja)")
    print("\nRecuerda: la fila 3 de cada hoja es un EJEMPLO — bórrala antes de importar.\n")


if __name__ == "__main__":
    generar_plantilla()
